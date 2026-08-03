"""
Plantillas Excel — que se generen, que se vean como deben y que el importador
acepte lo que la propia plantilla produce (ida y vuelta).

El chequeo de ida y vuelta es el que de verdad protege: hasta ahora nada
garantizaba que la plantilla que se descarga fuese la que su importador espera.
"""
from datetime import date
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient

import main
import plantillas
from core import auth, config
from parsers.plantilla_pu import parsear_plantilla_pu
from plantillas._estilo import NIVELES

DATOS = {
    "fases": [("11", "Obras civiles"), ("40", "Estructuras")],
    "cargos": ["OPERARIO", "PEON"],
    "unidades": ["m3", "kg"],
    "areas": ["PLANTA", "MINA"],
}


def _libro(clave: str):
    wb, archivo = plantillas.generar(clave, DATOS, "OBRA DE PRUEBA", date(2026, 8, 1))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf), archivo


def _claves():
    return [k for k in plantillas.CATALOGO if k != "pu"]


# ── Estructura ───────────────────────────────────────────────
@pytest.mark.parametrize("clave", list(plantillas.CATALOGO))
def test_todas_se_generan_y_abren(clave):
    wb, archivo = _libro(clave)
    assert archivo.endswith(".xlsx")
    assert len(wb.sheetnames) >= 2      # datos + instrucciones (+ catálogos)


@pytest.mark.parametrize("clave", _claves())
def test_cabeceras_exactas_y_en_orden(clave):
    """La cabecera es el CONTRATO con el importador: si cambia, los archivos que
    ya circulan dejan de importar."""
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]
    fila = next(r for r in range(1, 40)
                if ws.cell(r, 2).value == p.cols[0].clave)
    leidas = [ws.cell(fila, 2 + i).value for i in range(len(p.cols))]
    assert leidas == [c.clave for c in p.cols]


@pytest.mark.parametrize("clave", _claves())
def test_lo_esencial_del_formato(clave):
    """Sin cuadrícula, con paneles congelados y con la cabecera coloreada: es lo
    que separa un documento de una hoja de cálculo en bruto."""
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]
    assert ws.sheet_view.showGridLines is False
    assert ws.freeze_panes, "sin congelar, la cabecera se pierde al bajar"
    assert ws.auto_filter.ref
    fila = next(r for r in range(1, 40) if ws.cell(r, 2).value == p.cols[0].clave)
    for i, col in enumerate(p.cols):
        cel = ws.cell(fila, 2 + i)
        assert cel.fill.fgColor.rgb[-6:] == NIVELES[col.nivel]["bg"], col.clave
        assert cel.comment is not None, f"{col.clave} sin ayuda emergente"


@pytest.mark.parametrize("clave", _claves())
def test_las_obligatorias_se_avisan_en_rojo(clave):
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]
    n_oblig = sum(1 for c in p.cols if c.nivel == "obligatorio")
    assert len(ws.conditional_formatting._cf_rules) == n_oblig


@pytest.mark.parametrize("clave", _claves())
def test_los_valores_cerrados_son_desplegables(clave):
    """Un valor cerrado tecleado a mano solo falla DESPUÉS de subir el archivo."""
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]
    con_lista = [c for c in p.cols if c.lista or c.catalogo]
    assert len(ws.data_validations.dataValidation) == len(con_lista)


def test_el_paso_1_apunta_a_la_fila_real_del_ejemplo():
    """El paso 1 dice «borra la fila N». Si N no es la del ejemplo, la
    instrucción manda a borrar datos buenos."""
    for clave in _claves():
        wb, _ = _libro(clave)
        p = plantillas.definir(clave, DATOS)
        ws = wb[p.hoja]
        texto = next(ws.cell(r, 2).value for r in range(1, 40)
                     if str(ws.cell(r, 2).value or "").startswith("1."))
        n = int("".join(ch for ch in texto.split("(la ")[1].split(",")[0] if ch.isdigit()))
        assert ws.cell(n, 2).font.italic, f"{clave}: la fila {n} no es la del ejemplo"


def test_las_fechas_de_ejemplo_son_fechas_no_texto():
    """Si el ejemplo fuera texto, la columna acabaría con dos tipos mezclados
    según cómo teclee cada persona — el defecto que esta tanda cierra."""
    for clave in _claves():
        p = plantillas.definir(clave, DATOS)
        for col in p.cols:
            if "yyyy" in col.formato and col.ejemplo != "":
                assert isinstance(col.ejemplo, date), f"{clave}.{col.clave}"


def test_las_formulas_apuntan_a_columnas_existentes():
    """Las fórmulas nombran columnas (`{HITO1_PESO}`), no letras: una letra a
    mano ya apuntó a la columna equivocada sin que nada fallara."""
    for clave in _claves():
        p = plantillas.definir(clave, DATOS)
        claves = {c.clave for c in p.cols}
        for col in p.cols:
            if not col.formula:
                continue
            refs = {t.split("}")[0] for t in col.formula.split("{")[1:]}
            assert refs - {"f"} <= claves, f"{clave}.{col.clave}: {refs - claves}"


def test_el_padre_del_ejemplo_va_sin_fase_y_sin_hitos():
    """Las dos filas de ejemplo de PARTIDAS enseñan la única regla que no se
    puede adivinar: fase vacía = la fila solo agrupa."""
    wb, _ = _libro("partidas")
    ws = wb["PARTIDAS"]
    fila = next(r for r in range(1, 40) if ws.cell(r, 2).value == "CODIGO") + 1
    p = plantillas.definir("partidas", DATOS)
    col = {c.clave: 2 + i for i, c in enumerate(p.cols)}
    assert ws.cell(fila, col["FASE"]).value in (None, "")
    assert ws.cell(fila, col["HITO1_DESC"]).value in (None, "")
    assert ws.cell(fila + 1, col["FASE"]).value, "la 2ª fila debe ser una partida real"
    assert ws.cell(fila + 1, col["HITO1_PESO"]).value == 0.10


def test_los_catalogos_del_proyecto_llegan_a_la_plantilla():
    wb, _ = _libro("costos")
    valores = [c.value for row in wb["CATALOGOS"].iter_rows() for c in row]
    assert "11" in valores and "Obras civiles" in valores


# ── El contrato con los importadores del panel ────────────────
# Estos importadores viven en TypeScript, así que ningún test de aquí los
# ejecuta — y ese hueco es justo por donde se coló el fallo del 2-ago: la
# cabecera se movió a la fila 14-16 al rediseñar las plantillas y los
# importadores seguían leyendo la fila 1, así que un archivo llenado tal como
# indican las instrucciones daba 309 filas con «CODIGO vacío».
#
# `lib/excel.ts` ya no asume la fila 1: BUSCA la cabecera por estas claves. Lo
# que se verifica aquí es su parte del trato — que las claves sigan estando, y
# dentro del rango en el que las busca.
CLAVES_IMPORTADOR = {
    # plantilla   → claves que busca el importador (y dónde vive)
    "partidas":    ["CODIGO", "DESCRIPCION", "METRADO_PRESUP"],   # ImportarPartidas.tsx
    "personal":    ["NOMBRE", "CARGO", "DNI"],                    # ImportarPersonal.tsx
    "proyectos":   ["NOMBRE", "AREA", "ESTADO"],                  # OTMs.tsx
    "presupuesto": ["CODIGO", "METRADO", "PRECIO_UNITARIO"],      # Presupuesto.tsx
    "costos_ro":   ["TIPO_RECURSO", "PERIODO", "MONTO"],          # Rentabilidad.tsx
    "costos":      ["TIPO_RECURSO", "FECHA", "MONTO"],            # routers/ro.py (API)
}

MAX_BUSQUEDA = 40      # el mismo tope que `lib/excel.ts`
MIN_ACIERTOS = 2       # el mismo umbral que `lib/excel.ts`


@pytest.mark.parametrize("clave,esperadas", sorted(CLAVES_IMPORTADOR.items()))
def test_el_importador_encuentra_la_cabecera(clave, esperadas):
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]

    mejor, aciertos_max = 0, 0
    for r in range(1, min(ws.max_row, MAX_BUSQUEDA) + 1):
        fila = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        aciertos = sum(1 for v in fila if v in esperadas)
        if aciertos > aciertos_max:
            mejor, aciertos_max = r, aciertos

    assert aciertos_max >= MIN_ACIERTOS, (
        f"{clave}: el importador no reconocería la cabecera en las primeras "
        f"{MAX_BUSQUEDA} filas. Faltan claves o se movieron demasiado abajo.")
    encontradas = {ws.cell(mejor, c).value for c in range(1, ws.max_column + 1)}
    assert set(esperadas) <= encontradas, (
        f"{clave}: faltan {set(esperadas) - encontradas} en la fila {mejor}. "
        f"Si se renombró una columna, hay que renombrarla también en el panel.")


@pytest.mark.parametrize("clave", sorted(CLAVES_IMPORTADOR))
def test_los_datos_empiezan_justo_debajo_de_la_cabecera(clave):
    """`leerHoja` lee desde la fila siguiente a la cabecera. Una fila intermedia
    (una nota, un separador) entraría como un dato vacío y todas las filas
    quedarían corridas."""
    wb, _ = _libro(clave)
    p = plantillas.definir(clave, DATOS)
    ws = wb[p.hoja]
    fila_cab = next(r for r in range(1, MAX_BUSQUEDA)
                    if ws.cell(r, 2).value == p.cols[0].clave)
    assert ws.cell(fila_cab + 1, 2).font.italic, (
        f"{clave}: debajo de la cabecera debería ir ya la fila de ejemplo")


def test_la_fila_de_ejemplo_lleva_su_marca_en_el_margen():
    """`leerHoja` descarta las filas marcadas con ▸: si la marca desaparece, los
    ejemplos se importarían como datos buenos en cuanto alguien no los borre."""
    for clave in _claves():
        wb, _ = _libro(clave)
        p = plantillas.definir(clave, DATOS)
        ws = wb[p.hoja]
        fila = next(r for r in range(1, MAX_BUSQUEDA)
                    if ws.cell(r, 2).value == p.cols[0].clave) + 1
        assert ws.cell(fila, 1).value == "▸", f"{clave}: ejemplo sin marca"


# ── Ida y vuelta: la plantilla PU se lee a sí misma ───────────
def test_la_plantilla_pu_la_lee_su_propio_parser():
    wb, _ = plantillas.generar("pu", {}, "OBRA DE PRUEBA", date(2026, 8, 1))
    buf = BytesIO()
    wb.save(buf)
    r = parsear_plantilla_pu(buf.getvalue())
    assert r.errores == []
    hojas = [p for p in r.partidas if p.es_hoja]
    assert len(hojas) == 2
    assert r.hh_dia == 10.0
    for h in hojas:
        assert h.recursos, f"{h.codigo} sin APU"
        assert abs(sum(x.parcial for x in h.recursos) - h.cud) < 0.02
        assert abs(h.cud - h.pu_meta) < 0.02


def test_el_parser_sigue_leyendo_los_xls_de_siempre():
    """El presupuesto real del ex-gerente es .xls de verdad: retirar esa vía
    dejaría sin importar la obra que ya existe."""
    from pathlib import Path
    fixture = Path(__file__).parent / "fixtures" / "pu_min.xls"
    r = parsear_plantilla_pu(fixture.read_bytes())
    assert [p.codigo for p in r.partidas if p.es_hoja] == ["01.01", "01.02"]


# ── Roles ────────────────────────────────────────────────────
@pytest.fixture()
def _prod(monkeypatch):
    monkeypatch.setattr(config, "API_KEY", "svc")
    monkeypatch.setattr(config, "ENV", "prod")


def _cli():
    return TestClient(main.app, raise_server_exceptions=False)


def test_plantillas_sin_credenciales_401(_prod):
    assert _cli().get("/ev/plantillas").status_code == 401


def test_supervisor_no_descarga_plantillas(_prod):
    tk = auth.make_token("u-sup", "supervisor", "sup", {"sup_id": "01"})
    r = _cli().get("/ev/plantillas/personal", headers={"Authorization": "Bearer " + tk})
    assert r.status_code == 403


def test_plantilla_inexistente_404(_prod):
    tk = auth.make_token("u-of", "oficina", "oficina")
    r = _cli().get("/ev/plantillas/inventada", headers={"Authorization": "Bearer " + tk})
    assert r.status_code == 404


def test_el_catalogo_describe_cada_plantilla():
    filas = plantillas.listar()
    assert len(filas) == len(plantillas.CATALOGO)
    for f in filas:
        assert f["titulo"] and f["proposito"] and f["archivo"].endswith(".xlsx")
