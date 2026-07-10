# ============================================================
# routers/ro_mensual.py — F2.5/F2.7: insumos del motor + endpoints
#
#   GET /ev/ro/mensual?proyecto_id&anio&mes → JSON del motor (T OBRA + R FASES)
#   GET /ev/ro/export?proyecto_id&anio&mes  → .xlsx (hojas T OBRA y R FASES)
#
# Este módulo SOLO arma datos (SQL agregado) y llama al motor puro (ro_motor).
# ============================================================
from collections import defaultdict
from datetime import timedelta
from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from core.db import db
from routers.ro_motor import ro_mensual as motor

router = APIRouter(prefix="/ev/ro", tags=["resultado-operativo"])


async def _insumos(con, proyecto_id: int, anio: int, mes: int) -> dict:
    periodos = [dict(r) for r in await con.fetch(
        "SELECT id, anio, mes, tipo_cambio::float AS tipo_cambio, estado "
        "FROM periodos WHERE proyecto_id=$1 ORDER BY anio, mes", proyecto_id)]
    corte = next((p for p in periodos if p["anio"] == anio and p["mes"] == mes), None)
    if not corte:
        raise HTTPException(404, f"No existe el periodo {anio}-{mes:02d} (créalo en Periodos)")

    docs = [dict(r) for r in await con.fetch(
        """SELECT periodo_id, fase, tipo_recurso, directo, SUM(monto)::float AS monto
           FROM costo_documentos WHERE proyecto_id=$1
           GROUP BY periodo_id, fase, tipo_recurso, directo""", proyecto_id)]

    # MO del tareo × tarifa, por mes y fase (mapeada a periodos existentes)
    tar = {r["cargo"]: float(r["costo_hh"]) for r in
           await con.fetch("SELECT cargo, costo_hh FROM ev_tarifas_cargo")}
    default = tar.get("(Default)")
    mo_rows = await con.fetch(
        """SELECT EXTRACT(YEAR FROM tp.fecha)::int AS anio, EXTRACT(MONTH FROM tp.fecha)::int AS mes,
                  ev.fase AS fase, COALESCE(t.cargo, '(Sin cargo)') AS cargo, SUM(tp.hh)::float AS hh
           FROM tareo_partida tp
           JOIN ev_partidas ev ON ev.id = tp.partida_id
           LEFT JOIN trabajadores t ON t.id = tp.trabajador_id
           WHERE tp.hh IS NOT NULL AND ev.fase IS NOT NULL
             AND ev.otm_id IN (SELECT id FROM otms WHERE proyecto_id = $1)
           GROUP BY 1, 2, ev.fase, t.cargo""", proyecto_id)
    per_por_mes = {(p["anio"], p["mes"]): p["id"] for p in periodos}
    mo_tareo_mes: dict = defaultdict(lambda: defaultdict(float))
    for r in mo_rows:
        per = per_por_mes.get((r["anio"], r["mes"]))
        if per is None:
            continue      # mes sin periodo creado: no entra al RO (crear el periodo lo incorpora)
        rate = tar.get(r["cargo"], default) or 0.0
        mo_tareo_mes[per][r["fase"]] += r["hh"] * rate

    # Venta contractual mensual: valorización APROBADA (F2.8) o, si no hay,
    # ev_valorizado × PU mensualizado vía fecha_base (semana → lunes → mes).
    venta_fase_mes: dict = defaultdict(lambda: defaultdict(float))
    val_aprob = []
    try:
        val_aprob = await con.fetch(
            """SELECT v.periodo_id, ev.fase, SUM(vl.cantidad * vl.pu)::float AS monto
               FROM valorizaciones v
               JOIN valorizacion_lineas vl ON vl.valorizacion_id = v.id
               JOIN ev_partidas ev ON ev.id = vl.partida_id
               WHERE v.proyecto_id=$1 AND v.estado='APROBADA'
               GROUP BY v.periodo_id, ev.fase""", proyecto_id)
    except Exception:
        val_aprob = []    # F2.8 aún no migrada
    periodos_con_val = set()
    for r in val_aprob:
        venta_fase_mes[r["periodo_id"]][r["fase"]] += r["monto"]
        periodos_con_val.add(r["periodo_id"])
    from routers.ev._datos import _fecha_base
    base = await _fecha_base(con)
    if base:
        val_rows = await con.fetch(
            """SELECT v.semana, ev.fase, SUM(v.cantidad_valorizada * ev.precio_unitario)::float AS monto
               FROM ev_valorizado v JOIN ev_partidas ev ON ev.id = v.partida_id
               WHERE ev.fase IS NOT NULL
                 AND ev.otm_id IN (SELECT id FROM otms WHERE proyecto_id = $1)
               GROUP BY v.semana, ev.fase""", proyecto_id)
        for r in val_rows:
            lunes = base + timedelta(days=(int(r["semana"]) - 1) * 7)
            per = per_por_mes.get((lunes.year, lunes.month))
            if per is None or per in periodos_con_val:
                continue  # la valorización APROBADA manda en su periodo
            venta_fase_mes[per][r["fase"]] += r["monto"]

    ajustes = [dict(r) for r in await con.fetch(
        """SELECT periodo_id, tipo, fase, monto::float AS monto,
                  margen_previsto::float AS margen_previsto
           FROM venta_ajustes WHERE proyecto_id=$1 AND periodo_id IS NOT NULL""", proyecto_id)]

    async def _costo_presu(tipo: str) -> dict:
        if tipo == "META":
            rows = await con.fetch(
                """SELECT pcm.fase, pcm.tipo_recurso, pcm.monto::float AS monto
                   FROM presupuesto_costo_meta pcm JOIN presupuestos p ON p.id = pcm.presupuesto_id
                   WHERE p.proyecto_id=$1 AND p.tipo='META' AND p.vigente""", proyecto_id)
            return {(r["fase"], r["tipo_recurso"]): r["monto"] for r in rows}
        rows = await con.fetch(
            """SELECT pp.fase, SUM(pp.metrado * pp.precio_unitario)::float AS monto
               FROM presupuesto_partidas pp JOIN presupuestos p ON p.id = pp.presupuesto_id
               WHERE p.proyecto_id=$1 AND p.tipo='CONTRACTUAL' AND p.vigente
               GROUP BY pp.fase""", proyecto_id)
        return {(r["fase"], "PU"): r["monto"] for r in rows}

    proyeccion = [dict(r) for r in await con.fetch(
        """SELECT periodo_id, fase, tipo_recurso, monto::float AS monto, origen
           FROM ro_proyeccion WHERE proyecto_id=$1""", proyecto_id)]

    prev_rows = await con.fetch(
        """SELECT fase, tipo_recurso, monto::float AS monto FROM ro_prev
           WHERE proyecto_id=$1 AND periodo_id=$2""", proyecto_id, corte["id"])
    prev = ({"costo": round(sum(r["monto"] for r in prev_rows), 2),
             "celdas": [dict(r) for r in prev_rows]} if prev_rows else None)

    contingencia = await con.fetchval(
        "SELECT contingencia FROM proyectos WHERE id=$1", proyecto_id) or 0

    fases_desc = {r["fase"]: r["descripcion"] for r in await con.fetch(
        "SELECT codigo AS fase, descripcion FROM ev_partidas WHERE codigo IN "
        "(SELECT DISTINCT fase FROM ev_partidas WHERE fase IS NOT NULL)")}

    return dict(periodos=periodos, corte_id=corte["id"], docs=docs,
                mo_tareo_mes={k: dict(v) for k, v in mo_tareo_mes.items()},
                venta_fase_mes={k: dict(v) for k, v in venta_fase_mes.items()},
                ajustes=ajustes, costo_meta=await _costo_presu("META"),
                costo_contractual=await _costo_presu("CONTRACTUAL"),
                proyeccion=proyeccion, prev=prev,
                contingencia=float(contingencia), fases_desc=fases_desc)


@router.get("/mensual")
async def ro_mensual_endpoint(proyecto_id: int = 1, anio: int = 0, mes: int = 0):
    pool = await db()
    async with pool.acquire() as con:
        if not anio or not mes:
            ult = await con.fetchrow(
                "SELECT anio, mes FROM periodos WHERE proyecto_id=$1 ORDER BY anio DESC, mes DESC LIMIT 1",
                proyecto_id)
            if not ult:
                raise HTTPException(404, "No hay periodos: crea el primero en la página Costos/Periodos")
            anio, mes = ult["anio"], ult["mes"]
        ins = await _insumos(con, proyecto_id, anio, mes)
    return motor(**ins)


@router.get("/export")
async def ro_export(proyecto_id: int = 1, anio: int = 0, mes: int = 0):
    """Export .xlsx con hojas T OBRA y R FASES (espejo del JSON del motor)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    pool = await db()
    async with pool.acquire() as con:
        if not anio or not mes:
            ult = await con.fetchrow(
                "SELECT anio, mes FROM periodos WHERE proyecto_id=$1 ORDER BY anio DESC, mes DESC LIMIT 1",
                proyecto_id)
            if not ult:
                raise HTTPException(404, "No hay periodos")
            anio, mes = ult["anio"], ult["mes"]
        ins = await _insumos(con, proyecto_id, anio, mes)
        nombre = await con.fetchval("SELECT nombre FROM proyectos WHERE id=$1", proyecto_id)
    out = motor(**ins)

    wb = Workbook()
    negrita = Font(bold=True)
    num = "#,##0.00"

    # ── hoja T OBRA ──
    ws = wb.active
    ws.title = "T OBRA"
    ws.append([f"RESULTADO OPERATIVO — {nombre or f'Proyecto {proyecto_id}'}"])
    ws.append([f"Mes de corte: {anio}-{mes:02d} · TC {out['corte']['tipo_cambio']}"])
    ws.append([])
    ws.append(["CONCEPTO", "MES", "ACUM", "MARGEN PREV. ACUM"])
    for c in ws[4]:
        c.font = negrita
    for cpt in out["t_obra"]["venta"]:
        ws.append([cpt["concepto"], cpt["mes"], cpt["acum"], cpt["margen_previsto_acum"]])
    vt = out["t_obra"]["venta_total"]
    ws.append(["TOTAL VENTA S/.", vt["mes"], vt["acum"], ""])
    ws.append([])
    ws.append(["COSTO", "MES", "ACUM"])
    for grupo, titulo in (("costo_directo", "DIRECTO"), ("costo_indirecto", "INDIRECTO")):
        for rec, v_mes in out["t_obra"][grupo]["mes"].items():
            v_acum = out["t_obra"][grupo]["acum"][rec]
            if v_mes or v_acum:
                ws.append([f"{titulo} · {rec}", v_mes, v_acum])
    ct = out["t_obra"]["costo_total"]
    ws.append(["TOTAL COSTO S/.", ct["mes"], ct["acum"]])
    ws.append(["SALDO DE OBRA (proyección)", "", out["t_obra"]["saldo_obra"]])
    ws.append([])
    t = out["totales"]
    for k, label in (("costo_aplicado_acum", "COSTO APLICADO"),
                     ("resultado_pendiente", "RESULTADO PENDIENTE"),
                     ("margen_economico_acum", "MARGEN ECONÓMICO (acum)"),
                     ("margen_total", "MARGEN TOTAL"), ("pct_margen", "% MARGEN"),
                     ("contingencia", "CONTINGENCIA"),
                     ("margen_con_contingencia", "MARGEN C/CONTINGENCIA")):
        ws.append([label, t[k]])

    # ── hoja R FASES ──
    ws2 = wb.create_sheet("R FASES")
    cab = ["FASE", "DESCRIPCIÓN", "MAT", "MO", "EQP", "EQT", "SUB", "DIR", "GG",
           "TOTAL", "VENTA", "MARGEN", "%", "META", "CONTRACTUAL"]
    ws2.append(cab)
    for c in ws2[1]:
        c.font = negrita
    for f in out["r_fases"]:
        ws2.append([f.get("fase") or "", f.get("descripcion") or "",
                    f["mat"], f["mo"], f["eqp"], f["eqt"], f["sub"], f["dir"], f["gg"],
                    f["total"], f["venta"], f["margen"], f["pct_margen"],
                    f["meta"], f["contractual"]])
        if f.get("fase") is None:
            for c in ws2[ws2.max_row]:
                c.font = negrita
    for hoja in (ws, ws2):
        for fila in hoja.iter_rows(min_row=2):
            for c in fila:
                if isinstance(c.value, float):
                    c.number_format = num

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"RO_{anio}-{mes:02d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})
