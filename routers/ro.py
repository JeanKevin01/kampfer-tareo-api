# ============================================================
# routers/ro.py — Fase 2: Resultado Operativo (RO) "a la fecha"
#
# Por fase: VENTA − COSTO = MARGEN (%), con el costo abierto en recursos.
#   · Costo MO   = HH del tareo (tareo_partida) × tarifa por cargo (ev_tarifas_cargo).
#   · Costo no-MO = tabla `costos` (MAT/EQP/EQT/SUB directos; DIR/GG indirectos).
#   · Venta      = Σ(cantidad valorizada × PU del presupuesto) + `venta_ajustes`.
#
# Se monta en main.py con require_role("oficina").
# ============================================================
import hashlib
from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Optional, List

from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel

# F0.4: este router usa el pool asyncpg único (core.db); la lib `databases` salió de aquí.
from core.db import db
from core.log import get_logger
from routers.periodos import exigir_abierto, periodo_para_movimiento

log = get_logger("ro")

router = APIRouter(prefix="/ev/ro", tags=["resultado-operativo"])

_TIPOS_REC = ("MAT", "EQP", "EQT", "SUB", "DIR", "GG", "MO")
_TIPOS_DOC = ("FACTURA", "OC", "VALE", "PLANILLA_AJUSTE", "OTRO")


# ── Modelos de carga ──────────────────────────────────────────
class CostoIn(BaseModel):
    fase: Optional[str] = None
    tipo_recurso: str                 # MAT | EQP | EQT | SUB | DIR | GG
    directo: bool = True
    periodo: str                      # 'YYYY-MM-DD' (día 1 del mes)
    monto: float = 0
    fuente: Optional[str] = None
    nota: Optional[str] = None


class CostosIn(BaseModel):
    proyecto_id: int = 1
    costos: List[CostoIn]


class AjusteIn(BaseModel):
    fase: Optional[str] = None
    # F2.4: los 6 conceptos del T OBRA ('ADICIONAL' se acepta 1 versión como alias)
    tipo: str        # CONTRACTUAL | DIF_METRADO | NUEVAS_PARTIDAS | POR_APROBAR | REAJUSTE | TERCEROS
    periodo: Optional[str] = None
    monto: float = 0
    margen_previsto: Optional[float] = None
    nota: Optional[str] = None


_TIPOS_VENTA = ("CONTRACTUAL", "DIF_METRADO", "NUEVAS_PARTIDAS", "POR_APROBAR",
                "REAJUSTE", "TERCEROS")
_ALIAS_VENTA = {"ADICIONAL": "NUEVAS_PARTIDAS"}   # compat 1 versión


class AjustesIn(BaseModel):
    proyecto_id: int = 1
    ajustes: List[AjusteIn]


def _d(s: Optional[str]) -> Optional[date]:
    return date.fromisoformat(s[:10]) if s else None


# ── Lógica pura (testeable sin BD) ────────────────────────────
def _calc_mo_por_fase(hh_rows, tarifas: dict, default) -> dict:
    """Costo de MO por fase = Σ (HH del cargo × tarifa). hh_rows: [{fase,cargo,hh}]."""
    out: dict = defaultdict(float)
    for r in hh_rows:
        rate = tarifas.get(r["cargo"])
        if rate is None:
            rate = default
        if rate is None:
            rate = 0.0
        out[r["fase"]] += float(r["hh"] or 0) * rate
    return dict(out)


def _armar_ro(mo_por_fase: dict, costos, venta_real: dict, ajustes, fases_desc: dict):
    """Arma el cuadro RO. Función pura → testeable.
    costos: [{fase,tipo_recurso,directo,monto}] · ajustes: [{fase,tipo,monto}]."""
    dir_por_fase = defaultdict(lambda: defaultdict(float))   # fase -> tipo -> monto (directos)
    ind_por_tipo = defaultdict(float)                        # tipo -> monto (indirectos)
    for c in costos:
        if c["directo"]:
            dir_por_fase[c["fase"]][c["tipo_recurso"]] += float(c["monto"] or 0)
        else:
            ind_por_tipo[c["tipo_recurso"]] += float(c["monto"] or 0)

    ajuste_por_fase = defaultdict(float)
    ajuste_sin_fase = 0.0
    for a in ajustes:
        if a["fase"]:
            ajuste_por_fase[a["fase"]] += float(a["monto"] or 0)
        else:
            ajuste_sin_fase += float(a["monto"] or 0)

    fases = (set(mo_por_fase) | set(dir_por_fase) | set(venta_real) | set(ajuste_por_fase))
    fases.discard(None)

    filas = []
    tot_costo_dir = tot_venta = 0.0
    for f in sorted(fases):
        d = dir_por_fase.get(f, {})
        mat, eqp, eqt, sub = (d.get("MAT", 0.0), d.get("EQP", 0.0), d.get("EQT", 0.0), d.get("SUB", 0.0))
        mo = float(mo_por_fase.get(f, 0.0))
        costo = round(mat + mo + eqp + eqt + sub, 2)
        venta = round(float(venta_real.get(f, 0.0)) + ajuste_por_fase.get(f, 0.0), 2)
        margen = round(venta - costo, 2)
        filas.append({
            "fase": f, "descripcion": fases_desc.get(f),
            "mat": round(mat, 2), "mo": round(mo, 2), "eqp": round(eqp, 2),
            "eqt": round(eqt, 2), "sub": round(sub, 2),
            "costo": costo, "venta": venta, "margen": margen,
            "pct_margen": round(margen / venta, 4) if venta > 0 else 0,
        })
        tot_costo_dir += costo
        tot_venta += venta

    costo_ind = round(sum(ind_por_tipo.values()), 2)
    venta_total = round(tot_venta + ajuste_sin_fase, 2)
    costo_total = round(tot_costo_dir + costo_ind, 2)
    margen_total = round(venta_total - costo_total, 2)
    return {
        "fases": filas,
        "indirectos": {"DIR": round(ind_por_tipo.get("DIR", 0.0), 2),
                       "GG": round(ind_por_tipo.get("GG", 0.0), 2),
                       "total": costo_ind},
        "totales": {
            "costo_directo": round(tot_costo_dir, 2), "costo_indirecto": costo_ind,
            "costo_total": costo_total, "venta": venta_total, "margen": margen_total,
            "pct_margen": round(margen_total / venta_total, 4) if venta_total > 0 else 0,
        },
    }


# SQL: HH del tareo por fase y cargo (join directo: 0009 garantiza ids con pad + FK).
_HH_FASE_CARGO_SQL = """
    SELECT ev.fase AS fase, COALESCE(t.cargo, '(Sin cargo)') AS cargo, SUM(tp.hh) AS hh
    FROM tareo_partida tp
    JOIN ev_partidas ev ON ev.id = tp.partida_id
    LEFT JOIN trabajadores t ON t.id = tp.trabajador_id
    WHERE tp.hh IS NOT NULL AND ev.fase IS NOT NULL
      AND ev.otm_id IN (SELECT id FROM otms WHERE proyecto_id = $1)
    GROUP BY ev.fase, t.cargo
"""
# SQL: venta real por fase = Σ(cantidad valorizada × PU).
_VENTA_FASE_SQL = """
    SELECT ev.fase AS fase, SUM(v.cantidad_valorizada * ev.precio_unitario) AS venta
    FROM ev_valorizado v
    JOIN ev_partidas ev ON ev.id = v.partida_id
    WHERE ev.fase IS NOT NULL
      AND ev.otm_id IN (SELECT id FROM otms WHERE proyecto_id = $1)
    GROUP BY ev.fase
"""


# ── Endpoints ─────────────────────────────────────────────────
@router.get("")
async def resultado_operativo(proyecto_id: int = 1):
    pool = await db()
    async with pool.acquire() as con:
        hh_rows = await con.fetch(_HH_FASE_CARGO_SQL, proyecto_id)
        tar_rows = await con.fetch("SELECT cargo, costo_hh FROM ev_tarifas_cargo")
        # F2.2: el costo no-MO viene de los documentos. La MO documental (ajustes de
        # planilla) se excluye aquí: este RO "a la fecha" calcula MO = tareo × tarifa.
        costos = [dict(r) for r in await con.fetch(
            "SELECT fase, tipo_recurso, directo, SUM(monto) AS monto FROM costo_documentos "
            "WHERE proyecto_id = $1 AND tipo_recurso <> 'MO' "
            "GROUP BY fase, tipo_recurso, directo", proyecto_id)]
        venta_rows = await con.fetch(_VENTA_FASE_SQL, proyecto_id)
        ajustes = [dict(r) for r in await con.fetch(
            "SELECT fase, tipo, SUM(monto) AS monto FROM venta_ajustes "
            "WHERE proyecto_id = $1 GROUP BY fase, tipo", proyecto_id)]
        fases_rows = await con.fetch(
            "SELECT codigo AS fase, descripcion FROM ev_partidas WHERE codigo IN "
            "(SELECT DISTINCT fase FROM ev_partidas WHERE fase IS NOT NULL)")

    tarifas = {r["cargo"]: float(r["costo_hh"]) for r in tar_rows}
    default = tarifas.get("(Default)")
    mo = _calc_mo_por_fase([dict(r) for r in hh_rows], tarifas, default)
    venta_real = {r["fase"]: float(r["venta"] or 0) for r in venta_rows}
    fases_desc = {r["fase"]: r["descripcion"] for r in fases_rows}

    return _armar_ro(mo, costos, venta_real, ajustes, fases_desc)


# ── F2.2: documentos de costo (nivel factura). Los paths /costos se conservan
#    por compatibilidad con el panel; ahora leen/escriben costo_documentos. ──
_DOC_SQL = """
    SELECT cd.*, p.anio, p.mes, p.estado AS periodo_estado
    FROM costo_documentos cd JOIN periodos p ON p.id = cd.periodo_id
    WHERE cd.proyecto_id = $1 {filtros}
    ORDER BY p.anio DESC, p.mes DESC, cd.id DESC
"""


@router.get("/costos")
async def listar_costos(proyecto_id: int = 1, periodo_id: Optional[int] = None,
                        tipo_recurso: Optional[str] = None, fase: Optional[str] = None):
    pool = await db()
    filtros, args = "", [proyecto_id]
    if periodo_id:
        args.append(periodo_id); filtros += f" AND cd.periodo_id = ${len(args)}"
    if tipo_recurso:
        args.append(tipo_recurso.upper()); filtros += f" AND cd.tipo_recurso = ${len(args)}"
    if fase:
        args.append(fase); filtros += f" AND cd.fase = ${len(args)}"
    rows = await pool.fetch(_DOC_SQL.format(filtros=filtros), *args)
    docs = [dict(r) for r in rows]
    total = sum(float(d["monto"] or 0) for d in docs)
    return {"documentos": docs, "total": round(total, 2), "n": len(docs)}


class DocIn(BaseModel):
    proyecto_id: int = 1
    tipo_doc: str = "FACTURA"
    proveedor: Optional[str] = None
    numero_doc: Optional[str] = None
    fecha: Optional[str] = None       # 'YYYY-MM-DD' → resuelve el periodo (lo crea si falta)
    periodo_id: Optional[int] = None  # alternativa explícita
    tipo_recurso: str
    directo: bool = True
    fase: Optional[str] = None
    partida_id: Optional[int] = None
    moneda: str = "PEN"
    monto: float
    glosa: Optional[str] = None


@router.post("/documentos")
async def crear_documento(body: DocIn):
    """Alta rápida de un documento de costo (formulario de la página Costos)."""
    if body.tipo_recurso.upper() not in _TIPOS_REC:
        raise HTTPException(400, f"tipo_recurso inválido (válidos: {_TIPOS_REC})")
    if body.tipo_doc.upper() not in _TIPOS_DOC:
        raise HTTPException(400, f"tipo_doc inválido (válidos: {_TIPOS_DOC})")
    if body.tipo_recurso.upper() == "MO" and body.tipo_doc.upper() != "PLANILLA_AJUSTE":
        raise HTTPException(400, "La MO solo entra como PLANILLA_AJUSTE (el resto sale del tareo × tarifa)")
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            per = await periodo_para_movimiento(con, body.proyecto_id, _d(body.fecha), body.periodo_id)
            did = await con.fetchval(
                """INSERT INTO costo_documentos
                   (proyecto_id, periodo_id, tipo_doc, proveedor, numero_doc, fecha,
                    tipo_recurso, directo, fase, partida_id, moneda, monto, glosa, fuente)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,'manual')
                   RETURNING id""",
                body.proyecto_id, per, body.tipo_doc.upper(), body.proveedor, body.numero_doc,
                _d(body.fecha), body.tipo_recurso.upper(), body.directo, body.fase,
                body.partida_id, body.moneda.upper(), body.monto, body.glosa)
    return {"ok": True, "id": did, "periodo_id": per}


@router.delete("/documentos/{did}")
async def borrar_documento(did: int):
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            per = await con.fetchval("SELECT periodo_id FROM costo_documentos WHERE id=$1", did)
            if per is None:
                raise HTTPException(404, "Documento no encontrado")
            await exigir_abierto(con, per)
            await con.execute("DELETE FROM costo_documentos WHERE id=$1", did)
    return {"ok": True}


@router.post("/costos")
async def cargar_costos(body: CostosIn):
    """Compat con el flujo anterior: cada costo agregado entra como documento 'OTRO'."""
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            for c in body.costos:
                per = await periodo_para_movimiento(con, body.proyecto_id, _d(c.periodo))
                await con.execute(
                    """INSERT INTO costo_documentos
                       (proyecto_id, periodo_id, tipo_doc, fecha, tipo_recurso, directo,
                        fase, monto, glosa, fuente)
                       VALUES ($1,$2,'OTRO',$3,$4,$5,$6,$7,$8,COALESCE($9,'manual'))""",
                    body.proyecto_id, per, _d(c.periodo), c.tipo_recurso, c.directo,
                    c.fase, c.monto, c.nota, c.fuente)
    return {"ok": True, "cargados": len(body.costos)}


@router.delete("/costos/{cid}")
async def borrar_costo(cid: int):
    return await borrar_documento(cid)


# ── F2.2: import masivo de documentos (.xlsx, plantilla propia) ──
_COLS_IMPORT = ["PROVEEDOR", "NUMERO_DOC", "FECHA", "TIPO_DOC", "TIPO_RECURSO",
                "DIRECTO", "FASE", "MONEDA", "MONTO", "GLOSA"]


def _parse_xlsx_costos(contenido: bytes):
    """(filas, errores) desde la hoja 1. Cada error lleva su nº de fila."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    filas, errores = [], []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = ["" if v is None else str(v).strip() for v in row]
        if header is None:
            if "TIPO_RECURSO" in [v.upper() for v in vals]:
                header = [v.upper() for v in vals]
            continue
        celda = dict(zip(header, vals))
        if not any(celda.values()):
            continue
        tr = celda.get("TIPO_RECURSO", "").upper()
        if tr not in _TIPOS_REC:
            errores.append(f"fila {i}: TIPO_RECURSO inválido '{tr}'")
            continue
        td = (celda.get("TIPO_DOC") or "FACTURA").upper()
        if td not in _TIPOS_DOC:
            errores.append(f"fila {i}: TIPO_DOC inválido '{td}'")
            continue
        if tr == "MO" and td != "PLANILLA_AJUSTE":
            errores.append(f"fila {i}: MO solo entra como PLANILLA_AJUSTE")
            continue
        try:
            monto = float(str(celda.get("MONTO", "")).replace(",", ""))
        except ValueError:
            errores.append(f"fila {i}: MONTO ilegible '{celda.get('MONTO')}'")
            continue
        fecha = _d(celda.get("FECHA") or None)
        if not fecha:
            errores.append(f"fila {i}: FECHA requerida (YYYY-MM-DD)")
            continue
        filas.append({
            "proveedor": celda.get("PROVEEDOR") or None,
            "numero_doc": celda.get("NUMERO_DOC") or None,
            "fecha": fecha, "tipo_doc": td, "tipo_recurso": tr,
            "directo": (celda.get("DIRECTO") or "SI").upper() not in ("NO", "FALSE", "0"),
            "fase": celda.get("FASE") or None,
            "moneda": (celda.get("MONEDA") or "PEN").upper(),
            "monto": monto, "glosa": celda.get("GLOSA") or None,
        })
    if header is None:
        errores.append("No se encontró la fila de encabezados (debe incluir TIPO_RECURSO). "
                       f"Columnas esperadas: {_COLS_IMPORT}")
    return filas, errores


@router.post("/costos/importar")
async def importar_costos(proyecto_id: int = 1, confirmar: bool = False,
                          file: UploadFile = File(...)):
    """Import masivo de documentos. Idempotente: la fuente es el hash del archivo;
    reimportarlo REEMPLAZA sus documentos anteriores (delete+insert en transacción)."""
    contenido = await file.read()
    try:
        filas, errores = _parse_xlsx_costos(contenido)
    except Exception:
        log.exception("importar costos: archivo ilegible")
        raise HTTPException(400, "El archivo no es un .xlsx legible")
    fuente = "import:" + hashlib.sha1(contenido).hexdigest()[:12]
    total = round(sum(f["monto"] for f in filas), 2)
    if not confirmar:
        return {"resumen": {"filas": len(filas), "total": total, "fuente": fuente,
                            "errores": errores},
                "muestra": [{**f, "fecha": f["fecha"].isoformat()} for f in filas[:10]]}
    if errores:
        raise HTTPException(409, "El archivo tiene errores; corrige y reintenta:\n· "
                                 + "\n· ".join(errores[:10]))
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            await con.execute(
                "DELETE FROM costo_documentos WHERE proyecto_id=$1 AND fuente=$2",
                proyecto_id, fuente)
            for f in filas:
                per = await periodo_para_movimiento(con, proyecto_id, f["fecha"])
                await con.execute(
                    """INSERT INTO costo_documentos
                       (proyecto_id, periodo_id, tipo_doc, proveedor, numero_doc, fecha,
                        tipo_recurso, directo, fase, moneda, monto, glosa, fuente)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)""",
                    proyecto_id, per, f["tipo_doc"], f["proveedor"], f["numero_doc"],
                    f["fecha"], f["tipo_recurso"], f["directo"], f["fase"],
                    f["moneda"], f["monto"], f["glosa"], fuente)
    return {"ok": True, "importados": len(filas), "total": total, "fuente": fuente}


@router.get("/venta-ajustes")
async def listar_ajustes(proyecto_id: int = 1):
    pool = await db()
    rows = await pool.fetch(
        "SELECT * FROM venta_ajustes WHERE proyecto_id = $1 ORDER BY id DESC", proyecto_id)
    return [dict(r) for r in rows]


@router.post("/venta-ajustes")
async def cargar_ajustes(body: AjustesIn):
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            for a in body.ajustes:
                tipo = _ALIAS_VENTA.get(a.tipo.upper(), a.tipo.upper())
                if tipo not in _TIPOS_VENTA:
                    raise HTTPException(400, f"tipo inválido '{a.tipo}' (válidos: {_TIPOS_VENTA})")
                fch = _d(a.periodo)
                per = await periodo_para_movimiento(con, body.proyecto_id, fch) if fch else None
                await con.execute(
                    """INSERT INTO venta_ajustes
                       (proyecto_id, fase, tipo, periodo, periodo_id, monto, margen_previsto, nota)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8)""",
                    body.proyecto_id, a.fase, tipo, fch, per, a.monto, a.margen_previsto, a.nota,
                )
    return {"ok": True, "cargados": len(body.ajustes)}


@router.delete("/venta-ajustes/{aid}")
async def borrar_ajuste(aid: int):
    pool = await db()
    async with pool.acquire() as con:
        async with con.transaction():
            row = await con.fetchrow("SELECT periodo_id FROM venta_ajustes WHERE id=$1", aid)
            if row is None:
                raise HTTPException(404, "Ajuste no encontrado")
            if row["periodo_id"]:
                await exigir_abierto(con, row["periodo_id"])
            await con.execute("DELETE FROM venta_ajustes WHERE id = $1", aid)
    return {"ok": True}


# ── F2.3: ajuste de MO contra la planilla real del mes ────────
def calcular_ajuste_mo(hh_rows, tarifas: dict, default, planilla_real: float,
                       distribuir_por_fase: bool = True) -> dict:
    """Función PURA (testeable): compara el costo MO del tareo (hh × tarifa) contra
    la planilla real y propone el/los documentos de ajuste.
    hh_rows: [{fase, cargo, hh}] del mes."""
    por_fase = _calc_mo_por_fase(hh_rows, tarifas, default)
    mo_tareo = round(sum(por_fase.values()), 2)
    delta = round(float(planilla_real) - mo_tareo, 2)
    if not distribuir_por_fase or mo_tareo <= 0:
        ajustes = [{"fase": None, "monto": delta}]
    else:
        ajustes, acum = [], 0.0
        fases = sorted(por_fase.items())
        for i, (fase, mo) in enumerate(fases):
            if i == len(fases) - 1:
                m = round(delta - acum, 2)      # el último absorbe el redondeo
            else:
                m = round(delta * (mo / mo_tareo), 2)
                acum += m
            ajustes.append({"fase": fase, "monto": m})
    return {"mo_tareo": mo_tareo, "planilla_real": round(float(planilla_real), 2),
            "delta": delta, "por_fase": {k: round(v, 2) for k, v in por_fase.items()},
            "ajustes": ajustes}


_HH_FASE_CARGO_MES_SQL = """
    SELECT ev.fase AS fase, COALESCE(t.cargo, '(Sin cargo)') AS cargo, SUM(tp.hh) AS hh
    FROM tareo_partida tp
    JOIN ev_partidas ev ON ev.id = tp.partida_id
    LEFT JOIN trabajadores t ON t.id = tp.trabajador_id
    WHERE tp.hh IS NOT NULL AND ev.fase IS NOT NULL
      AND ev.otm_id IN (SELECT id FROM otms WHERE proyecto_id = $1)
      AND EXTRACT(YEAR FROM tp.fecha)::int = $2 AND EXTRACT(MONTH FROM tp.fecha)::int = $3
    GROUP BY ev.fase, t.cargo
"""


@router.post("/ajuste-mo")
async def ajuste_mo(data: dict):
    """{proyecto_id, periodo_id, planilla_real, distribuir_por_fase?, confirmar?}
    Preview: {mo_tareo, planilla_real, delta, ajustes}. Con confirmar=true inserta
    los documentos PLANILLA_AJUSTE/MO (reemplaza el ajuste anterior del periodo)."""
    proyecto_id = int(data.get("proyecto_id") or 1)
    periodo_id = data.get("periodo_id")
    if not periodo_id:
        raise HTTPException(400, "periodo_id requerido")
    try:
        planilla_real = float(data.get("planilla_real"))
    except (TypeError, ValueError):
        raise HTTPException(400, "planilla_real numérico requerido")
    distribuir = bool(data.get("distribuir_por_fase", True))

    pool = await db()
    async with pool.acquire() as con:
        per = await con.fetchrow("SELECT * FROM periodos WHERE id=$1", int(periodo_id))
        if not per:
            raise HTTPException(404, "Periodo no encontrado")
        hh_rows = [dict(r) for r in await con.fetch(
            _HH_FASE_CARGO_MES_SQL, proyecto_id, per["anio"], per["mes"])]
        tar = {r["cargo"]: float(r["costo_hh"])
               for r in await con.fetch("SELECT cargo, costo_hh FROM ev_tarifas_cargo")}
        calc = calcular_ajuste_mo(hh_rows, tar, tar.get("(Default)"), planilla_real, distribuir)
        if not data.get("confirmar"):
            return calc
        await exigir_abierto(con, int(periodo_id))
        fuente = f"ajuste-mo:{periodo_id}"
        async with con.transaction():
            await con.execute(
                "DELETE FROM costo_documentos WHERE proyecto_id=$1 AND periodo_id=$2 AND fuente=$3",
                proyecto_id, int(periodo_id), fuente)
            for aj in calc["ajustes"]:
                if abs(aj["monto"]) < 0.005:
                    continue
                await con.execute(
                    """INSERT INTO costo_documentos
                       (proyecto_id, periodo_id, tipo_doc, tipo_recurso, directo, fase,
                        monto, glosa, fuente)
                       VALUES ($1,$2,'PLANILLA_AJUSTE','MO',true,$3,$4,$5,$6)""",
                    proyecto_id, int(periodo_id), aj["fase"], aj["monto"],
                    f"Ajuste MO vs planilla real {per['anio']}-{per['mes']:02d} "
                    f"(tareo {calc['mo_tareo']} vs planilla {calc['planilla_real']})",
                    fuente)
    return {"ok": True, **calc}
