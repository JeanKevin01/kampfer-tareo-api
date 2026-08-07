# -*- coding: utf-8 -*-
"""Validación end-to-end LOCAL del flujo tareo → ISP (hito de la Fase 0).

Simula el circuito completo contra una BD local migrada:
  supervisor envía tareo (enviar-con-partidas) → tareo_partida →
  captura de avance → /ev/reporte y /ev/isp muestran HH gastadas y ganadas.
También verifica la idempotencia del reenvío (mismo supervisor/OTM/día REEMPLAZA).

Uso (Windows Git Bash / Linux):
  1) Postgres local:  docker run -d --name kampfer-e2e -e POSTGRES_PASSWORD=test \
       -e POSTGRES_DB=kampfer_test -p 55432:5432 postgres:16
  2) Migrar:          DATABASE_URL=postgresql://postgres:test@localhost:55432/kampfer_test \
       python -m alembic upgrade head
  3) API local:       (mismo DATABASE_URL) ENV=dev python -m uvicorn main:app --port 8765
  4) Este script:     (mismo DATABASE_URL) API_URL=http://localhost:8765 \
       python scripts/validacion_e2e.py

Idempotente: limpia sus propios datos (OTM-E2E / E2E-%) al inicio.
Sale con código 0 si TODO pasa; 1 si algo falla.
"""
import json
import os
import sys

import httpx
import psycopg2

API = os.environ.get("API_URL", "http://localhost:8765")
DB = os.environ["DATABASE_URL"]

FECHA_BASE = "2026-06-01"   # lunes → semana 1 = 2026-06-01..07
FECHA_TAREO = "2026-06-02"  # martes de la semana 1

# ── Semilla del WBS: HOJAS (con fase) vs CABECERA (fase NULL) ────────────────
# Los totales del EV se calculan SOLO sobre hojas. Que estos números vivan aquí
# permite que EV3 compruebe la suma exacta en vez de afirmarlo de palabra.
HOJAS_EV = (("E2E-001", 100, 10), ("E2E-002", 50, 5))   # (codigo, hh_presup, metrado)
HH_PARTIDA_PROG = 100000        # las 12 partidas desechables de programación
N_PARTIDAS_PROG = 12
HH_CABECERA_WBS = 777           # HH de la cabecera del WBS. NO puede aparecer
METRADO_CABECERA_WBS = 999      # en ningún total: ver el check EV3.

_fallas = []


def check(nombre, cond, detalle=""):
    estado = "OK  " if cond else "FALLA"
    print(f"[{estado}] {nombre}" + (f" — {detalle}" if detalle and not cond else ""))
    if not cond:
        _fallas.append(nombre)


def limpiar_y_sembrar(cur):
    cur.execute("DELETE FROM campo_fotos WHERE reporte_id IN "
                "(SELECT id FROM campo_reportes WHERE otm_id='OTM-E2E')")
    cur.execute("DELETE FROM campo_reportes WHERE otm_id='OTM-E2E'")
    # La traza de restricciones es append-only y NO tiene FK a la restricción
    # (esa es su gracia: sobrevive al borrado), así que hay que limpiarla ANTES
    # de que el DELETE de actividades se lleve las restricciones por cascada —
    # después ya no habría por dónde encontrarlas.
    cur.execute("DELETE FROM prog_restriccion_eventos WHERE actividad_id IN "
                "(SELECT id FROM prog_actividades WHERE titulo LIKE 'E2E %' OR otm_id='OTM-E2E')")
    cur.execute("DELETE FROM prog_actividades WHERE titulo LIKE 'E2E %' OR otm_id='OTM-E2E'")
    cur.execute("DELETE FROM prog_feriados WHERE proyecto_id=1")
    cur.execute("DELETE FROM prog_config WHERE proyecto_id=1")
    # Compromiso y bitácora de las semanas que toca el humo (0041). La bitácora
    # es append-only y NO se puede borrar por API —esa es su gracia—, así que el
    # humo la limpia por SQL: si no, el segundo run encuentra los eventos del
    # primero y el conteo de «veces comprometida» sale acumulado.
    cur.execute("DELETE FROM prog_semana_plan WHERE proyecto_id=1")
    cur.execute("DELETE FROM prog_semana_eventos WHERE proyecto_id=1")
    cur.execute("DELETE FROM fases WHERE codigo LIKE 'E2E-%'")
    # El responsable de humo se borra DESPUÉS de las actividades: sus
    # restricciones lo referencian y el DELETE de arriba ya se las llevó.
    cur.execute("DELETE FROM prog_responsables WHERE nombre IN ('LOGISTICA','INGENIERIA')")
    cur.execute("DELETE FROM tareo_ediciones WHERE otm_id='OTM-E2E'")
    cur.execute("DELETE FROM tareo_partida WHERE otm_id='OTM-E2E'")
    cur.execute("DELETE FROM registros WHERE otm_id='OTM-E2E'")
    cur.execute("DELETE FROM sesion_trabajadores WHERE sesion_id IN "
                "(SELECT id FROM sesiones WHERE otm_id='OTM-E2E')")
    cur.execute("DELETE FROM sesiones WHERE otm_id='OTM-E2E'")
    # El segundo supervisor de P85 es de humo y NO tiene ficha de trabajador: si
    # sobrevive al run, P46 (todo supervisor activo tiene ficha) lo encuentra en
    # la siguiente pasada y falla. Se borra con su acceso, que se lo crea sola la
    # semilla del arranque del API.
    cur.execute("DELETE FROM usuarios WHERE supervisor_id='SUPE2F'")
    cur.execute("DELETE FROM cuadrillas WHERE supervisor_id='SUPE2F'")
    # Cuadrillas de F-CUADRILLA y F-CATALOGO (los miembros se van por cascada).
    # Se filtra por NOMBRE y no por quién la creó: desde 0048 las cuadrillas son
    # libres y las del catálogo no tienen autor, así que un filtro por supervisor
    # dejaba basura viva y la pasada siguiente chocaba de nombre contra ella.
    cur.execute("DELETE FROM cuadrilla_grupos WHERE nombre LIKE 'E2E %'")
    cur.execute("DELETE FROM cuadrilla_grupos WHERE creada_por IN ('SUPE2E','SUPE2F')")
    cur.execute("DELETE FROM supervisores WHERE id='SUPE2F'")
    cur.execute("DELETE FROM ev_avances WHERE hito_id IN (SELECT id FROM ev_hitos WHERE "
                "partida_id IN (SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%'))")
    cur.execute("DELETE FROM ev_hh_gastadas WHERE partida_id IN "
                "(SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%')")
    cur.execute("DELETE FROM ev_hitos WHERE partida_id IN "
                "(SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%')")
    # Residuos de smokes de F2 en BDs locales reutilizadas:
    cur.execute("DELETE FROM ev_avances_diarios WHERE partida_id IN "
                "(SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%')")
    cur.execute("DELETE FROM valorizacion_lineas WHERE partida_id IN "
                "(SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%')")
    cur.execute("DELETE FROM ev_valorizado WHERE partida_id IN "
                "(SELECT id FROM ev_partidas WHERE codigo LIKE 'E2E-%')")
    cur.execute("DELETE FROM ev_partidas WHERE codigo LIKE 'E2E-%'")

    # Usuarios y supervisores creados por los checks del padrón (P42-P44):
    # los usuarios van PRIMERO (referencian supervisores por FK).
    cur.execute("DELETE FROM usuarios WHERE username IN ('supe2e','e2etrab','mpacheco','lticona') "
                "OR supervisor_id IN (SELECT id FROM supervisores WHERE nombre LIKE '%E2E%')")
    cur.execute("DELETE FROM supervisores WHERE trabajador_id IN ('901','902') "
                "OR nombre LIKE 'PACHECO ROJAS%' OR nombre LIKE 'TICONA HUANCA%'")
    cur.execute("DELETE FROM trabajadores WHERE nombre IN "
                "('PACHECO ROJAS MARIO E2E','TICONA HUANCA LUIS E2E')")

    cur.execute("INSERT INTO trabajadores (id, nombre, cargo) VALUES "
                "('900','Supervisor E2E','SUPERVISOR'),"
                "('901','Trabajador E2E Uno','OFICIAL'),('902','Trabajador E2E Dos','OPERARIO') "
                "ON CONFLICT (id) DO NOTHING")
    # Padrón unificado: el supervisor de prueba también tiene su ficha (0031)
    cur.execute("INSERT INTO supervisores (id, nombre, trabajador_id) "
                "VALUES ('SUPE2E','Supervisor E2E','900') "
                "ON CONFLICT (id) DO UPDATE SET trabajador_id = '900'")
    cur.execute("DELETE FROM otms WHERE descripcion LIKE 'E2E PROYECTO%'")
    # El ÁREA es del proyecto/OTM (0033): el parte de campo ya no la digita,
    # la hereda. Lo que el supervisor escribe ahora es el FRENTE.
    cur.execute("INSERT INTO otms (id, descripcion, proyecto_id, area) VALUES "
                "('OTM-E2E','OTM de validación E2E',1,'PLANTA SX / EW') "
                "ON CONFLICT (id) DO UPDATE SET area = 'PLANTA SX / EW'")
    cur.execute("INSERT INTO ev_config (clave, valor) VALUES ('fecha_base', %s) "
                "ON CONFLICT (clave) DO UPDATE SET valor=%s", (FECHA_BASE, FECHA_BASE))

    partidas = {}
    for codigo, hh_presup, metrado in HOJAS_EV:
        cur.execute(
            "INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, metrado_presup, "
            "hh_presup, otm_id) VALUES (%s,'F-E2E',%s,'und',%s,%s,'OTM-E2E') RETURNING id",
            (codigo, f"Partida {codigo}", metrado, hh_presup))
        pid = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
            "VALUES (%s,1,'Ejecución',1.0,true) RETURNING id", (pid,))
        partidas[codigo] = {"id": pid, "hito_id": cur.fetchone()[0]}

    # CABECERA del WBS: partida SIN fase (fase NULL). Es la forma que tiene un
    # nodo padre en producción —los crea el importador de partidas, no el POST
    # de la API, que exige fase— y es lo que distingue una HOJA de un padre:
    # `fase is not None`. Sin una cabecera sembrada, los datos del humo no tienen
    # la forma de los de producción y endpoints enteros pueden estar caídos sin
    # que nada avise (fue el caso de /ev/curva-fase). Ver check EV1.
    #
    # Lleva HH ADREDE (HH_CABECERA_WBS, un número que no sale por casualidad):
    # en producción las cabeceras traen `hh_presup > 0` —lo verificó la auditoría
    # sobre la BD real— y si la semilla las pusiera a cero, EV3 no podría
    # detectar que se cuelan en los totales: contaminar con cero no se nota.
    cur.execute(
        "INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, metrado_presup, "
        "hh_presup, otm_id) VALUES ('E2E-WBS', NULL, 'Cabecera de WBS E2E', 'und', %s, %s, "
        "'OTM-E2E') RETURNING id",
        (METRADO_CABECERA_WBS, HH_CABECERA_WBS))
    partidas["E2E-WBS"] = {"id": cur.fetchone()[0]}

    # Partidas DESECHABLES para las pruebas de programación (calendario, plazos,
    # vínculos): desde 0034 una actividad con metrado exige partida, y esas
    # pruebas necesitan una limpia —sin avances registrados— para que el
    # prorrateo salga predecible. Una por actividad, para que no se estorben.
    for i in range(1, N_PARTIDAS_PROG + 1):
        cur.execute(
            "INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, metrado_presup, "
            "hh_presup, otm_id) VALUES (%s,'F-E2E',%s,'und',%s,%s,'OTM-E2E') RETURNING id",
            (f"E2E-L{i:02d}", f"Partida de programación {i}",
             HH_PARTIDA_PROG, HH_PARTIDA_PROG))
        pid = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
            "VALUES (%s,1,'Ejecución',1.0,true)", (pid,))
        partidas[f"L{i:02d}"] = {"id": pid}
    return partidas


def fila(reporte, codigo):
    return next((p for p in reporte["partidas"] if p.get("codigo") == codigo), None)


def main():
    con = psycopg2.connect(DB)
    con.autocommit = True
    cur = con.cursor()
    c = httpx.Client(timeout=30)

    print(f"== Validación E2E tareo->ISP contra {API} ==")
    partidas = limpiar_y_sembrar(cur)
    p1, p2 = partidas["E2E-001"], partidas["E2E-002"]

    # T1 — salud
    r = c.get(f"{API}/health")
    check("T1 /health responde", r.status_code == 200 and r.json().get("status") == "ok")

    # T2 — enviar tareo del día (2 trabajadores, 3 asignaciones)
    payload = {
        "supervisor_id": "SUPE2E", "otm_id": "OTM-E2E", "fecha": FECHA_TAREO,
        "trabajadores": [
            {"trab_id": "901", "via": "e2e", "asignaciones": [
                {"partida_id": p1["id"], "hh": 5.0}, {"partida_id": p2["id"], "hh": 4.5}]},
            {"trab_id": "902", "via": "e2e", "asignaciones": [
                {"partida_id": p1["id"], "hh": 9.5}]},
        ],
    }
    r = c.post(f"{API}/api/sesion/enviar-con-partidas", json=payload)
    ok = r.status_code == 200 and r.json().get("ok") is True
    check("T2 enviar-con-partidas acepta el tareo", ok, f"status={r.status_code} body={r.text[:200]}")

    # T3 — el tareo quedó en BD con la semana correcta
    cur.execute("SELECT count(*), COALESCE(sum(hh),0), COALESCE(min(semana),0), "
                "COALESCE(max(semana),0) FROM tareo_partida WHERE otm_id='OTM-E2E'")
    n, suma, smin, smax = cur.fetchone()
    check("T3 tareo_partida: 3 filas, 19.0 HH, semana 1",
          n == 3 and float(suma) == 19.0 and smin == 1 and smax == 1,
          f"filas={n} hh={suma} semanas={smin}..{smax}")

    # T3b — /campo/tareo-dia devuelve el día partida por partida. La app de campo
    # lo usa para rellenar su espejo antes de reenviar el tareo con un parte: sin
    # esto, un parte en un teléfono que no tarearon (otro equipo, caché borrado)
    # borraría las horas de las demás partidas del día.
    r = c.get(f"{API}/campo/tareo-dia", params={
        "fecha": FECHA_TAREO, "otm_id": "OTM-E2E", "supervisor_id": "SUPE2E"})
    td = r.json() if r.status_code == 200 else []
    por_part = {}
    for x in td:
        por_part.setdefault(x["partida_id"], {})[x["trabajador_id"]] = x["hh"]
    r2 = c.get(f"{API}/campo/tareo-dia", params={
        "fecha": "no-es-fecha", "otm_id": "OTM-E2E", "supervisor_id": "SUPE2E"})
    check("T3b tareo-dia: las 3 asignaciones con sus HH por partida (y 422 si la fecha no es fecha)",
          r.status_code == 200 and len(td) == 3
          and por_part.get(p1["id"], {}).get("901") == 5.0
          and por_part.get(p1["id"], {}).get("902") == 9.5
          and por_part.get(p2["id"], {}).get("901") == 4.5
          and all(x["via"] == "e2e" for x in td)
          and r2.status_code == 422,
          f"status={r.status_code} filas={td} fecha_mala={r2.status_code}")

    # T4 — captura de avance: 50% de E2E-001 (5 de 10 und)
    r = c.post(f"{API}/ev/captura", json={
        "semana": 1, "avances": [{"hito_id": p1["hito_id"], "cantidad_acum": 5}], "hh_gastadas": []})
    check("T4 captura de avance 50% E2E-001", r.status_code == 200)

    # T5 — /ev/reporte semana 1: HH gastadas por partida + ganadas por avance
    r = c.get(f"{API}/ev/reporte", params={"semana": 1, "otm": "OTM-E2E"})
    check("T5a /ev/reporte responde", r.status_code == 200, r.text[:200])
    rep = r.json()
    f1, f2 = fila(rep, "E2E-001"), fila(rep, "E2E-002")
    check("T5b E2E-001 gastadas=14.5",
          f1 is not None and abs(f1.get("hh_gastadas_acum", -1) - 14.5) < 0.01, f"fila={f1}")
    check("T5c E2E-002 gastadas=4.5",
          f2 is not None and abs(f2.get("hh_gastadas_acum", -1) - 4.5) < 0.01, f"fila={f2}")
    check("T5d E2E-001 ganadas=50.0 (50% de 100 HH presup)",
          f1 is not None and abs(f1.get("hh_ganadas_acum", -1) - 50.0) < 0.01,
          f"claves={list(f1.keys()) if f1 else None}")
    tot = rep.get("totales", {})
    check("T5e totales.gastadas=19.0", abs(tot.get("hh_gastadas_acum", -1) - 19.0) < 0.01,
          f"totales={tot}")

    # T6 — idempotencia: reenviar el MISMO día reemplaza (no acumula)
    payload["trabajadores"] = [
        {"trab_id": "901", "via": "e2e", "asignaciones": [
            {"partida_id": p1["id"], "hh": 2.0}, {"partida_id": p2["id"], "hh": 2.0}]},
        {"trab_id": "902", "via": "e2e", "asignaciones": [
            {"partida_id": p1["id"], "hh": 9.5}]},
    ]
    r = c.post(f"{API}/api/sesion/enviar-con-partidas", json=payload)
    cur.execute("SELECT count(*), COALESCE(sum(hh),0) FROM tareo_partida WHERE otm_id='OTM-E2E'")
    n, suma = cur.fetchone()
    check("T6 reenvío del día REEMPLAZA (3 filas, 13.5 HH -- no 32.5)",
          r.status_code == 200 and n == 3 and float(suma) == 13.5, f"filas={n} hh={suma}")

    # T7 — /ev/isp incluye la semana con nuestras partidas
    r = c.get(f"{API}/ev/isp", params={"otm": "OTM-E2E"})
    ok = r.status_code == 200
    cods = [p.get("codigo") for p in r.json().get("partidas", [])] if ok else []
    check("T7 /ev/isp incluye E2E-001 y E2E-002",
          ok and "E2E-001" in cods and "E2E-002" in cods, f"status={r.status_code} codigos={cods[:5]}")

    # ── EV-WBS — los endpoints del EV con una CABECERA de WBS presente ──────
    # Una partida sin fase (nodo padre) no es un caso raro: es lo normal en
    # cualquier proyecto importado. Estos checks existen porque /ev/curva-fase
    # llevaba caído en producción con 500 —`sorted()` sobre un set que mezclaba
    # None con str— y ningún test lo veía: el humo solo sembraba hojas.
    # PRUEBA DE CONTROL: quitar el filtro `if p["fase"] is not None` de
    # isp.py::curva_fase y EV1 tiene que caer con 500.
    r = c.get(f"{API}/ev/curva-fase", params={"hasta": 3, "otm": "OTM-E2E"})
    cf = r.json() if r.status_code == 200 else {}
    check("EV1 /ev/curva-fase no revienta con una cabecera de WBS (fase NULL)",
          r.status_code == 200 and isinstance(cf.get("fases"), list)
          and all(f is not None for f in cf.get("fases", [])),
          f"status={r.status_code} fases={cf.get('fases')} body={r.text[:120]}")

    # Y el resto de la familia, con la misma cabecera viva: si alguno hereda el
    # mismo descuido, cae aquí y no en producción.
    caidos = []
    for ruta, params in (("/ev/reporte", {"semana": 3, "otm": "OTM-E2E"}),
                         ("/ev/isp", {"otm": "OTM-E2E"}),
                         ("/ev/curva", {"hasta": 3, "otm": "OTM-E2E"}),
                         ("/ev/performance", {"hasta": 3, "otm": "OTM-E2E"}),
                         ("/ev/programacion/historial-grid", {"otm": "OTM-E2E"})):
        rr = c.get(f"{API}{ruta}", params=params)
        if rr.status_code != 200:
            caidos.append(f"{ruta}->{rr.status_code}")
    check("EV2 ningun endpoint del EV cae por la cabecera de WBS",
          not caidos, f"caidos={caidos}")

    # La cabecera NO debe contaminar los totales: el detalle la incluye, pero las
    # sumas se calculan SOLO sobre hojas. Es la garantía de que filtrar por
    # `fase` sigue siendo el criterio en todo el motor.
    #
    # La 1ª versión de este check comparaba "E2E-WBS" contra `por_fase`, que
    # agrupa por FASE y no por código: la condición no podía fallar jamás, ni
    # quitando el filtro. Un check que no puede fallar da confianza sin haberla
    # ganado — exactamente lo que dejó /ev/curva-fase caído semanas. Ahora se
    # comprueban las dos cosas que sí distinguen:
    #   · si la cabecera se colara en las hojas, `_agrupar(hojas,"fase")` la
    #     metería en el grupo "SIN ASIGNAR" (`f[clave] or "SIN ASIGNAR"`);
    #   · y sus 777 HH aparecerían en `totales.hh_presup`.
    #
    # El total NO se compara contra una constante: otros checks siembran más
    # partidas en OTM-E2E (la fase CIV, por ejemplo), así que la cifra absoluta
    # depende del orden de ejecución. Se compara contra la suma de las HOJAS de
    # la propia respuesta, que es la definición de lo que el total debe ser.
    # PRUEBA DE CONTROL: cambiar `hojas` por `filas` en isp.py::reporte y EV3
    # tiene que caer por los dos motivos a la vez.
    r = c.get(f"{API}/ev/reporte", params={"semana": 3, "otm": "OTM-E2E"})
    rep = r.json() if r.status_code == 200 else {}
    detalle = rep.get("partidas", [])
    cods_rep = [p.get("codigo") for p in detalle]
    grupos_fase = [g.get("grupo") for g in rep.get("por_fase", [])]
    hh_presup_tot = (rep.get("totales") or {}).get("hh_presup")
    suma_hojas = round(sum(p.get("hh_presup") or 0
                           for p in detalle if p.get("fase") is not None), 2)
    # Sin esta línea el check pasaría también con una cabecera de 0 HH, que es
    # justo el agujero que tenía la 1ª versión: no se puede detectar
    # contaminación cuando lo que contamina vale cero.
    cabeceras_con_hh = [p.get("codigo") for p in detalle
                        if p.get("fase") is None and (p.get("hh_presup") or 0) > 0]
    check("EV3 la cabecera sale en el detalle y NO suma en los totales",
          r.status_code == 200
          and "E2E-WBS" in cods_rep                 # está en el detalle
          and "E2E-WBS" in cabeceras_con_hh         # y con HH, si no no prueba nada
          and "SIN ASIGNAR" not in grupos_fase      # no se coló como hoja sin fase
          and hh_presup_tot == suma_hojas,          # el total = solo hojas
          f"status={r.status_code} en_detalle={'E2E-WBS' in cods_rep} "
          f"cabeceras_con_hh={cabeceras_con_hh} grupos={grupos_fase[:6]} "
          f"hh_presup={hh_presup_tot} suma_hojas={suma_hojas}")

    # F-FASES — catálogo de fases (migración 0018 + CRUD)
    r = c.get(f"{API}/ev/fases", params={"proyecto_id": 1})
    fases = r.json() if r.status_code == 200 else []
    check("F1 /ev/fases con seed de disciplinas (>=11, incluye FAB)",
          r.status_code == 200 and len(fases) >= 11
          and any(f["codigo"] == "FAB" for f in fases),
          f"status={r.status_code} n={len(fases)}")

    r = c.post(f"{API}/ev/fases", json={"codigo": " e2e-99 ", "nombre": "Fase E2E"})
    fase_nueva = r.json() if r.status_code == 200 else {}
    check("F2 POST /ev/fases crea y normaliza el codigo a E2E-99",
          r.status_code == 200 and fase_nueva.get("codigo") == "E2E-99",
          f"status={r.status_code} body={r.text[:120]}")

    r = c.post(f"{API}/ev/fases", json={"codigo": "E2E-99", "nombre": "Duplicada"})
    check("F3 POST fase duplicada -> 409", r.status_code == 409, f"status={r.status_code}")

    r = c.put(f"{API}/ev/fases/{fase_nueva.get('id', 0)}", json={"activo": False})
    r2 = c.get(f"{API}/ev/fases", params={"proyecto_id": 1})
    activas = [f["codigo"] for f in r2.json()] if r2.status_code == 200 else []
    check("F4 PUT desactiva y el GET por defecto la oculta",
          r.status_code == 200 and "E2E-99" not in activas, f"activas={len(activas)}")

    # La ruta se conserva por compatibilidad, pero desde 2026-08-01 sirve la
    # plantilla .xlsx generada (con formato e instrucciones) en vez del .xls
    # estático, que era el fixture de los tests.
    r = c.get(f"{API}/ev/presupuesto/plantilla-pu")
    check("F5 la ruta antigua de la plantilla PU sirve el .xlsx nuevo",
          r.status_code == 200 and r.content[:2] == b"PK",
          f"status={r.status_code}")

    # M — matriz histórica (fechas × partidas/trabajadores)
    r = c.get(f"{API}/ev/matriz", params={"desde": FECHA_BASE, "hasta": "2026-06-07",
                                          "modo": "partidas", "otm": "OTM-E2E"})
    mz = r.json() if r.status_code == 200 else {}
    fila_p1 = next((f for f in mz.get("filas", []) if f["etiqueta"].startswith("E2E-001")), None)
    check("M1 matriz partidas: E2E-001 con 11.5 HH el martes y total de columna 13.5",
          r.status_code == 200 and fila_p1 is not None
          and abs(fila_p1["celdas"].get(FECHA_TAREO, 0) - 11.5) < 0.01
          and abs(mz.get("tot_col", {}).get(FECHA_TAREO, 0) - 13.5) < 0.01,
          f"status={r.status_code} fila={fila_p1} tot={mz.get('tot_col')}")

    r = c.get(f"{API}/ev/matriz", params={"desde": FECHA_BASE, "hasta": "2026-06-07",
                                          "modo": "trabajadores", "otm": "OTM-E2E"})
    mz = r.json() if r.status_code == 200 else {}
    check("M2 matriz trabajadores: 2 filas y semanas etiquetadas",
          r.status_code == 200 and len(mz.get("filas", [])) == 2
          and len(mz.get("semanas", [])) >= 1,
          f"filas={len(mz.get('filas', []))}")

    # F-PROG — calendario de programación + reportes de campo con fotos (0019)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
        "titulo": "E2E Hormigonado losa", "responsable": "Cuadrilla 1"})
    act = r.json() if r.status_code == 200 else {}
    check("P1 crear actividad programada", r.status_code == 200 and act.get("estado") == "PROGRAMADO",
          f"status={r.status_code} body={r.text[:150]}")

    from io import BytesIO
    from PIL import Image
    buf = BytesIO()
    Image.new("RGB", (900, 600), (180, 90, 30)).save(buf, "JPEG")
    r = c.post(f"{API}/campo/reportes",
               data={"proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
                     "supervisor_id": "SUPE2E", "descripcion": "Losa vaciada al 100%",
                     "actividad_id": act.get("id")},
               files=[("fotos", ("losa.jpg", buf.getvalue(), "image/jpeg"))])
    check("P2 reporte de campo con foto", r.status_code == 200 and r.json().get("fotos") == 1,
          f"status={r.status_code} body={r.text[:150]}")

    r = c.get(f"{API}/ev/programacion/semana", params={"proyecto_id": 1, "lunes": FECHA_TAREO})
    sem = r.json() if r.status_code == 200 else {}
    act_sem = next((a for a in sem.get("actividades", []) if a["id"] == act.get("id")), None)
    rep_sem = next((x for x in sem.get("reportes", []) if x["otm_id"] == "OTM-E2E"), None)
    check("P3 semana: actividad EJECUTADO (reporte la ejecuta) + reporte con foto",
          act_sem is not None and act_sem["estado"] == "EJECUTADO"
          and rep_sem is not None and len(rep_sem["fotos"]) == 1
          and rep_sem["fotos"][0]["url"], f"act={act_sem and act_sem.get('estado')}")

    url_foto = rep_sem["fotos"][0]["url"] if rep_sem and rep_sem["fotos"] else ""
    r = c.get(f"{API}{url_foto}")
    check("P4 URL firmada sirve el JPEG", r.status_code == 200 and r.content[:2] == b"\xff\xd8",
          f"status={r.status_code}")
    r = c.get(f"{API}{url_foto[:-4]}0000")
    check("P5 firma corrupta -> 403", r.status_code == 403, f"status={r.status_code}")

    r = c.get(f"{API}/ev/programacion/media-uso", params={"proyecto_id": 1})
    uso = r.json() if r.status_code == 200 else []
    sem_iso = next((u for u in uso if u["n_fotos"] >= 1 and u["bytes_en_disco"] > 0), None)
    check("P6 media-uso reporta bytes por semana", sem_iso is not None, f"uso={uso}")

    r = c.delete(f"{API}/ev/programacion/actividades/{act.get('id')}")
    check("P7 DELETE actividad con reporte -> 409", r.status_code == 409, f"status={r.status_code}")

    r = c.post(f"{API}/ev/programacion/purgar",
               json={"proyecto_id": 1, "semana_iso": sem_iso["semana_iso"] if sem_iso else ""})
    purga = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}{url_foto}")
    check("P8 purga libera bytes y la foto ya no se sirve",
          r.status_code == 200 and purga.get("fotos_purgadas", 0) >= 1 and r2.status_code == 404,
          f"purga={purga} get={r2.status_code}")

    # Flujo del supervisor: actividad asignada + causa de no cumplimiento (0020)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
        "titulo": "E2E Montaje faja", "supervisor_id": "SUPE2E"})
    act2 = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/campo/mis-actividades",
               params={"fecha": FECHA_TAREO, "supervisor_id": "SUPE2E"})
    mias = r2.json() if r2.status_code == 200 else []
    check("P9 actividad asignada aparece en mis-actividades del supervisor",
          r.status_code == 200 and any(a["id"] == act2.get("id") for a in mias),
          f"status={r.status_code}/{r2.status_code} n={len(mias)}")

    r = c.post(f"{API}/campo/actividades/{act2.get('id')}/no-cumplida",
               json={"supervisor_id": "SUPE2E", "causa_cat": "MATERIALES",
                     "causa": "No llegó el acero"})
    r2 = c.get(f"{API}/ev/programacion/semana", params={"proyecto_id": 1, "lunes": FECHA_TAREO})
    a2 = next((a for a in r2.json().get("actividades", []) if a["id"] == act2.get("id")), {})
    check("P10 no-cumplida registra estado, categoría CNC y detalle",
          r.status_code == 200 and a2.get("estado") == "NO_CUMPLIDA"
          and a2.get("causa_nc_cat") == "MATERIALES"
          and a2.get("causa_nc") == "No llegó el acero",
          f"status={r.status_code} act={a2.get('estado')}/{a2.get('causa_nc_cat')}")

    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-NO-EXISTE-99",
        "titulo": "E2E OTM inválida"})
    check("P11 OTM inexistente -> 400 claro (no 500 sin CORS)",
          r.status_code == 400, f"status={r.status_code}")

    # Last Planner: partida + restricciones (lookahead) + PPC/CNC (0021)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
        "titulo": "E2E Con partida y restricción", "partida_id": p1["id"],
        "supervisor_id": "SUPE2E"})
    act3 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/ev/programacion/actividades/{act3.get('id')}/restricciones",
                json={"descripcion": "Llega el acero", "tipo": "MATERIALES",
                      "responsable": "Logística"})
    rest = r2.json() if r2.status_code == 200 else {}
    r3 = c.get(f"{API}/ev/programacion/lookahead",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 3})
    la = r3.json() if r3.status_code == 200 else {}
    a3 = next((a for s in la.get("semanas", []) for a in s["actividades"]
               if a["id"] == act3.get("id")), {})
    check("P12 lookahead: actividad con partida y 1 restricción pendiente",
          a3.get("rest_pend") == 1 and a3.get("partida_codigo") == "E2E-001"
          and len(la.get("semanas", [])) == 3,
          f"act={a3.get('rest_pend')}/{a3.get('partida_codigo')} sem={len(la.get('semanas', []))}")

    r = c.put(f"{API}/ev/programacion/restricciones/{rest.get('id')}", json={"liberada": True})
    r2 = c.get(f"{API}/ev/programacion/lookahead",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    a3 = next((a for s in r2.json().get("semanas", []) for a in s["actividades"]
               if a["id"] == act3.get("id")), {})
    check("P13 liberar la restricción deja la actividad lista (0 pendientes)",
          r.status_code == 200 and r.json().get("liberada") is True
          and a3.get("rest_pend") == 0, f"rest={a3.get('rest_pend')}")

    r = c.get(f"{API}/ev/programacion/ppc", params={"proyecto_id": 1, "semanas": 26})
    d = r.json() if r.status_code == 200 else {}
    cnc_mat = next((x for x in d.get("cnc", []) if x["causa"] == "MATERIALES"), None)
    sem_ppc = next((x for x in d.get("semanal", []) if x["comprometidas"] >= 3), None)
    check("P14 PPC: Pareto registra MATERIALES y la semana calcula su PPC",
          r.status_code == 200 and cnc_mat is not None and cnc_mat["n"] >= 1
          and sem_ppc is not None and sem_ppc["ppc"] is not None,
          f"cnc={d.get('cnc')} sem={sem_ppc}")

    # Lookahead-grid con metrado diario (0022 — plantillas Anexo 01 / F030b)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "fecha_fin": "2026-06-04",
        "otm_id": "OTM-E2E", "titulo": "E2E Relleno con metrado",
        "partida_id": p2["id"], "metrado_prog": 90})
    act4 = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 2})
    grid = r2.json() if r2.status_code == 200 else {}
    ga = next((a for g in grid.get("grupos", []) for a in g["actividades"]
               if a["id"] == act4.get("id")), {})
    check("P15 lookahead-grid distribuye 90 en 3 días (30/30/30) agrupado por OTM",
          r.status_code == 200 and ga.get("metrado_prog") == 90
          and ga.get("prog", {}).get(FECHA_TAREO) == 30
          and ga.get("prog", {}).get("2026-06-04") == 30
          and len(grid.get("fechas", [])) == 14,
          f"status={r.status_code} prog={ga.get('prog')}")

    # P16 (nueva semántica 0027): replanificar un día NO cambia el META; la
    # celda queda MANUAL y el saldo se re-prorratea en los demás días.
    r = c.put(f"{API}/ev/programacion/actividades/{act4.get('id')}/metrado-dias",
              json={"dias": {"2026-06-03": 50}})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g4 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act4.get("id")), {})
    check("P16 replanificar un día: celda manual 50, META intacto (90) y saldo 40 en 20/20",
          r.status_code == 200 and r.json().get("metrado_prog") == 90
          and g4.get("prog", {}).get("2026-06-03") == 50
          and g4.get("prog", {}).get(FECHA_TAREO) == 20
          and g4.get("prog", {}).get("2026-06-04") == 20
          and "2026-06-03" in g4.get("prog_manual", []),
          f"status={r.status_code} meta={r.json().get('metrado_prog') if r.status_code == 200 else '-'} "
          f"prog={g4.get('prog')} manual={g4.get('prog_manual')}")

    r = c.post(f"{API}/ev/programacion/avance-dia",
               json={"partida_id": p2["id"], "fecha": FECHA_TAREO, "cantidad": 2.5})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    ga = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act4.get("id")), {})
    r3 = c.get(f"{API}/ev/matriz", params={"desde": FECHA_TAREO, "hasta": FECHA_TAREO,
                                           "modo": "partidas", "celda": "cantidad",
                                           "otm": "OTM-E2E"})
    fila_ev = next((f for f in r3.json().get("filas", []) if f["id"] == str(p2["id"])), {})
    cant_ev = fila_ev.get("celdas", {}).get(FECHA_TAREO)
    check("P17 avance real: aparece en el grid Y en /ev/matriz (las 2 vías, un dato)",
          ga.get("real", {}).get(FECHA_TAREO) == 2.5 and cant_ev == 2.5
          and ga.get("saldo") == 2.5,   # metrado_presup 5 - 2.5
          f"real={ga.get('real')} ev={cant_ev} saldo={ga.get('saldo')}")

    # Programación por lote: N partidas → N actividades, metrado default = presupuesto
    r = c.post(f"{API}/ev/programacion/actividades-lote", json={
        "proyecto_id": 1, "otm_id": "OTM-E2E", "responsable": "Cuadrilla E2E",
        "items": [
            {"partida_id": p1["id"], "fecha": FECHA_TAREO, "fecha_fin": "2026-06-03"},
            {"partida_id": p2["id"], "fecha": FECHA_TAREO, "metrado_prog": 4},
        ]})
    lote = r.json() if r.status_code == 200 else {}
    a_l1 = next((a for a in lote.get("actividades", []) if a["partida_id"] == p1["id"]), {})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    gl = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == a_l1.get("id")), {})
    check("P18 lote: 2 partidas -> 2 actividades; sin metrado usa el del presupuesto (10) y prorratea 5/5",
          lote.get("creadas") == 2 and a_l1.get("metrado_prog") == 10
          and gl.get("prog", {}).get(FECHA_TAREO) == 5
          and gl.get("prog", {}).get("2026-06-03") == 5
          and a_l1.get("titulo") == "Partida E2E-001",
          f"status={r.status_code} lote={lote.get('creadas')} prog={gl.get('prog')}")

    # Calendario laboral + saltos intencionales + re-prorrateo con avance (0023)
    r = c.put(f"{API}/ev/programacion/config", json={"proyecto_id": 1, "dias_semana": [1, 2, 3, 4, 5, 6]})
    r2 = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-05", "fecha_fin": "2026-06-08",
        "otm_id": "OTM-E2E", "titulo": "E2E Con calendario y salto", "partida_id": partidas["L01"]["id"],
        "metrado_prog": 90, "dias_salto": ["2026-06-06"]})
    act5 = r2.json() if r2.status_code == 200 else {}
    r3 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-05", "semanas": 2})
    g5 = next((a for g in r3.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act5.get("id")), {})
    check("P19 calendario L-S + salto del sábado: 90 cae solo en vie y lun (45/45)",
          r.status_code == 200 and g5.get("prog") == {"2026-06-05": 45, "2026-06-08": 45}
          and r3.json().get("dias_semana") == [1, 2, 3, 4, 5, 6],
          f"cfg={r.status_code} prog={g5.get('prog')}")

    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "fecha_fin": "2026-06-04",
        "otm_id": "OTM-E2E", "titulo": "E2E Linea base", "partida_id": p1["id"],
        "metrado_prog": 90})
    act6 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/ev/programacion/actividades/{act6.get('id')}/avance-dia",
                json={"fecha": FECHA_TAREO, "cantidad": 12})
    r3 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g6 = next((a for g in r3.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act6.get("id")), {})
    check("P20 avance 12 vs prog 30: el día queda congelado y el saldo 78 se re-prorratea 39/39",
          r2.status_code == 200 and g6.get("prog", {}).get(FECHA_TAREO) == 30
          and g6.get("real", {}).get(FECHA_TAREO) == 12
          and g6.get("prog", {}).get("2026-06-03") == 39
          and g6.get("prog", {}).get("2026-06-04") == 39,
          f"status={r2.status_code} prog={g6.get('prog')} real={g6.get('real')}")

    r = c.put(f"{API}/ev/programacion/actividades/{act6.get('id')}",
              json={"fecha_fin": "2026-06-05"})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g6 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act6.get("id")), {})
    check("P21 ampliar F.Fin re-prorratea el saldo en 3 días (26/26/26) sin tocar el congelado",
          r.status_code == 200 and g6.get("prog", {}).get(FECHA_TAREO) == 30
          and g6.get("prog", {}).get("2026-06-03") == 26
          and g6.get("prog", {}).get("2026-06-05") == 26,
          f"prog={g6.get('prog')}")

    # P22: registrar avance en un día intermedio NO toca los días anteriores
    # (act6: rango 02-05, prog {02:30 congelado, 03:26, 04:26, 05:26}, real 02=12)
    r = c.post(f"{API}/ev/programacion/actividades/{act6.get('id')}/avance-dia",
               json={"fecha": "2026-06-04", "cantidad": 26})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g6 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act6.get("id")), {})
    check("P22 avance en día intermedio: los anteriores quedan intactos y el saldo cae después (05=52)",
          r.status_code == 200
          and g6.get("prog", {}).get("2026-06-03") == 26      # anterior: intacto
          and g6.get("prog", {}).get("2026-06-04") == 26      # el registrado: congelado
          and g6.get("prog", {}).get("2026-06-05") == 52      # 90-12-26 en el único día posterior
          and g6.get("real", {}).get("2026-06-04") == 26,
          f"prog={g6.get('prog')} real={g6.get('real')}")

    # restaurar el calendario para no contaminar corridas parciales
    c.put(f"{API}/ev/programacion/config", json={"proyecto_id": 1, "dias_semana": [1, 2, 3, 4, 5, 6, 7]})
    for fx in (FECHA_TAREO, "2026-06-04"):
        c.post(f"{API}/ev/programacion/actividades/{act6.get('id')}/avance-dia",
               json={"fecha": fx, "cantidad": None})

    # P23 — F1 LookAhead v2: el avance por la VÍA DEL EV (/ev/avance-diario)
    # re-prorratea la actividad vinculada igual que la vía de programación,
    # y la semana se calcula con core.tiempo.semana_de (base = lunes 06-01).
    r = c.post(f"{API}/ev/avance-diario",
               json={"partida_id": p1["id"], "fecha": "2026-06-03", "cantidad_dia": 5})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g6 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act6.get("id")), {})
    cur.execute("SELECT semana FROM ev_avances_diarios WHERE partida_id=%s AND fecha='2026-06-03'",
                (p1["id"],))
    fila_sem = cur.fetchone()
    r3 = c.post(f"{API}/ev/avance-diario",
                json={"partida_id": p1["id"], "fecha": "2026-06-03", "cantidad_dia": None})
    r4 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    g6b = next((a for g in r4.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act6.get("id")), {})
    check("P23 avance vía EV re-prorratea el LookAhead (42.5/42.5), semana canónica=1 y el null borra",
          r.status_code == 200 and g6.get("real", {}).get("2026-06-03") == 5
          and g6.get("prog", {}).get("2026-06-04") == 42.5
          and g6.get("prog", {}).get("2026-06-05") == 42.5
          and fila_sem is not None and fila_sem[0] == 1
          and r3.status_code == 200 and "2026-06-03" not in g6b.get("real", {}),
          f"real={g6.get('real')} prog={g6.get('prog')} sem={fila_sem}")

    # P24 — F3 v2: causa de no cumplimiento del PLANNER, separada de la de campo;
    # en el Pareto manda la del planner (act2 pasa de MATERIALES a PROGRAMACION).
    r = c.put(f"{API}/ev/programacion/actividades/{act2.get('id')}",
              json={"causa_nc_planner_cat": "PROGRAMACION",
                    "causa_nc_planner": "Secuencia mal estimada"})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": FECHA_TAREO, "semanas": 1})
    a2g = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act2.get("id")), {})
    r3 = c.get(f"{API}/ev/programacion/ppc", params={"proyecto_id": 1, "semanas": 26})
    pareto = {x["causa"]: x["n"] for x in r3.json().get("cnc", [])}
    check("P24 causa del planner: se guarda, sale en el grid y manda en el Pareto",
          r.status_code == 200
          and a2g.get("causa_nc_planner_cat") == "PROGRAMACION"
          and a2g.get("causa_nc_planner") == "Secuencia mal estimada"
          and a2g.get("causa_nc_cat") == "MATERIALES"
          and pareto.get("PROGRAMACION", 0) >= 1,
          f"act={a2g.get('causa_nc_planner_cat')}/{a2g.get('causa_nc_cat')} pareto={pareto}")

    # P25 — F4 v2: medio día pesa 0.5 en el prorrateo (90 en L/M◐/X → 36/18/36)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-08", "fecha_fin": "2026-06-10",
        "otm_id": "OTM-E2E", "titulo": "E2E Medio dia", "partida_id": partidas["L02"]["id"], "metrado_prog": 90,
        "dias_medio": ["2026-06-09"]})
    act7 = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-08", "semanas": 1})
    g7 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act7.get("id")), {})
    check("P25 medio dia pesa 0.5: 90 en 3 días con M◐ -> 36/18/36 y el grid devuelve dias_medio",
          r.status_code == 200
          and g7.get("prog", {}).get("2026-06-08") == 36
          and g7.get("prog", {}).get("2026-06-09") == 18
          and g7.get("prog", {}).get("2026-06-10") == 36
          and g7.get("dias_medio") == ["2026-06-09"],
          f"status={r.status_code} prog={g7.get('prog')} medios={g7.get('dias_medio')}")

    # P26 — F5a v2: dependencia FS con anti-ciclo y retorno en el grid
    # act7 (08-10) será antecesora de act8 (11-12); el ciclo inverso da 409.
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-11", "fecha_fin": "2026-06-12",
        "otm_id": "OTM-E2E", "titulo": "E2E Sucesora", "partida_id": partidas["L03"]["id"], "metrado_prog": 40})
    act8 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/ev/programacion/actividades/{act8.get('id')}/dependencias",
                json={"predecesora_id": act7.get("id"), "lag_dias": 0})
    r3 = c.post(f"{API}/ev/programacion/actividades/{act7.get('id')}/dependencias",
                json={"predecesora_id": act8.get("id")})       # ciclo → 409
    r4 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-08", "semanas": 1})
    g8 = next((a for g in r4.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act8.get("id")), {})
    g7b = next((a for g in r4.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act7.get("id")), {})
    check("P26 dependencia FS: se crea, el ciclo inverso da 409 y el grid trae PRED./sucesoras",
          r2.status_code == 200 and r3.status_code == 409
          and [p["id"] for p in g8.get("predecesoras", [])] == [act7.get("id")]
          and g8.get("dep_total") == 1
          and g7b.get("sucesoras") == [act8.get("id")],
          f"dep={r2.status_code} ciclo={r3.status_code} preds={g8.get('predecesoras')}")

    # P27 — F5b v2: mover la F.Fin de la antecesora EMPUJA a la sucesora
    # act7 termina ahora el 11 (era 10) → act8 (11-12, dur 2 hábiles) debe
    # arrancar el 12 y terminar el 13, con su metrado re-prorrateado (20/20).
    r = c.put(f"{API}/ev/programacion/actividades/{act7.get('id')}",
              json={"fecha_fin": "2026-06-11"})
    movidas = r.json().get("movidas", []) if r.status_code == 200 else []
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-08", "semanas": 1})
    g8 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act8.get("id")), {})
    check("P27 cascada: ampliar la antecesora empuja la sucesora al 12-13 y re-prorratea 20/20",
          r.status_code == 200 and movidas == [act8.get("id")]
          and g8.get("fecha") == "2026-06-12" and g8.get("fecha_fin") == "2026-06-13"
          and g8.get("prog", {}).get("2026-06-12") == 20
          and g8.get("prog", {}).get("2026-06-13") == 20,
          f"movidas={movidas} rango={g8.get('fecha')}..{g8.get('fecha_fin')} prog={g8.get('prog')}")

    # P28 — la cascada nunca ADELANTA: acortar la antecesora no mueve a nadie
    r = c.put(f"{API}/ev/programacion/actividades/{act7.get('id')}",
              json={"fecha_fin": "2026-06-09"})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-08", "semanas": 1})
    g8b = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act8.get("id")), {})
    check("P28 la cascada nunca adelanta: acortar la antecesora deja a la sucesora en 12-13",
          r.status_code == 200 and r.json().get("movidas") == []
          and g8b.get("fecha") == "2026-06-12" and g8b.get("fecha_fin") == "2026-06-13",
          f"movidas={r.json().get('movidas')} rango={g8b.get('fecha')}..{g8b.get('fecha_fin')}")

    # ── Plazo y tipos de vínculo FS/SS/FF (0034, planner 2026-07-26) ──
    # Bloque autocontenido: pone el calendario en L-S y lo restaura al final.
    # Junio 2026: el 15 es lunes, el 19 viernes, el 20 sábado y el 21 domingo.
    c.put(f"{API}/ev/programacion/config", json={"proyecto_id": 1, "dias_semana": [1, 2, 3, 4, 5, 6]})

    # P54 — INICIO + PLAZO 1.5: el fin se deriva y el medio día cae al final,
    # así que el metrado se reparte 2:1 (lo que el planner pidió poder hacer).
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-15", "otm_id": "OTM-E2E",
        "titulo": "E2E Plazo dia y medio", "partida_id": partidas["L04"]["id"], "metrado_prog": 90, "plazo_dias": 1.5})
    act54 = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-15", "semanas": 2})
    g54 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act54.get("id")), {})
    check("P54 inicio+plazo 1.5: fin derivado al 16, medio día al final y 90 -> 60/30",
          r.status_code == 200 and g54.get("fecha") == "2026-06-15"
          and g54.get("fecha_fin") == "2026-06-16" and g54.get("plazo_dias") == 1.5
          and g54.get("dias_medio") == ["2026-06-16"]
          and g54.get("prog", {}).get("2026-06-15") == 60
          and g54.get("prog", {}).get("2026-06-16") == 30,
          f"status={r.status_code} rango={g54.get('fecha')}..{g54.get('fecha_fin')} "
          f"plazo={g54.get('plazo_dias')} prog={g54.get('prog')}")

    # P55 — el plazo se cuenta en días HÁBILES: 3 días desde el viernes 19
    # llegan al lunes 22 porque el domingo 21 no se trabaja.
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-19", "otm_id": "OTM-E2E",
        "titulo": "E2E Plazo salta domingo", "partida_id": partidas["L05"]["id"], "metrado_prog": 30, "plazo_dias": 3})
    act55 = r.json() if r.status_code == 200 else {}
    check("P55 el plazo cuenta días hábiles: 3 días desde el vie 19 terminan el lun 22",
          r.status_code == 200 and str(act55.get("fecha_fin")) == "2026-06-22",
          f"status={r.status_code} fin={act55.get('fecha_fin')}")

    # P56 — mover el INICIO desplaza la barra sin estirarla (modo INICIO_PLAZO)
    # y las marcas de medio día viejas no quedan colgando fuera del rango.
    r = c.put(f"{API}/ev/programacion/actividades/{act54.get('id')}",
              json={"fecha": "2026-06-17"})
    a54b = r.json() if r.status_code == 200 else {}
    check("P56 mover el inicio conserva el plazo: 17-18 y el medio día se recoloca",
          r.status_code == 200 and str(a54b.get("fecha_fin")) == "2026-06-18"
          and float(a54b.get("plazo_dias") or 0) == 1.5
          and [str(d) for d in (a54b.get("dias_medio") or [])] == ["2026-06-18"],
          f"status={r.status_code} fin={a54b.get('fecha_fin')} "
          f"plazo={a54b.get('plazo_dias')} medios={a54b.get('dias_medio')}")

    # P57 — vínculo SS con lag 1: la sucesora no arranca con la antecesora
    # sino un día hábil después, conservando SU plazo (traslape típico de obra).
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-15", "otm_id": "OTM-E2E",
        "titulo": "E2E SS sucesora", "partida_id": partidas["L06"]["id"], "metrado_prog": 40, "plazo_dias": 2})
    act57 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/ev/programacion/actividades/{act57.get('id')}/dependencias",
                json={"predecesora_id": act55.get("id"), "tipo": "SS", "lag_dias": 1})
    r3 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-15", "semanas": 2})
    g57 = next((a for g in r3.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act57.get("id")), {})
    check("P57 vínculo SS+1: arranca 1 día hábil tras el inicio de la antecesora (20-22)",
          r2.status_code == 200 and g57.get("fecha") == "2026-06-20"
          and g57.get("fecha_fin") == "2026-06-22"       # sáb 20 + lun 22 (dom no)
          and (g57.get("predecesoras") or [{}])[0].get("tipo") == "SS",
          f"dep={r2.status_code} rango={g57.get('fecha')}..{g57.get('fecha_fin')} "
          f"preds={g57.get('predecesoras')}")

    # P58 — ENCADENAR en secuencia: el gesto masivo del planner. 3 actividades
    # sueltas quedan 1→2→3 en FS con un solo POST, y la cascada las ordena.
    ids58 = []
    for i in range(3):
        r = c.post(f"{API}/ev/programacion/actividades", json={
            "proyecto_id": 1, "fecha": "2026-06-15", "otm_id": "OTM-E2E",
            "titulo": f"E2E Cadena {i + 1}", "partida_id": partidas[f"L{i + 7:02d}"]["id"],
            "metrado_prog": 10, "plazo_dias": 1})
        ids58.append(r.json().get("id") if r.status_code == 200 else None)
    r = c.post(f"{API}/ev/programacion/dependencias/encadenar",
               json={"ids": ids58, "tipo": "FS", "lag_dias": 0})
    enc = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-15", "semanas": 2})
    porid58 = {a["id"]: a for g in r2.json().get("grupos", []) for a in g["actividades"]}
    fechas58 = [porid58.get(i, {}).get("fecha") for i in ids58]
    check("P58 encadenar: un POST crea los 2 vínculos y la cadena queda 15/16/17",
          r.status_code == 200 and enc.get("vinculos") == 2 and enc.get("omitidos") == []
          and fechas58 == ["2026-06-15", "2026-06-16", "2026-06-17"],
          f"status={r.status_code} enc={enc} fechas={fechas58}")

    # P59 — encadenar es idempotente y no acepta ciclos: repetir la secuencia
    # invertida sobre la misma cadena no debe romper nada.
    r = c.post(f"{API}/ev/programacion/dependencias/encadenar",
               json={"ids": list(reversed(ids58))})
    inv = r.json() if r.status_code == 200 else {}
    check("P59 encadenar al revés no crea ciclos: los 2 pares se informan omitidos",
          r.status_code == 200 and inv.get("vinculos") == 0
          and len(inv.get("omitidos") or []) == 2,
          f"status={r.status_code} inv={inv}")

    # P60 — CAMBIAR EL TIPO DE UN VÍNCULO REPROGRAMA (bug que reportó Jean:
    # «si se le cambia el tipo no hace nada»). La causa era la regla de
    # arrastre «nunca adelanta»: al pasar de FS a SS la restricción nueva es
    # MÁS TEMPRANA y se descartaba. Editar el vínculo a propósito ahora
    # reprograma la sucesora EXACTAMENTE sobre él, también hacia atrás.
    c.put(f"{API}/ev/programacion/config", json={"proyecto_id": 1, "dias_semana": [1, 2, 3, 4, 5, 6, 7]})
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-07", "otm_id": "OTM-E2E",
        "titulo": "E2E Vinculo A", "partida_id": partidas["L10"]["id"], "metrado_prog": 100, "plazo_dias": 3})
    a60 = r.json() if r.status_code == 200 else {}
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-07", "otm_id": "OTM-E2E",
        "titulo": "E2E Vinculo B", "partida_id": partidas["L11"]["id"], "metrado_prog": 100, "plazo_dias": 2})
    b60 = r.json() if r.status_code == 200 else {}
    rangos = {}
    for tipo in ("FS", "SS", "FF"):
        rr = c.post(f"{API}/ev/programacion/actividades/{b60.get('id')}/dependencias",
                    json={"predecesora_id": a60.get("id"), "tipo": tipo, "lag_dias": 0})
        g = c.get(f"{API}/ev/programacion/lookahead-grid",
                  params={"proyecto_id": 1, "desde": "2026-09-07", "semanas": 2})
        b = next((x for gr in g.json().get("grupos", []) for x in gr["actividades"]
                  if x["id"] == b60.get("id")), {})
        rangos[tipo] = (rr.status_code, b.get("fecha"), b.get("fecha_fin"),
                        (b.get("predecesoras") or [{}])[0].get("tipo"))
    check("P60 cambiar el tipo del vínculo reprograma: FS 10-11 → SS 07-08 → FF 08-09",
          rangos["FS"] == (200, "2026-09-10", "2026-09-11", "FS")
          and rangos["SS"] == (200, "2026-09-07", "2026-09-08", "SS")
          and rangos["FF"] == (200, "2026-09-08", "2026-09-09", "FF"),
          f"rangos={rangos}")

    # P61 — pero el ARRASTRE sigue sin adelantar: acortar la antecesora no
    # trae de vuelta a la sucesora (protege el plan del planner).
    r = c.put(f"{API}/ev/programacion/actividades/{a60.get('id')}",
              json={"plazo_dias": 1})
    g = c.get(f"{API}/ev/programacion/lookahead-grid",
              params={"proyecto_id": 1, "desde": "2026-09-07", "semanas": 2})
    b61 = next((x for gr in g.json().get("grupos", []) for x in gr["actividades"]
                if x["id"] == b60.get("id")), {})
    check("P61 el arrastre no adelanta: acortar la antecesora deja a la sucesora en 08-09",
          r.status_code == 200 and b61.get("fecha") == "2026-09-08"
          and b61.get("fecha_fin") == "2026-09-09",
          f"status={r.status_code} rango={b61.get('fecha')}..{b61.get('fecha_fin')}")

    # P62 — METRADO SIN PARTIDA: el error silencioso que castigaba el PPC.
    # Con metrado exige partida (no hay dónde anotar el avance real y el PPC la
    # cuenta como no cumplida); SIN metrado es una actividad de apoyo legítima.
    r_mal = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-07", "otm_id": "OTM-E2E",
        "titulo": "E2E Metrado huerfano", "metrado_prog": 90})
    r_ok = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-07", "otm_id": "OTM-E2E",
        "titulo": "E2E Charla de seguridad"})
    check("P62 metrado sin partida da 400; la actividad de apoyo sin metrado se crea",
          r_mal.status_code == 400 and "partida" in r_mal.json().get("detail", "").lower()
          and r_ok.status_code == 200 and r_ok.json().get("partida_id") is None,
          f"mal={r_mal.status_code} ok={r_ok.status_code} detalle={r_mal.text[:120]}")

    # P63 — ADICIONAL no presupuestado: se crea su partida con HH 0 (el dato
    # llega al aprobarlo) y el grid devuelve la señal para pintarla en rojo;
    # al cargar las HH la señal desaparece.
    r = c.post(f"{API}/ev/partidas", json={
        "codigo": "E2E-ADIC", "otm_id": "OTM-E2E", "descripcion": "Adicional E2E",
        "unidad": "m3", "fase": "F-E2E", "metrado_presup": 50, "hh_presup": 0,
        "naturaleza": "ADICIONAL",
        "hitos": [{"numero": 1, "descripcion": "Ejecución", "peso": 1, "es_principal": True}]})
    pad = r.json().get("id") if r.status_code == 200 else None
    r2 = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-07", "otm_id": "OTM-E2E",
        "titulo": "E2E Trabajo adicional", "metrado_prog": 50,
        "partida_id": pad, "plazo_dias": 2})
    aad = r2.json() if r2.status_code == 200 else {}

    def _fila_adic():
        g = c.get(f"{API}/ev/programacion/lookahead-grid",
                  params={"proyecto_id": 1, "desde": "2026-09-07", "semanas": 2})
        return next((x for gr in g.json().get("grupos", []) for x in gr["actividades"]
                     if x["id"] == aad.get("id")), {})
    antes = _fila_adic()
    c.put(f"{API}/ev/partidas/{pad}", json={
        "codigo": "E2E-ADIC", "otm_id": "OTM-E2E", "descripcion": "Adicional E2E",
        "unidad": "m3", "fase": "F-E2E", "metrado_presup": 50, "hh_presup": 320,
        "naturaleza": "ADICIONAL",
        "hitos": [{"numero": 1, "descripcion": "Ejecución", "peso": 1, "es_principal": True}]})
    despues = _fila_adic()
    check("P63 adicional: nace con HH 0 marcado en rojo y deja de estarlo al cargarlas",
          r.status_code == 200 and r2.status_code == 200
          and antes.get("partida_naturaleza") == "ADICIONAL"
          and antes.get("partida_hh_presup") == 0
          and despues.get("partida_hh_presup") == 320,
          f"crear={r.status_code} act={r2.status_code} antes={antes.get('partida_hh_presup')} "
          f"nat={antes.get('partida_naturaleza')} despues={despues.get('partida_hh_presup')}")

    c.put(f"{API}/ev/programacion/config",           # se restaura el calendario
          json={"proyecto_id": 1, "dias_semana": [1, 2, 3, 4, 5, 6, 7]})

    # ── Hitos + fuente única (migración 0025, encargo 2026-07-18) ──

    # P29 — ROLLUP: cada registro diario deriva ev_avances del hito principal
    # (la ENTRADA del motor EV) = Σ diario de la semana canónica.
    r = c.post(f"{API}/ev/avance-diario",
               json={"partida_id": p1["id"], "fecha": "2026-06-04", "cantidad_dia": 2})
    cur.execute("SELECT COALESCE(SUM(cantidad_dia),0) FROM ev_avances_diarios "
                "WHERE partida_id=%s AND hito_id IS NULL AND semana=1", (p1["id"],))
    suma_diaria = float(cur.fetchone()[0])
    cur.execute("SELECT cantidad_acum FROM ev_avances WHERE hito_id=%s AND semana=1",
                (p1["hito_id"],))
    fila_acum = cur.fetchone()
    check("P29 rollup fuente única: ev_avances(sem 1) del hito principal = suma del diario",
          r.status_code == 200 and fila_acum is not None
          and abs(float(fila_acum[0]) - suma_diaria) < 0.001,
          f"acum={fila_acum} sum_diario={suma_diaria}")

    # P30 — ETAPAS: partida con 2 hitos desplegada por hitos en el lote, con
    # FS encadenado; el diario de la etapa NO principal alimenta SU hito y no
    # aparece en semana-grid (que es solo cantidad instalada = principal).
    cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                "metrado_presup, hh_presup, otm_id) VALUES "
                "('E2E-003','F-E2E','Partida etapas','und',20,80,'OTM-E2E') RETURNING id")
    p3 = cur.fetchone()[0]
    cur.execute("INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
                "VALUES (%s,1,'Habilitación',0.4,false) RETURNING id", (p3,))
    h1 = cur.fetchone()[0]
    cur.execute("INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
                "VALUES (%s,2,'Montaje',0.6,true) RETURNING id", (p3,))
    h2 = cur.fetchone()[0]
    r = c.post(f"{API}/ev/programacion/actividades-lote", json={
        "proyecto_id": 1, "otm_id": "OTM-E2E", "encadenar_hitos": True,
        "items": [
            {"partida_id": p3, "hito_id": h1, "fecha": "2026-06-15", "fecha_fin": "2026-06-16"},
            {"partida_id": p3, "hito_id": h2, "fecha": "2026-06-17", "fecha_fin": "2026-06-18"},
        ]})
    lote = r.json() if r.status_code == 200 else {}
    acts_etapa = lote.get("actividades", [])
    act_h1 = next((a for a in acts_etapa if a.get("hito_id") == h1), {})
    act_h2 = next((a for a in acts_etapa if a.get("hito_id") == h2), {})
    cur.execute("SELECT count(*) FROM prog_dependencias WHERE actividad_id=%s "
                "AND predecesora_id=%s", (act_h2.get("id"), act_h1.get("id")))
    fs_creado = cur.fetchone()[0]
    r2 = c.post(f"{API}/ev/programacion/actividades/{act_h1.get('id')}/avance-dia",
                json={"fecha": "2026-06-15", "cantidad": 4})
    cur.execute("SELECT cantidad_acum FROM ev_avances WHERE hito_id=%s", (h1,))
    acum_h1 = cur.fetchone()
    r3 = c.get(f"{API}/ev/matriz", params={"desde": "2026-06-15", "hasta": "2026-06-15",
                                           "modo": "partidas", "celda": "cantidad",
                                           "otm": "OTM-E2E"})
    fila3 = next((f for f in r3.json().get("filas", []) if f["id"] == str(p3)), {})
    cant_grid = fila3.get("celdas", {}).get("2026-06-15")
    r4 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-15", "semanas": 1})
    ge = next((a for g in r4.json().get("grupos", []) for a in g["actividades"]
               if a["id"] == act_h1.get("id")), {})
    check("P30 etapas: lote por hitos con FS auto, el diario alimenta SU hito y "
          "la matriz (cantidad instalada) solo muestra el principal",
          r.status_code == 200 and len(acts_etapa) == 2 and fs_creado == 1
          and r2.status_code == 200 and acum_h1 is not None
          and abs(float(acum_h1[0]) - 4) < 0.001
          and cant_grid is None
          and ge.get("hito_desc") == "Habilitación"
          and ge.get("real", {}).get("2026-06-15") == 4,
          f"lote={len(acts_etapa)} fs={fs_creado} acum_h1={acum_h1} "
          f"grid_cant={cant_grid} hito_desc={ge.get('hito_desc')}")

    # P31 — CHECKPOINT: etapa manual sin diario acepta pct; la etapa con
    # diario lo rechaza (409, la gobierna el rollup).
    cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                "metrado_presup, hh_presup, otm_id) VALUES "
                "('E2E-004','F-E2E','Partida checkpoint','und',10,40,'OTM-E2E') RETURNING id")
    p4 = cur.fetchone()[0]
    cur.execute("INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
                "VALUES (%s,1,'QC',0.3,false) RETURNING id", (p4,))
    h4 = cur.fetchone()[0]
    cur.execute("INSERT INTO ev_hitos (partida_id, numero, descripcion, peso, es_principal) "
                "VALUES (%s,2,'Ejecución',0.7,true)", (p4,))
    r = c.post(f"{API}/ev/programacion/hitos/{h4}/checkpoint",
               json={"fecha": FECHA_TAREO, "pct": 1})
    cur.execute("SELECT cantidad_acum FROM ev_avances WHERE hito_id=%s", (h4,))
    acum_h4 = cur.fetchone()
    r2 = c.post(f"{API}/ev/programacion/hitos/{h1}/checkpoint", json={"pct": 1})
    check("P31 checkpoint: etapa manual guarda pct*metrado y la etapa con diario da 409",
          r.status_code == 200 and acum_h4 is not None
          and abs(float(acum_h4[0]) - 10) < 0.001 and r2.status_code == 409,
          f"chk={r.status_code} acum={acum_h4} con_diario={r2.status_code}")

    # P32 — HISTORIAL-GRID: arranca en el primer registro, trae las etapas y
    # marca sin_registros la partida que aún no tiene diario ni tareo.
    r = c.get(f"{API}/ev/programacion/historial-grid", params={"otm": "OTM-E2E"})
    hg = r.json() if r.status_code == 200 else {}
    hg3 = next((p for p in hg.get("partidas", []) if p["id"] == p3), {})
    hg4 = next((p for p in hg.get("partidas", []) if p["id"] == p4), {})
    etapas3 = {e.get("hito_id") for e in hg3.get("etapas", [])}
    check("P32 historial-grid: rango desde el primer registro, etapas por hito y sin_registros",
          r.status_code == 200 and hg.get("desde") == "2026-06-01"
          and h1 in etapas3
          and hg3.get("sin_registros") is False
          and hg4.get("sin_registros") is True,
          f"desde={hg.get('desde')} etapas3={etapas3} sr4={hg4.get('sin_registros')}")

    # P33 — PERFORMANCE: la serie semanal sale del motor con los datos ya
    # registrados (sin llenar nada) y refleja HH ganadas > 0.
    r = c.get(f"{API}/ev/performance", params={"hasta": 3, "otm": "OTM-E2E"})
    pf = r.json() if r.status_code == 200 else {}
    serie = pf.get("serie", [])
    check("P33 performance: serie semanal auto-alimentada con HH ganadas y presupuesto",
          r.status_code == 200 and len(serie) >= 1
          and pf.get("hh_presup_total", 0) > 0
          and any(s.get("hh_ganadas_acum", 0) > 0 for s in serie),
          f"status={r.status_code} n={len(serie)} presup={pf.get('hh_presup_total')}")

    # P34 — PROYECTOS (2ª tanda): id PROY-#### automático con F.Fin calculada,
    # el similar (mismo nombre / monto ±100) responde 409, y forzar crea igual.
    r = c.post(f"{API}/admin/otm", json={
        "nombre": "E2E PROYECTO NUEVO", "moneda": "USD", "estado": "EJECUCION",
        "fecha_inicio": "2026-06-01", "plazo": 30, "monto_contractual": 5000})
    j1 = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT fecha_fin, moneda FROM otms WHERE id = %s", (j1.get("id", ""),))
    fila_p = cur.fetchone()
    r2 = c.post(f"{API}/admin/otm", json={
        "nombre": "E2E PROYECTO NUEVO", "monto_contractual": 5050})
    sim = (r2.json().get("detail") or {}).get("similares", []) if r2.status_code == 409 else []
    r3 = c.post(f"{API}/admin/otm", json={
        "nombre": "E2E PROYECTO NUEVO 2", "monto_contractual": 5050, "forzar": True})
    check("P34 proyectos: PROY-#### auto + F.Fin=inicio+plazo, similar da 409 y forzar crea",
          r.status_code == 200 and str(j1.get("id", "")).startswith("PROY-")
          and fila_p is not None and str(fila_p[0]) == "2026-07-01"
          and fila_p[1] == "USD"
          and r2.status_code == 409 and len(sim) >= 1
          and r3.status_code == 200,
          f"id={j1.get('id')} fin={fila_p} sim={r2.status_code}/{len(sim)} forzar={r3.status_code}")

    # P35 — REPROGRAMAR NO RESUCITA LO HECHO: act_h1 (metrado 20, real 4 el
    # 15/06) se mueve a un rango posterior → el programado nuevo reparte solo
    # el SALDO (16) y la celda del día ya avanzado queda congelada (10).
    r = c.put(f"{API}/ev/programacion/actividades/{act_h1.get('id')}",
              json={"fecha": "2026-06-22", "fecha_fin": "2026-06-23"})
    r2 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-15", "semanas": 2})
    g35 = next((a for g in r2.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act_h1.get("id")), {})
    prog35 = g35.get("prog", {})
    suma_nueva = sum(v for f, v in prog35.items() if f >= "2026-06-22")
    check("P35 reprogramar descuenta lo anotado: saldo 16 repartido y celda avanzada congelada",
          r.status_code == 200 and abs(suma_nueva - 16) < 0.001
          and prog35.get("2026-06-15") == 10
          and g35.get("saldo") == 16,
          f"put={r.status_code} prog={prog35} saldo={g35.get('saldo')}")

    # P36 — PPC AUTOMÁTICO por metrado (al cierre + SI anticipado): en una
    # semana YA CERRADA (2026-05-04), la actividad que alcanzó su programado
    # cuenta cumplida sola y la que no llegó cuenta no cumplida sola.
    cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                "metrado_presup, hh_presup, otm_id) VALUES "
                "('E2E-005','F-E2E','Partida ppc ok','und',10,40,'OTM-E2E') RETURNING id")
    p5 = cur.fetchone()[0]
    cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                "metrado_presup, hh_presup, otm_id) VALUES "
                "('E2E-006','F-E2E','Partida ppc atrasada','und',20,80,'OTM-E2E') RETURNING id")
    p6 = cur.fetchone()[0]
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-05-04", "fecha_fin": "2026-05-05",
        "otm_id": "OTM-E2E", "titulo": "E2E PPC cumplida", "partida_id": p5,
        "metrado_prog": 10})
    act_ok = r.json() if r.status_code == 200 else {}
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-05-04", "fecha_fin": "2026-05-05",
        "otm_id": "OTM-E2E", "titulo": "E2E PPC atrasada", "partida_id": p6,
        "metrado_prog": 20})
    act_no = r.json() if r.status_code == 200 else {}
    c.post(f"{API}/ev/programacion/actividades/{act_ok.get('id')}/avance-dia",
           json={"fecha": "2026-05-04", "cantidad": 10})
    c.post(f"{API}/ev/programacion/actividades/{act_no.get('id')}/avance-dia",
           json={"fecha": "2026-05-04", "cantidad": 5})
    r = c.get(f"{API}/ev/programacion/ppc", params={"proyecto_id": 1, "semanas": 26})
    s36 = next((s for s in r.json().get("semanal", []) if s["lunes"] == "2026-05-04"), {})
    check("P36 PPC automático: alcanzó → cumplida sola; semana cerrada sin llegar → no cumplida",
          r.status_code == 200 and s36.get("comprometidas") == 2
          and s36.get("cumplidas") == 1 and s36.get("no_cumplidas") == 1
          and s36.get("ppc") == 0.5,
          f"status={r.status_code} semana={s36}")

    # P37 — CELDA MANUAL PROTEGIDA: sobrevive al re-prorrateo por avance real
    # y al cambio de fechas; el saldo siempre la respeta.
    cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                "metrado_presup, hh_presup, otm_id) VALUES "
                "('E2E-007','F-E2E','Partida manual','und',30,120,'OTM-E2E') RETURNING id")
    p7 = cur.fetchone()[0]
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-06-22", "fecha_fin": "2026-06-24",
        "otm_id": "OTM-E2E", "titulo": "E2E Manual protegida", "partida_id": p7,
        "metrado_prog": 30})
    act_m = r.json() if r.status_code == 200 else {}
    c.put(f"{API}/ev/programacion/actividades/{act_m.get('id')}/metrado-dias",
          json={"dias": {"2026-06-23": 4}})
    c.post(f"{API}/ev/programacion/actividades/{act_m.get('id')}/avance-dia",
           json={"fecha": "2026-06-22", "cantidad": 6})
    r2 = c.put(f"{API}/ev/programacion/actividades/{act_m.get('id')}",
               json={"fecha_fin": "2026-06-25"})
    r3 = c.get(f"{API}/ev/programacion/lookahead-grid",
               params={"proyecto_id": 1, "desde": "2026-06-22", "semanas": 1})
    g37 = next((a for g in r3.json().get("grupos", []) for a in g["actividades"]
                if a["id"] == act_m.get("id")), {})
    check("P37 celda manual protegida: tras real y ampliación de F.Fin sigue en 4 y el saldo 20 va 10/10",
          r2.status_code == 200
          and g37.get("prog", {}).get("2026-06-23") == 4
          and g37.get("prog", {}).get("2026-06-24") == 10
          and g37.get("prog", {}).get("2026-06-25") == 10
          and "2026-06-23" in g37.get("prog_manual", []),
          f"prog={g37.get('prog')} manual={g37.get('prog_manual')}")

    # P38 — CANDADO FUENTE ÚNICA (Fase S): la captura semanal manual NO puede
    # escribir un hito gobernado por el diario en sus semanas (409); un hito
    # sin registro diario sigue aceptando captura (checkpoint/histórico).
    r = c.post(f"{API}/ev/captura", json={
        "semana": 1, "avances": [{"hito_id": p1["hito_id"], "cantidad_acum": 99}],
        "hh_gastadas": []})
    cur.execute("SELECT id FROM ev_hitos WHERE partida_id=%s AND descripcion='Montaje'", (p3,))
    h2_id = cur.fetchone()[0]
    r2 = c.post(f"{API}/ev/captura", json={
        "semana": 3, "avances": [{"hito_id": h2_id, "cantidad_acum": 7}],
        "hh_gastadas": []})
    check("P38 candado: captura manual sobre hito con diario da 409; sin diario sigue permitida",
          r.status_code == 409 and r2.status_code == 200,
          f"con_diario={r.status_code} sin_diario={r2.status_code}")

    # P39 — HISTOGRAMA MO + RATIOS (Fase S·S6, espejo del Anexo 01): el tareo
    # de la semana 1 aparece como HH/trabajadores por día y alguna partida
    # trae su ratio semanal.
    r = c.get(f"{API}/ev/programacion/histograma",
              params={"desde": "2026-06-01", "semanas": 1, "otm": "OTM-E2E"})
    hg39 = r.json() if r.status_code == 200 else {}
    dia_t = next((x for x in hg39.get("dias", []) if x["fecha"] == FECHA_TAREO), {})
    con_hh = [p for p in hg39.get("ratios", [])
              if any(s.get("hh", 0) > 0 for s in p.get("semanas", {}).values())]
    check("P39 histograma MO: HH y trabajadores del tareo por día + ratios por partida",
          r.status_code == 200 and dia_t.get("hh", 0) > 0
          and dia_t.get("trabajadores", 0) >= 1 and len(con_hh) >= 1,
          f"status={r.status_code} dia={dia_t} ratios_con_hh={len(con_hh)}")

    # P40 — IDEMPOTENCIA DEL OUTBOX (F4, 0029): el mismo reporte reenviado con
    # su id_local NO se duplica — el segundo intento devuelve el existente.
    buf40 = BytesIO()
    Image.new("RGB", (400, 300), (30, 90, 180)).save(buf40, "JPEG")
    dat40 = {"proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
             "supervisor_id": "SUPE2E", "descripcion": "E2E outbox reintento",
             "id_local": "e2e-uuid-p40"}
    r1 = c.post(f"{API}/campo/reportes", data=dat40,
                files=[("fotos", ("f.jpg", buf40.getvalue(), "image/jpeg"))])
    r2 = c.post(f"{API}/campo/reportes", data=dat40,
                files=[("fotos", ("f.jpg", buf40.getvalue(), "image/jpeg"))])
    cur.execute("SELECT COUNT(*) FROM campo_reportes WHERE id_local='e2e-uuid-p40'")
    n40 = cur.fetchone()[0]
    check("P40 outbox: reenviar el mismo reporte (id_local) no duplica",
          r1.status_code == 200 and not r1.json().get("duplicado")
          and r2.status_code == 200 and r2.json().get("duplicado") is True
          and r2.json().get("id") == r1.json().get("id") and n40 == 1,
          f"r1={r1.status_code} r2={r2.status_code} filas={n40}")

    # P41 — TTL POR ROL (F4): el login de un usuario supervisor emite token de
    # ~7 días (la app de campo offline no puede re-loguear sin señal).
    import base64 as _b64
    import json as _json
    import time as _time
    c.post(f"{API}/api/admin/usuarios",
           json={"username": "supe2e", "password": "clave-e2e",
                 "rol": "supervisor", "supervisor_id": "SUPE2E"})  # 409 si ya existe: ok
    r = c.post(f"{API}/api/auth/login",
               json={"username": "supe2e", "password": "clave-e2e"})
    tk41 = r.json().get("token", "") if r.status_code == 200 else ""
    exp41 = 0
    if tk41:
        cuerpo = tk41.split(".")[0]
        cuerpo += "=" * (-len(cuerpo) % 4)
        exp41 = _json.loads(_b64.urlsafe_b64decode(cuerpo)).get("exp", 0)
    restante41 = exp41 - _time.time()
    check("P41 login supervisor emite token de ~7 dias (outbox sin re-login)",
          r.status_code == 200 and restante41 > 6.9 * 24 * 3600,
          f"status={r.status_code} restante_h={restante41/3600:.1f}")

    # P42 — ALTA DE SUPERVISOR: entra al PADRÓN DE TRABAJADORES (regla de
    # padrón unificado) y además recibe rol + acceso con clave 1234; el login
    # trae su identidad (sup_id) para saltarse "¿Quién eres?".
    r = c.post(f"{API}/admin/supervisor", json={"nombre": "PACHECO ROJAS MARIO E2E"})
    alta = r.json() if r.status_code == 200 else {}
    rl = c.post(f"{API}/api/auth/login",
                json={"username": alta.get("usuario", "x"), "password": "1234"})
    lg = rl.json() if rl.status_code == 200 else {}
    cur.execute("SELECT id FROM trabajadores WHERE nombre = 'PACHECO ROJAS MARIO E2E'")
    ficha = cur.fetchone()
    check("P42 alta de supervisor: ficha en trabajadores + rol + acceso (clave 1234)",
          r.status_code == 200 and alta.get("usuario") == "mpacheco"
          and alta.get("password") == "1234" and ficha is not None
          and alta.get("trabajador_id") == ficha[0] and rl.status_code == 200
          and lg.get("rol") == "supervisor" and lg.get("supervisor_id") == alta.get("id"),
          f"alta={alta} login={rl.status_code} ficha={ficha}")

    # P42b — IMPORT DE PERSONAL: una sola puerta (/admin/trabajador). Con
    # es_supervisor=true la persona queda en el padrón Y con acceso; repetir la
    # importación REUTILIZA su perfil (no duplica ni cambia su contraseña).
    body = {"nombre": "TICONA HUANCA LUIS E2E", "cargo": "SUPERVISOR DE CAMPO",
            "dni": "77777777", "tipo": "INDIRECTO", "es_supervisor": True}
    r = c.post(f"{API}/admin/trabajador", json=body)
    imp1 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/admin/trabajador", json=body)
    imp2 = r2.json() if r2.status_code == 200 else {}
    cur.execute("SELECT COUNT(*) FROM trabajadores WHERE nombre = 'TICONA HUANCA LUIS E2E'")
    n_fichas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE supervisor_id = %s", (imp1.get("supervisor_id"),))
    n_users = cur.fetchone()[0]
    check("P42b import: todos al padrón; reimportar reutiliza el perfil sin duplicar",
          r.status_code == 200 and imp1.get("nuevo") is True and imp1.get("usuario") == "lticona"
          and r2.status_code == 200 and imp2.get("nuevo") is False
          and imp2.get("id") == imp1.get("id") and imp2.get("usuario") is None
          and n_fichas == 1 and n_users == 1,
          f"1ra={imp1} 2da={imp2} fichas={n_fichas} usuarios={n_users}")

    # P43 — PROMOVER A UN TRABAJADOR: elegirlo desde el panel Usuarios lo
    # registra como supervisor (ligado a su ficha) y le crea el acceso; el
    # segundo intento avisa que ya lo tiene.
    r = c.post(f"{API}/api/admin/usuarios/desde-personal",
               json={"origen": "TRABAJADOR", "id": "901", "username": "e2etrab"})
    pro = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/api/admin/usuarios/desde-personal",
                json={"origen": "TRABAJADOR", "id": "901"})
    cur.execute("SELECT id FROM supervisores WHERE trabajador_id = '901'")
    sup_pro = cur.fetchone()
    rl = c.post(f"{API}/api/auth/login", json={"username": "e2etrab", "password": "1234"})
    check("P43 promover trabajador a supervisor: crea padrón + acceso; repetir da 409",
          r.status_code == 200 and pro.get("promovido") is True
          and pro.get("username") == "e2etrab" and sup_pro is not None
          and pro.get("supervisor_id") == sup_pro[0] and r2.status_code == 409
          and rl.status_code == 200 and rl.json().get("supervisor_id") == sup_pro[0],
          f"status={r.status_code} pro={pro} repetido={r2.status_code} sup={sup_pro}")

    # P44 — PERSONAL ELEGIBLE: el promovido sale como SUPERVISOR con su
    # usuario y ya no como TRABAJADOR; el resto trae username sugerido.
    r = c.get(f"{API}/api/admin/personal-elegible")
    pers = r.json() if r.status_code == 200 else []
    p901 = [p for p in pers if p["origen"] == "TRABAJADOR" and p["id"] == "901"]
    p902 = next((p for p in pers if p["origen"] == "TRABAJADOR" and p["id"] == "902"), None)
    psup = next((p for p in pers if p["origen"] == "SUPERVISOR"
                 and p.get("username") == "e2etrab"), None)
    check("P44 personal elegible: promovido ya no duplica y los demás traen usuario sugerido",
          r.status_code == 200 and not p901 and psup is not None
          and p902 is not None and bool(p902.get("username_sugerido"))
          and p902.get("username") is None,
          f"status={r.status_code} dup901={len(p901)} sup={psup} t902={p902}")

    # P45 — SINCRONIZAR: crea de golpe los accesos faltantes y es idempotente.
    r = c.post(f"{API}/api/admin/usuarios/sincronizar-supervisores")
    s1 = r.json() if r.status_code == 200 else {}
    r2 = c.post(f"{API}/api/admin/usuarios/sincronizar-supervisores")
    s2 = r2.json() if r2.status_code == 200 else {}
    check("P45 sincronizar supervisores: crea los faltantes y no duplica al repetir",
          r.status_code == 200 and s1.get("clave_inicial") == "1234"
          and r2.status_code == 200 and len(s2.get("creados", [])) == 0
          and s2.get("ya_tenian", 0) >= 1,
          f"1ra={len(s1.get('creados', []))} creados / 2da={s2}")

    # P46 — INVARIANTE DEL PADRÓN UNIFICADO (0031): no puede quedar ningún
    # supervisor activo sin su ficha en trabajadores (antes el import los
    # creaba solo en `supervisores` y no aparecían en el padrón).
    cur.execute("SELECT COUNT(*) FROM supervisores WHERE activo AND trabajador_id IS NULL")
    sueltos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM supervisores s JOIN trabajadores t ON t.id = s.trabajador_id "
                "WHERE s.activo AND NOT t.activo")
    inactivos = cur.fetchone()[0]
    check("P46 padrón unificado: todo supervisor activo tiene ficha de trabajador activa",
          sueltos == 0 and inactivos == 0,
          f"sin_ficha={sueltos} con_ficha_inactiva={inactivos}")

    # P47 — REPORTE ESTRUCTURADO (0032): el supervisor manda viñetas de lo
    # hecho + restricciones con categoría CNC + área/turno; la actividad queda
    # EJECUTADA y el parte del día sale armado para copiar al grupo.
    buf47 = BytesIO()
    Image.new("RGB", (500, 400), (60, 120, 60)).save(buf47, "JPEG")
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
        "titulo": "E2E Cerco perimetral", "supervisor_id": "SUPE2E"})
    act47 = r.json() if r.status_code == 200 else {}
    r = c.post(f"{API}/campo/reportes",
               data={"proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
                     "supervisor_id": "SUPE2E", "actividad_id": act47.get("id"),
                     # el área NO se manda: se hereda de la OTM. El frente sí,
                     # y se normaliza a MAYÚSCULAS sin espacios repetidos.
                     "frente": "  bahia  4 ", "turno": "DIA",
                     "anotaciones": json.dumps(["Corte de esparragos",
                                                "Instalacion del cerco completo"]),
                     "restricciones": json.dumps([
                         {"cat": "EQUIPOS", "detalle": "No hubo camion grua"}])},
               files=[("fotos", ("cerco.jpg", buf47.getvalue(), "image/jpeg"))])
    rep47 = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT area, turno, anotaciones, restricciones, descripcion, frente "
                "FROM campo_reportes WHERE id = %s", (rep47.get("id"),))
    fila47 = cur.fetchone()
    check("P47 reporte estructurado: área heredada, frente normalizado, viñetas y CNC",
          r.status_code == 200 and fila47 is not None
          and fila47[0] == "PLANTA SX / EW" and fila47[1] == "DIA"
          and len(fila47[2]) == 2 and fila47[3][0]["cat"] == "EQUIPOS"
          and "• Corte de esparragos" in (fila47[4] or "")
          and fila47[5] == "BAHIA 4",
          f"status={r.status_code} fila={fila47}")

    # P47b — el catálogo de frentes se autoalimenta con lo ya escrito.
    r = c.get(f"{API}/campo/frentes", params={"otm_id": "OTM-E2E"})
    frentes = r.json() if r.status_code == 200 else []
    check("P47b catálogo de frentes: se autoalimenta con lo que ya se usó",
          r.status_code == 200 and "BAHIA 4" in frentes,
          f"status={r.status_code} frentes={frentes}")

    # P48 — PARTE DIARIO listo para WhatsApp (mismo formato que la app).
    r = c.get(f"{API}/ev/programacion/reporte-dia",
              params={"fecha": FECHA_TAREO, "supervisor_id": "SUPE2E"})
    partes = r.json().get("partes", []) if r.status_code == 200 else []
    txt = partes[0]["texto"] if partes else ""
    check("P48 parte diario: cabecera, personal del tareo, área, frente y restricciones",
          r.status_code == 200 and "Turno: DIA" in txt
          and "CANTIDAD TOTAL PERSONAL:" in txt and "AREA: PLANTA SX / EW" in txt
          and "FRENTE: BAHIA 4" in txt
          and "* Corte de esparragos" in txt
          and "RESTRICCIONES." in txt and "camion grua" in txt,
          f"status={r.status_code} texto={txt[:200]!r}")

    # P49 — PLANTILLA: al reportar otra vez la misma partida/hito, la app
    # ofrece lo escrito la vez anterior para no volver a tipearlo.
    cur.execute("SELECT partida_id, hito_id FROM prog_actividades WHERE id = %s",
                (act47.get("id"),))
    ph = cur.fetchone()
    cur.execute("UPDATE prog_actividades SET partida_id = %s WHERE id = %s",
                (p1["id"], act47.get("id")))
    r0 = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": FECHA_TAREO, "otm_id": "OTM-E2E",
        "titulo": "E2E Cerco perimetral dia 2", "supervisor_id": "SUPE2E",
        "partida_id": p1["id"]})
    act49 = r0.json() if r0.status_code == 200 else {}
    r = c.get(f"{API}/campo/reporte-plantilla", params={"actividad_id": act49.get("id")})
    pl = r.json() if r.status_code == 200 else {}
    check("P49 plantilla: el reporte anterior de esa partida se ofrece para reusar",
          r.status_code == 200 and pl.get("frente") == "BAHIA 4"
          and "Corte de esparragos" in (pl.get("anotaciones") or [])
          and (pl.get("restricciones") or [{}])[0].get("cat") == "EQUIPOS",
          f"status={r.status_code} plantilla={pl} ph={ph}")

    # P49b — La misma ayuda para una partida que el planner NO programó: cuando
    # el supervisor abre el parte todavía no hay actividad de la que colgarse,
    # así que se pregunta por la partida. Sin argumentos, 422.
    r = c.get(f"{API}/campo/reporte-plantilla", params={"partida_id": p1["id"]})
    plp = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/campo/reporte-plantilla")
    check("P49b plantilla por partida: sirve sin actividad previa (y 422 sin argumentos)",
          r.status_code == 200 and plp.get("frente") == "BAHIA 4"
          and "Corte de esparragos" in (plp.get("anotaciones") or [])
          and r2.status_code == 422,
          f"status={r.status_code} plantilla={plp} sin_args={r2.status_code}")

    # P50 — RESTRICCIONES EN EL PPC: las que reportó el supervisor (con la
    # actividad SÍ ejecutada) salen por actividad para la tabla F030b y en su
    # propio Pareto, SIN mezclarse con las causas de no cumplimiento.
    r = c.get(f"{API}/ev/programacion/ppc",
              params={"proyecto_id": 1, "desde": FECHA_BASE, "hasta": "2026-06-30"})
    ppc50 = r.json() if r.status_code == 200 else {}
    rest_act = (ppc50.get("restricciones") or {}).get(str(act47.get("id")), [])
    par_rest = ppc50.get("pareto_restricciones") or []
    cnc_cats = [x["causa"] for x in (ppc50.get("cnc") or [])]
    check("P50 PPC: restricciones por actividad y Pareto propio (separado del CNC)",
          r.status_code == 200 and len(rest_act) == 1
          and rest_act[0]["cat"] == "EQUIPOS" and "camion grua" in rest_act[0]["detalle"]
          and any(x["causa"] == "EQUIPOS" and x["n"] >= 1 for x in par_rest)
          and "EQUIPOS" not in cnc_cats,
          f"status={r.status_code} act={rest_act} pareto={par_rest} cnc={cnc_cats}")

    # P51 — REPORTE POR PARTIDA (sustento de valorización): cabecera con las
    # cifras de la partida + partes en orden cronológico con sus fotos.
    r = c.get(f"{API}/ev/programacion/reporte-partida", params={"partidas": p1["id"]})
    rp = r.json() if r.status_code == 200 else {}
    bloque = (rp.get("partidas") or [{}])[0]
    part = bloque.get("partida", {})
    reps51 = bloque.get("reportes", [])
    fechas51 = [x["fecha"] for x in reps51]
    txt51 = reps51[0].get("texto") or "" if reps51 else ""
    fotos51 = [f for x in reps51 for f in x.get("fotos", [])]
    check("P51 reporte por partida: cifras de la partida + partes cronológicos con fotos",
          r.status_code == 200 and part.get("codigo") == "E2E-001"
          and part.get("metrado_presup") == 10 and part.get("hh_gastadas", 0) > 0
          and len(reps51) >= 1 and fechas51 == sorted(fechas51)
          and "CANTIDAD TOTAL PERSONAL" in txt51
          and len(fotos51) > 0
          # La galería del PDF arma las filas de altura uniforme con la forma de
          # cada foto: ancho/alto deben viajar (Pillow los guardó al subirla).
          and all(f.get("ancho") and f.get("alto") for f in fotos51),
          f"status={r.status_code} partida={part} n_fotos={len(fotos51)} "
          f"dims={[(f.get('ancho'), f.get('alto')) for f in fotos51]}")

    # P53 — COHERENCIA DEL SUSTENTO: la cuadrilla del parte sale del tareo de
    # ESA partida (no del día entero del supervisor), cada parte trae sus HH,
    # las actividades realizadas aparecen y las restricciones NO (el sustento
    # acredita lo ejecutado; las restricciones viven en el PPC).
    per51 = [x for x in reps51 if x.get("hh_dia", 0) > 0]
    check("P53 sustento coherente: cuadrilla y HH de la partida, sin restricciones",
          "CANTIDAD TOTAL PERSONAL: 0" not in txt51
          and len(per51) >= 1 and part.get("sin_tareo") is False
          and part.get("hh_rango", 0) > 0
          and "* Corte de esparragos" in txt51
          and "RESTRICCIONES." not in txt51 and "camion grua" not in txt51,
          f"hh_dia={[x.get('hh_dia') for x in reps51]} sin_tareo={part.get('sin_tareo')} "
          f"texto={txt51[:220]!r}")

    # P52 — el rango de fechas filtra (vacío = todo el historial)
    r = c.get(f"{API}/ev/programacion/reporte-partida",
              params={"partidas": p1["id"], "desde": "2027-01-01", "hasta": "2027-12-31"})
    vacio = (r.json().get("partidas") or [{}])[0] if r.status_code == 200 else {}
    check("P52 reporte por partida: el rango de fechas filtra los partes",
          r.status_code == 200 and len(vacio.get("reportes", [])) == 0
          and vacio.get("partida", {}).get("codigo") == "E2E-001",
          f"status={r.status_code} reportes={len(vacio.get('reportes', []))}")

    # P64 — DOS TRAMOS de la misma partida: el avance de uno no puede
    # atribuirse al otro (0035). Antes las dos filas mostraban el mismo real,
    # el segundo tramo se quedaba sin plan al re-prorratearse y en el PPC las
    # dos se daban por cumplidas con el trabajo de una sola.
    r = c.post(f"{API}/ev/partidas", json={
        "codigo": "E2E-TRAMOS", "otm_id": "OTM-E2E", "descripcion": "Partida en dos tramos",
        "unidad": "m2", "fase": "F-E2E", "metrado_presup": 100, "hh_presup": 200,
        "hitos": [{"numero": 1, "descripcion": "Ejecución", "peso": 1, "es_principal": True}]})
    ptr = r.json().get("id") if r.status_code == 200 else None
    tr = []
    for tit, ini, fin in (("E2E Tramo A", "2026-09-14", "2026-09-16"),
                          ("E2E Tramo B", "2026-09-17", "2026-09-19")):
        rr = c.post(f"{API}/ev/programacion/actividades", json={
            "proyecto_id": 1, "fecha": ini, "fecha_fin": fin, "otm_id": "OTM-E2E",
            "titulo": tit, "partida_id": ptr, "metrado_prog": 50})
        tr.append(rr.json().get("id") if rr.status_code == 200 else None)
    c.post(f"{API}/ev/programacion/actividades/{tr[0]}/avance-dia",
           json={"fecha": "2026-09-15", "cantidad": 50})

    def _tramos():
        g = c.get(f"{API}/ev/programacion/lookahead-grid",
                  params={"proyecto_id": 1, "desde": "2026-09-14", "semanas": 1})
        f = {x["id"]: x for gr in g.json().get("grupos", []) for x in gr["actividades"]}
        return f.get(tr[0], {}), f.get(tr[1], {})
    ta, tb = _tramos()
    # Tocar B fuerza su re-prorrateo: es donde antes perdía todo su plan.
    c.put(f"{API}/ev/programacion/actividades/{tr[1]}",
          json={"fecha": "2026-09-17", "fecha_fin": "2026-09-19"})
    _ta2, tb2 = _tramos()
    check("P64 dos tramos de una partida: el real es de su tramo y el otro conserva su plan",
          "2026-09-15" in ta.get("real", {}) and "2026-09-15" not in tb.get("real", {})
          and ta.get("tramos") == 2
          and abs(sum(v for v in tb2.get("prog", {}).values()) - 50) < 0.01,
          f"A.real={ta.get('real')} B.real={tb.get('real')} B.prog={tb2.get('prog')}")

    # P65 — WBS de las partidas creadas fuera del importador (bandeja «por
    # ubicar»): nacen colgadas de su padre y una sin OTM se puede ubicar
    # después, con 409 explicado si el código ya existe en la OTM destino.
    r = c.post(f"{API}/ev/partidas", json={
        "codigo": "E2E-001.09", "otm_id": "OTM-E2E", "descripcion": "Hija de E2E-001",
        "unidad": "m", "fase": "F-E2E", "metrado_presup": 5, "hh_presup": 5,
        "hitos": [{"numero": 1, "descripcion": "Ejecución", "peso": 1, "es_principal": True}]})
    hija = next((x for x in c.get(f"{API}/ev/partidas?otm=OTM-E2E").json()
                 if x["codigo"] == "E2E-001.09"), {})
    r2 = c.post(f"{API}/ev/partidas", json={
        "codigo": "E2E-SINOTM", "descripcion": "Sin OTM todavia", "unidad": "glb",
        "fase": "F-E2E", "metrado_presup": 3, "hh_presup": 0,
        "hitos": [{"numero": 1, "descripcion": "Ejecución", "peso": 1, "es_principal": True}]})
    suelta = r2.json().get("id") if r2.status_code == 200 else None
    band = c.get(f"{API}/ev/partidas-por-ubicar").json()
    en_bandeja = next((x for x in band if x["id"] == suelta), {})
    r3 = c.put(f"{API}/ev/partidas/{suelta}/ubicar",
               json={"otm_id": "OTM-E2E", "parent_codigo": "E2E-001"})
    r4 = c.put(f"{API}/ev/partidas/{suelta}/ubicar",
               json={"otm_id": "OTM-E2E", "codigo": "E2E-001"})
    check("P65 partida nueva cuelga del WBS; la de la bandeja se ubica y el código repetido da 409",
          r.status_code == 200 and hija.get("parent_codigo") == "E2E-001"
          and hija.get("nivel") == 2
          and "SIN_OTM" in en_bandeja.get("motivos", []) and r3.status_code == 200
          and r3.json().get("parent_codigo") == "E2E-001" and r4.status_code == 409,
          f"hija={hija.get('nivel')}/{hija.get('parent_codigo')} bandeja={en_bandeja.get('motivos')} "
          f"ubicar={r3.status_code} choque={r4.status_code}")

    # P66 — correlativo sugerido: el que sigue entre los hijos del padre y la
    # serie propia de los adicionales (el planner ya no inventa códigos).
    def _sig(**kw):
        return c.get(f"{API}/ev/partidas/siguiente-codigo",
                     params={"otm": "OTM-E2E", **kw}).json().get("codigo")
    # E2E-001.09 la creó P65 -> el siguiente hijo es el .10
    sig_hijo = _sig(parent_codigo="E2E-001")
    sig_adic = _sig(naturaleza="ADICIONAL")
    check("P66 correlativo sugerido: hijo que sigue del padre y serie propia del adicional",
          sig_hijo == "E2E-001.10" and str(sig_adic or "").startswith("ADIC-"),
          f"hijo={sig_hijo} adicional={sig_adic}")

    # P67 — datos maestros: desactivar una partida CON trabajo colgado avisa qué
    # tiene (409) en vez de hacerlo callado; confirmando sí se desactiva.
    r = c.delete(f"{API}/ev/partidas/{p1['id']}")
    uso = c.get(f"{API}/ev/partidas/{p1['id']}/uso").json()
    check("P67 desactivar una partida con trabajo registrado avisa antes (409) y detalla el uso",
          r.status_code == 409 and "registrado" in r.text.lower()
          and (uso.get("actividades", 0) or uso.get("dias_avance", 0)) > 0,
          f"status={r.status_code} uso={uso} detalle={r.text[:120]}")

    # P68 — un documento de costo ya cargado se puede CORREGIR (antes solo
    # borrar y rehacer el Excel entero).
    r = c.post(f"{API}/ev/ro/documentos", json={
        "proyecto_id": 1, "tipo_doc": "FACTURA", "proveedor": "E2E PROVEEDOR",
        "numero_doc": "E2E-DOC-1", "fecha": "2026-06-10", "tipo_recurso": "MAT",
        "directo": True, "fase": "F-E2E", "moneda": "PEN", "monto": 1000})
    did = r.json().get("id") if r.status_code == 200 else None
    r2 = c.put(f"{API}/ev/ro/documentos/{did}", json={
        "proyecto_id": 1, "tipo_doc": "FACTURA", "proveedor": "E2E CORREGIDO",
        "numero_doc": "E2E-DOC-1", "fecha": "2026-06-10", "tipo_recurso": "EQP",
        "directo": False, "fase": "F-E2E", "moneda": "PEN", "monto": 2500})
    doc = next((d for d in c.get(f"{API}/ev/ro/costos?proyecto_id=1").json().get("documentos", [])
                if d["id"] == did), {})
    r3 = c.put(f"{API}/ev/ro/documentos/{did}", json={
        "proyecto_id": 1, "tipo_doc": "FACTURA", "numero_doc": "E2E-DOC-1",
        "fecha": "2026-06-10", "tipo_recurso": "MO", "directo": True,
        "moneda": "PEN", "monto": 10})
    check("P68 editar un documento de costo cargado; la MO sigue vetada como factura",
          r2.status_code == 200 and doc.get("proveedor") == "E2E CORREGIDO"
          and float(doc.get("monto") or 0) == 2500 and doc.get("tipo_recurso") == "EQP"
          and r3.status_code == 400,
          f"put={r2.status_code} doc={ {k: doc.get(k) for k in ('proveedor', 'monto', 'tipo_recurso')} } mo={r3.status_code}")

    # P69 — histograma de personal por día / semana / mes. La agrupación por
    # mes es la que faltaba: el histograma de MO del Anexo 01 solo veía días.
    def _hist(agrupar):
        return c.get(f"{API}/api/histograma-personal",
                     params={"agrupar": agrupar, "desde": FECHA_TAREO, "hasta": FECHA_TAREO})
    hd, hm = _hist("dia"), _hist("mes")
    pd_ = (hd.json().get("periodos") or [{}])[0]
    pm_ = (hm.json().get("periodos") or [{}])[0]
    hmal = c.get(f"{API}/api/histograma-personal", params={"agrupar": "trimestre"})
    check("P69 histograma de personal: agrupa por día y por mes, y valida la agrupación",
          hd.status_code == 200 and hm.status_code == 200
          and pd_.get("periodo") == FECHA_TAREO
          and pm_.get("periodo") == FECHA_TAREO[:8] + "01"      # 1º del mes
          and pd_.get("trabajadores", 0) > 0
          and pd_.get("pico") == pd_.get("trabajadores")        # en un día coinciden
          and bool(pd_.get("por_cargo"))
          and hmal.status_code == 422,
          f"dia={pd_} mes={pm_} invalida={hmal.status_code}")

    # ── P70-P72 · Plantillas Excel ────────────────────────────
    # Lo que se comprueba no es el aspecto (eso son los tests unitarios) sino el
    # CONTRATO: que lo que la plantilla produce lo acepte su propio importador.
    # Hasta ahora nada lo garantizaba, y una cabecera renombrada no daba error
    # en ningún sitio: el import simplemente no encontraba filas válidas.
    import io as _io
    import openpyxl as _op

    esperadas = {
        "personal":    ["NOMBRE", "CARGO", "DNI", "TIPO", "ES_SUPERVISOR"],
        "proyectos":   ["NOMBRE", "AREA", "ESTADO", "CENTRO_COSTO", "MONEDA", "PLAZO",
                        "FECHA DE INICIO", "MONTO CONTRACTUAL", "MONTO VALORIZADO", "ID"],
        "presupuesto": ["CODIGO", "DESCRIPCION", "UNIDAD", "FASE", "SUB_FASE",
                        "METRADO", "PRECIO_UNITARIO", "HH_META"],
        "costos_ro":   ["FASE", "TIPO_RECURSO", "DIRECTO", "PERIODO", "MONTO",
                        "FUENTE", "NOTA"],
        "costos":      ["PROVEEDOR", "NUMERO_DOC", "FECHA", "TIPO_DOC", "TIPO_RECURSO",
                        "DIRECTO", "FASE", "MONEDA", "MONTO", "GLOSA"],
    }

    def _cabeceras(clave: str) -> list:
        r = c.get(f"{API}/ev/plantillas/{clave}")
        if r.status_code != 200 or r.content[:2] != b"PK":
            return []
        ws = _op.load_workbook(_io.BytesIO(r.content)).worksheets[0]
        for fila in range(1, 40):
            if ws.cell(fila, 2).value == esperadas[clave][0]:
                return [ws.cell(fila, 2 + i).value for i in range(len(esperadas[clave]))]
        return []

    malas = {k: _cabeceras(k) for k in esperadas if _cabeceras(k) != esperadas[k]}
    check("P70 las plantillas traen las cabeceras EXACTAS que espera cada importador",
          not malas, f"no coinciden: {malas}")

    cat = c.get(f"{API}/ev/plantillas")
    claves = {p["clave"] for p in cat.json()} if cat.status_code == 200 else set()
    check("P71 el catálogo de plantillas describe las 7",
          cat.status_code == 200 and claves == {"personal", "partidas", "proyectos",
                                                "presupuesto", "costos", "costos_ro", "pu"},
          f"status={cat.status_code} claves={claves}")

    # Ida y vuelta de verdad: la plantilla de costos se parsea en el SERVIDOR,
    # así que se puede descargar y volver a subir tal cual. Con solo la fila de
    # ejemplo dentro, el import debe leerla sin errores.
    rp = c.get(f"{API}/ev/plantillas/costos")
    rimp = c.post(f"{API}/ev/ro/costos/importar",
                  params={"proyecto_id": 1, "confirmar": False},
                  files={"file": ("plantilla.xlsx", rp.content,
                                  "application/vnd.openxmlformats-officedocument."
                                  "spreadsheetml.sheet")})
    res = rimp.json().get("resumen", {}) if rimp.status_code == 200 else {}
    check("P72 ida y vuelta: el importador de costos lee la plantilla que él mismo reparte",
          rimp.status_code == 200 and res.get("filas", 0) >= 2 and not res.get("errores"),
          f"status={rimp.status_code} resumen={res}")

    # ── P73-P77 · D1/D2: el compromiso congelado (migración 0041) ──────────
    # El defecto: `_redistribuir` borra las celdas de prog_metrado_dia sin filtro
    # de fecha, así que correr la F.Inicio de una actividad que NO se hizo la
    # dejaba con comprometido 0 y la sacaba del PPC. «Si no cumpliste, muévela.»
    #
    # Se reusa la semana de P36 (2026-05-04, ya cerrada en el calendario, PPC
    # 0.5: una cumplida y una atrasada). Es el escenario exacto del defecto.
    LUN_D1 = "2026-05-04"

    def _ppc_semana(lunes=LUN_D1):
        rr = c.get(f"{API}/ev/programacion/ppc", params={"proyecto_id": 1, "semanas": 26})
        if rr.status_code != 200:
            return {}
        return next((s for s in rr.json().get("semanal", []) if s["lunes"] == lunes), {})

    def _nueva_partida(codigo, desc, metrado, hh):
        cur.execute("INSERT INTO ev_partidas (codigo, fase, descripcion, unidad, "
                    "metrado_presup, hh_presup, otm_id) VALUES "
                    f"('{codigo}','F-E2E','{desc}','und',{metrado},{hh},'OTM-E2E') "
                    "RETURNING id")
        return cur.fetchone()[0]

    # D1A — comprometida y NUNCA tocada: es la que el defecto hacía desaparecer.
    # Sin avance registrado no hay ninguna celda «intacta», así que al mover la
    # fecha el DELETE de `_redistribuir` se lleva TODAS sus celdas de mayo.
    p_d1a = _nueva_partida("E2E-D1A", "Partida D1 sin avance", 15, 60)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": LUN_D1, "fecha_fin": "2026-05-05",
        "otm_id": "OTM-E2E", "titulo": "E2E D1 sin avance",
        "partida_id": p_d1a, "metrado_prog": 15})
    act_d1a = r.json() if r.status_code == 200 else {}

    # D1B — comprometida con 25 y solo 5 hechos. El metrado congelado es lo
    # único que impide darla por cumplida bajándole el compromiso.
    p_d1b = _nueva_partida("E2E-D1B", "Partida D1 avance parcial", 25, 100)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": LUN_D1, "fecha_fin": "2026-05-08",
        "otm_id": "OTM-E2E", "titulo": "E2E D1 avance parcial",
        "partida_id": p_d1b, "metrado_prog": 25})
    act_d1b = r.json() if r.status_code == 200 else {}
    c.post(f"{API}/ev/programacion/actividades/{act_d1b.get('id')}/avance-dia",
           json={"fecha": LUN_D1, "cantidad": 5})

    base = _ppc_semana()
    r = c.post(f"{API}/ev/programacion/plan-semana",
               json={"proyecto_id": 1, "lunes": LUN_D1,
                     "comprometido_por": "e2e", "nota": "compromiso E2E"})
    comp = r.json() if r.status_code == 200 else {}
    tras_comp = _ppc_semana()
    check("P73 comprometer la semana congela el metrado y no cambia el PPC",
          r.status_code == 200 and comp.get("comprometidas", 0) >= 4
          and comp.get("metrado_comprometido", 0) > 0
          and tras_comp.get("ppc") == base.get("ppc")
          and tras_comp.get("comprometidas") == base.get("comprometidas")
          and tras_comp.get("origen") == "COMPROMETIDO",
          f"status={r.status_code} post={comp} antes={base} despues={tras_comp}")

    # EL CASO D1. La que no se tocó se mueve DOS MESES adelante. Sin 0041 sus
    # celdas de mayo desaparecían, la actividad salía del denominador y el PPC
    # de la semana SUBÍA solo: «si no cumpliste, muévela».
    r = c.put(f"{API}/ev/programacion/actividades/{act_d1a.get('id')}",
              json={"fecha": "2026-07-06", "fecha_fin": "2026-07-07"})
    movida = r.status_code
    tras_mover = _ppc_semana()
    check("P73b D1 CERRADO: mover una comprometida sin avance no la saca del PPC",
          movida == 200
          and tras_mover.get("comprometidas") == tras_comp.get("comprometidas")
          and tras_mover.get("no_cumplidas") == tras_comp.get("no_cumplidas")
          and tras_mover.get("ppc") == tras_comp.get("ppc"),
          f"put={movida} antes={tras_comp} despues={tras_mover}")

    # LA MISMA FUGA POR LA OTRA PUERTA: en vez de mover la fecha, bajar el
    # compromiso a lo que sí se hizo (25 → 5, con 5 registrados). Contra el plan
    # vigente pasaría a cumplida; contra el congelado sigue debiendo 20.
    r = c.put(f"{API}/ev/programacion/actividades/{act_d1b.get('id')}",
              json={"metrado_prog": 5})
    bajada = r.status_code
    tras_bajar = _ppc_semana()
    check("P73c D1 CERRADO: bajar el metrado comprometido no vuelve cumplida la actividad",
          bajada == 200
          and tras_bajar.get("cumplidas") == tras_mover.get("cumplidas")
          and tras_bajar.get("no_cumplidas") == tras_mover.get("no_cumplidas")
          and tras_bajar.get("ppc") == tras_mover.get("ppc"),
          f"put={bajada} antes={tras_mover} despues={tras_bajar}")

    # EL CASO D2. Trabajo nuevo programado HACIA ATRÁS, dentro de una semana ya
    # comprometida: no puede entrar al denominador (no se prometió).
    p_d2 = _nueva_partida("E2E-D2", "Partida D2 tardia", 9, 36)
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": LUN_D1, "fecha_fin": "2026-05-05",
        "otm_id": "OTM-E2E", "titulo": "E2E D2 programada tarde",
        "partida_id": p_d2, "metrado_prog": 9})
    tras_d2 = _ppc_semana()
    check("P74 D2 CERRADO: programar hacia atrás no reescribe el PPC comprometido",
          r.status_code == 200
          and tras_d2.get("comprometidas") == tras_bajar.get("comprometidas")
          and tras_d2.get("ppc") == tras_bajar.get("ppc")
          and tras_d2.get("no_planificadas", 0) >= 1,
          f"status={r.status_code} antes={tras_bajar} despues={tras_d2}")

    # La otra puerta del mismo defecto: cancelar en vez de mover.
    r = c.put(f"{API}/ev/programacion/actividades/{act_ok.get('id')}",
              json={"estado": "CANCELADO"})
    tras_cancel = _ppc_semana()
    check("P75 cancelar una comprometida no la saca del PPC: cuenta como no cumplida",
          r.status_code == 200
          and tras_cancel.get("comprometidas") == tras_d2.get("comprometidas")
          and tras_cancel.get("no_cumplidas", 0) > tras_d2.get("no_cumplidas", 0),
          f"put={r.status_code} antes={tras_d2} despues={tras_cancel}")

    # La bitácora: el registro de cuándo se congeló y quién lo hizo.
    r = c.get(f"{API}/ev/programacion/semana-historial",
              params={"proyecto_id": 1, "lunes": LUN_D1})
    h = r.json() if r.status_code == 200 else {}
    evs = h.get("eventos", [])
    check("P76 la bitácora registra el compromiso con actor, metrado y fecha",
          r.status_code == 200 and len(evs) >= 1
          and evs[0]["evento"] == "COMPROMETIDA" and evs[0]["actor"] == "e2e"
          and evs[0]["metrado"] > 0 and len(evs[0].get("detalle") or []) >= 4
          and (h.get("resumen") or {}).get("veces_comprometida") == 1,
          f"status={r.status_code} eventos={evs} resumen={h.get('resumen')}")

    # Descomprometer NO borra la historia: quedan los dos eventos. Es el gesto
    # que más podría usarse para maquillar el indicador.
    c.delete(f"{API}/ev/programacion/plan-semana",
             params={"proyecto_id": 1, "lunes": LUN_D1, "actor": "e2e"})
    r = c.get(f"{API}/ev/programacion/semana-historial",
              params={"proyecto_id": 1, "lunes": LUN_D1})
    h2 = r.json() if r.status_code == 200 else {}
    tipos = [e["evento"] for e in h2.get("eventos", [])]
    sin_comp = _ppc_semana()
    check("P77 descomprometer deja rastro y devuelve la semana al plan vigente",
          r.status_code == 200 and tipos == ["COMPROMETIDA", "DESCOMPROMETIDA"]
          and (h2.get("resumen") or {}).get("descompromisos") == 1
          and sin_comp.get("origen") == "VIGENTE",
          f"tipos={tipos} semana={sin_comp}")

    # ── P78-P80 · Trabajo de terceros en el LookAhead (0042) ──────────────
    # Otra empresa da su plazo y de eso dependen actividades nuestras. La fila
    # tiene que arrastrar nuestras fechas con los vínculos de siempre pero NO
    # contar en nuestro PPC: su atraso no es nuestro incumplimiento.
    SEM_EXT = "2026-09-07"                    # lunes futuro, sin datos de otros checks
    ppc_ext_antes = _ppc_semana(SEM_EXT)

    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": SEM_EXT, "otm_id": "OTM-E2E",
        "titulo": "ELECTRO SAC — Montaje de bandejas",
        "externa": True, "empresa": "ELECTRO SAC", "plazo_dias": 4})
    ext = r.json() if r.status_code == 200 else {}
    # Con plazo, la F.Fin se deriva saltando los no laborables: no hay que
    # contar días a mano (es como dan el dato: «nos toma 4 días»).
    r2 = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": SEM_EXT, "fecha_fin": "2026-09-09",
        "otm_id": "OTM-E2E", "titulo": "E2E depende del tercero",
        "partida_id": p_d1a, "metrado_prog": 6})
    nuestra = r2.json() if r2.status_code == 200 else {}
    check("P78 fila de terceros: se crea sin partida ni metrado, con plazo y empresa",
          r.status_code == 200 and ext.get("externa") is True
          and ext.get("empresa") == "ELECTRO SAC"
          and ext.get("partida_id") is None and not ext.get("metrado_prog")
          and ext.get("fecha_fin") and ext.get("fecha_fin") > SEM_EXT,
          f"status={r.status_code} act={ext}")

    # Vincular: la nuestra va DESPUÉS de la del tercero (FS).
    rv = c.post(f"{API}/ev/programacion/actividades/{nuestra.get('id')}/dependencias",
                json={"predecesora_id": ext.get("id"), "tipo": "FS", "lag_dias": 0})
    # El tercero avisa que se atrasa 5 días: al mover SU fecha, la nuestra debe
    # correrse sola. Es la razón de modelarlo como actividad y no como
    # restricción (una restricción tiene fecha, no duración, y no arrastra).
    rm = c.put(f"{API}/ev/programacion/actividades/{ext.get('id')}",
               json={"fecha": "2026-09-14", "fecha_fin": "2026-09-17"})
    movidas = (rm.json() or {}).get("movidas") if rm.status_code == 200 else None
    # OJO: no nombrar `fila` a esta variable — choca con la función `fila()` del
    # propio script y la vuelve inaccesible en TODO main (UnboundLocalError).
    cur.execute("SELECT fecha FROM prog_actividades WHERE id=%s", (nuestra.get("id"),))
    row_ext = cur.fetchone()
    nueva_fecha = str(row_ext[0]) if row_ext else ""
    check("P79 el atraso del tercero arrastra nuestra actividad por el vínculo FS",
          rv.status_code == 200 and rm.status_code == 200
          and nuestra.get("id") in (movidas or [])
          and nueva_fecha > "2026-09-09",
          f"dep={rv.status_code} put={rm.status_code} movidas={movidas} nuestra={nueva_fecha}")

    ppc_ext = _ppc_semana("2026-09-14")
    r = c.get(f"{API}/ev/programacion/empresas", params={"proyecto_id": 1})
    emps = [e["empresa"] for e in r.json()] if r.status_code == 200 else []
    # La semana del tercero NO gana una comprometida por su culpa. Se compara
    # contra la semana donde quedó tras moverse: si entrara al PPC, aparecería.
    r2 = c.get(f"{API}/ev/programacion/cierre-semana",
               params={"proyecto_id": 1, "lunes": "2026-09-14"})
    cierre_ext = r2.json() if r2.status_code == 200 else {}
    ids_cierre = [x.get("actividad_id") for x in cierre_ext.get("actividades", [])]
    check("P80 la fila de terceros NO entra al PPC ni al cierre; la empresa se autocataloga",
          ext.get("id") not in ids_cierre
          and (ppc_ext.get("comprometidas") or 0) == (ppc_ext_antes.get("comprometidas") or 0)
          and "ELECTRO SAC" in emps,
          f"ids_cierre={ids_cierre} externa={ext.get('id')} ppc={ppc_ext} empresas={emps}")

    # El CHECK de la BD es la última línea: ninguna ruta futura puede crear la
    # combinación prohibida aunque se salte la validación del router.
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": SEM_EXT, "otm_id": "OTM-E2E",
        "titulo": "E2E externa invalida", "externa": True,
        "partida_id": p_d1a, "metrado_prog": 5})
    check("P81 una fila de terceros con partida o metrado se rechaza con 400 explicado",
          r.status_code == 400 and "partida" in str(r.json().get("detail", "")).lower(),
          f"status={r.status_code} detail={r.json().get('detail')}")

    # ── P82 · Actividad LIBRE sin metrado, dada por plazo ─────────────────
    # Es el camino del medio del selector «Programar actividad»: trabajo NUESTRO
    # que no está en el presupuesto (andamios, traslados, pruebas). Sin partida y
    # sin metrado el alta tiene que pasar igual —_exigir_partida solo prohíbe
    # metrado SIN partida—, la F.Fin se deriva del plazo, y a diferencia de la
    # externa SÍ cuenta en nuestro PPC: es un compromiso nuestro.
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-09-21", "otm_id": "OTM-E2E",
        "titulo": "E2E libre — montaje de andamios", "plazo_dias": 3})
    libre = r.json() if r.status_code == 200 else {}
    r2 = c.get(f"{API}/ev/programacion/cierre-semana",
               params={"proyecto_id": 1, "lunes": "2026-09-21"})
    ids_libre = [x.get("actividad_id") for x in (r2.json() or {}).get("actividades", [])]
    check("P82 actividad libre: sin partida ni metrado, F.Fin por plazo y SÍ entra al PPC",
          r.status_code == 200 and libre.get("externa") is False
          and libre.get("partida_id") is None and not libre.get("metrado_prog")
          and libre.get("fecha_fin") == "2026-09-23"
          and libre.get("id") in ids_libre,
          f"status={r.status_code} act={libre} ids_cierre={ids_libre}")

    # ── P83 · El catálogo de empresas ordena por PRIMERA APARICIÓN ────────
    # De este orden sale el color de la barra de cada empresa en el LookAhead.
    # Si ordenara por frecuencia, la barra cambiaría de color sola en cuanto una
    # empresa adelantara a otra en número de actividades: el color tiene que
    # seguir a la empresa, no a su ranking.
    for i, (emp, n) in enumerate([("E2E ZETA SAC", 1), ("E2E ALFA SAC", 3)]):
        for k in range(n):
            c.post(f"{API}/ev/programacion/actividades", json={
                "proyecto_id": 1, "fecha": "2026-09-28", "otm_id": "OTM-E2E",
                "titulo": f"{emp} — trabajo {k}", "externa": True,
                "empresa": emp, "plazo_dias": 2})
    r = c.get(f"{API}/ev/programacion/empresas", params={"proyecto_id": 1})
    lista = r.json() if r.status_code == 200 else []
    solo = [e for e in lista if e["empresa"].startswith("E2E ")]
    check("P83 el catálogo de empresas ordena por primera aparición, no por frecuencia",
          r.status_code == 200 and len(solo) == 2
          and solo[0]["empresa"] == "E2E ZETA SAC"   # apareció primero pese a tener 1
          and solo[1]["empresa"] == "E2E ALFA SAC"   # tiene 3, va después
          and [e["orden"] for e in lista] == list(range(len(lista))),
          f"status={r.status_code} lista={lista}")

    # ── P84 · Metrado comprometido que no se pudo repartir en ningún día ──
    # Si TODOS los días del rango son saltos ∅, `_redistribuir` se va sin
    # escribir una sola celda: el metrado queda comprometido y sin plan diario.
    # La actividad sigue contando en el PPC, así que el grid tiene que poder
    # distinguirla de una terminada — para eso está `prog_total`.
    r = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-10-05", "fecha_fin": "2026-10-07",
        "otm_id": "OTM-E2E", "titulo": "E2E sin dias habiles",
        "partida_id": p_d1a, "metrado_prog": 90,
        "dias_salto": ["2026-10-05", "2026-10-06", "2026-10-07"]})
    rota = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT COALESCE(SUM(cantidad),0) FROM prog_metrado_dia WHERE actividad_id=%s",
                (rota.get("id"),))
    celdas = float(cur.fetchone()[0])
    # Y una sana, para que el check distinga en vez de dar siempre lo mismo.
    r2 = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-10-12", "fecha_fin": "2026-10-14",
        "otm_id": "OTM-E2E", "titulo": "E2E con dias habiles",
        "partida_id": p_d1a, "metrado_prog": 90})
    sana = r2.json() if r2.status_code == 200 else {}
    g = c.get(f"{API}/ev/programacion/lookahead-grid",
              params={"proyecto_id": 1, "desde": "2026-10-05", "semanas": 3})
    filas = {a["id"]: a for gr in (g.json() or {}).get("grupos", [])
             for a in gr.get("actividades", [])}
    f_rota, f_sana = filas.get(rota.get("id"), {}), filas.get(sana.get("id"), {})
    check("P84 el grid distingue el metrado que no se pudo repartir (prog_total=0) del que sí",
          r.status_code == 200 and celdas == 0
          and f_rota.get("prog_total") == 0 and (f_rota.get("metrado_prog") or 0) > 0
          and (f_sana.get("prog_total") or 0) > 0,
          f"celdas={celdas} rota={f_rota.get('prog_total')} sana={f_sana.get('prog_total')}")

    # ── P85-P87 · Hoja semanal de HH y corrección desde oficina (0043) ────
    # Escenario: DOS supervisores tarean al mismo trabajador el mismo día. Como
    # el reemplazo del reenvío es por (supervisor, OTM, fecha), las dos sesiones
    # conviven y suman 14.5 HH contra una jornada de 10. Es el caso que la hoja
    # existe para cazar y arreglar.
    F_HOJA = "2026-06-03"   # miércoles de la semana 1 → jornada 10 por defecto
    cur.execute("INSERT INTO supervisores (id, nombre, activo) VALUES ('SUPE2F','SUP E2E DOS',true) "
                "ON CONFLICT (id) DO NOTHING")
    for sup, pid, hh in (("SUPE2E", p1["id"], 9.5), ("SUPE2F", p2["id"], 5.0)):
        c.post(f"{API}/api/sesion/enviar-con-partidas", json={
            "supervisor_id": sup, "otm_id": "OTM-E2E", "fecha": F_HOJA,
            "trabajadores": [{"trab_id": "901", "via": "e2e",
                              "asignaciones": [{"partida_id": pid, "hh": hh}]}]})

    r = c.get(f"{API}/ev/hoja-semanal", params={"lunes": "2026-06-01"})
    hoja = r.json() if r.status_code == 200 else {}
    proy = next((p for p in hoja.get("proyectos", []) if p["otm_id"] == "OTM-E2E"), {})
    tot = hoja.get("totales_persona_dia", {}).get(f"901|{F_HOJA}", {})
    check("P85 la hoja agrupa OTM->partida->persona y ve el exceso del día sumando sesiones",
          r.status_code == 200 and len(proy.get("partidas", [])) >= 2
          and abs(tot.get("hh", 0) - 14.5) < 0.01 and tot.get("estado") == "extra"
          and any(a["trab_id"] == "901" and a["fecha"] == F_HOJA for a in hoja.get("avisos", [])),
          f"partidas={len(proy.get('partidas', []))} tot={tot}")

    # Y filtrando un solo proyecto el total del día NO puede encogerse: el exceso
    # casi siempre nace de estar en dos sitios a la vez.
    r = c.get(f"{API}/ev/hoja-semanal", params={"lunes": "2026-06-01", "otm": "OTM-E2E"})
    tot_f = (r.json() or {}).get("totales_persona_dia", {}).get(f"901|{F_HOJA}", {})
    check("P85b el total por persona/día no cambia al filtrar por proyecto",
          abs(tot_f.get("hh", 0) - 14.5) < 0.01, f"tot_filtrado={tot_f}")

    cur.execute("SELECT id FROM tareo_partida WHERE otm_id='OTM-E2E' AND fecha=%s "
                "AND trabajador_id='901' AND supervisor_id='SUPE2E'", (F_HOJA,))
    fila_sup = cur.fetchone()
    linea_id = fila_sup[0] if fila_sup else 0
    r = c.patch(f"{API}/ev/tareo-linea/{linea_id}", json={"hh": 6.0, "motivo": "E2E doble tareo"})
    cur.execute("SELECT hh, editado_por FROM tareo_partida WHERE id=%s", (linea_id,))
    hh_post, editor = cur.fetchone() if linea_id else (None, None)
    cur.execute("SELECT accion, hh_antes, hh, usuario FROM tareo_ediciones WHERE tareo_id=%s", (linea_id,))
    traza = cur.fetchone()
    check("P86 corregir HH desde oficina baja el dato, marca la línea y deja traza",
          r.status_code == 200 and float(hh_post or 0) == 6.0 and editor
          and traza and traza[0] == "editar" and float(traza[1]) == 9.5 and float(traza[2]) == 6.0,
          f"status={r.status_code} hh={hh_post} editor={editor} traza={traza}")

    # El check que de verdad importa: la corrección de oficina gana sobre el
    # reenvío del supervisor. Sin la marca `editado_por`, este reenvío borraría
    # la línea y la repondría con 9.5, deshaciendo la corrección en silencio.
    r = c.post(f"{API}/api/sesion/enviar-con-partidas", json={
        "supervisor_id": "SUPE2E", "otm_id": "OTM-E2E", "fecha": F_HOJA,
        "trabajadores": [{"trab_id": "901", "via": "e2e",
                          "asignaciones": [{"partida_id": p1["id"], "hh": 9.5}]}]})
    respetados = (r.json() or {}).get("tareo_respetados", 0)
    cur.execute("SELECT COALESCE(SUM(hh),0) FROM tareo_partida WHERE otm_id='OTM-E2E' "
                "AND fecha=%s AND trabajador_id='901' AND supervisor_id='SUPE2E'", (F_HOJA,))
    hh_tras_reenvio = float(cur.fetchone()[0])
    # Control: el otro supervisor, cuya línea NO está corregida, sí se reemplaza.
    c.post(f"{API}/api/sesion/enviar-con-partidas", json={
        "supervisor_id": "SUPE2F", "otm_id": "OTM-E2E", "fecha": F_HOJA,
        "trabajadores": [{"trab_id": "901", "via": "e2e",
                          "asignaciones": [{"partida_id": p2["id"], "hh": 2.0}]}]})
    cur.execute("SELECT COALESCE(SUM(hh),0) FROM tareo_partida WHERE otm_id='OTM-E2E' "
                "AND fecha=%s AND trabajador_id='901' AND supervisor_id='SUPE2F'", (F_HOJA,))
    hh_sin_marca = float(cur.fetchone()[0])
    check("P87 el reenvío del supervisor respeta la línea corregida y sí reemplaza la que no lo está",
          hh_tras_reenvio == 6.0 and respetados == 1 and hh_sin_marca == 2.0,
          f"corregida={hh_tras_reenvio} respetados={respetados} sin_marca={hh_sin_marca}")

    # Anular deja la línea en 0 CON la marca (no la borra): si se borrara, el
    # siguiente reenvío la repondría tal cual.
    r = c.delete(f"{API}/ev/tareo-linea/{linea_id}")
    c.post(f"{API}/api/sesion/enviar-con-partidas", json={
        "supervisor_id": "SUPE2E", "otm_id": "OTM-E2E", "fecha": F_HOJA,
        "trabajadores": [{"trab_id": "901", "via": "e2e",
                          "asignaciones": [{"partida_id": p1["id"], "hh": 9.5}]}]})
    cur.execute("SELECT hh FROM tareo_partida WHERE id=%s", (linea_id,))
    tras_anular = cur.fetchone()
    check("P88 anular deja la línea en 0 y el reenvío tampoco la repone",
          r.status_code == 200 and tras_anular and float(tras_anular[0]) == 0.0,
          f"fila={tras_anular}")

    # ── P89-P92 · Libro mayor de fiabilidad de restricciones (0044) ───────
    # El circuito completo: catálogo de responsables → restricción ligada →
    # liberación con FECHA REAL → latencia medida → traza que sobrevive.
    r = c.post(f"{API}/ev/responsables", json={"nombre": " logistica ", "tipo": "INTERNA"})
    resp_log = r.json() if r.status_code == 200 else {}
    # Con tilde: es el duplicado que de verdad se cuela si solo se hace upper().
    r2 = c.post(f"{API}/ev/responsables", json={"nombre": "Logística"})
    check("P89 el catálogo normaliza (tildes incluidas) y reutiliza en vez de duplicar",
          r.status_code == 200 and resp_log.get("nombre") == "LOGISTICA"
          and r2.status_code == 200 and r2.json().get("id") == resp_log.get("id"),
          f"1={resp_log} 2={r2.json() if r2.status_code == 200 else r2.status_code}")

    act = c.post(f"{API}/ev/programacion/actividades", json={
        "proyecto_id": 1, "fecha": "2026-11-02", "fecha_fin": "2026-11-06",
        "otm_id": "OTM-E2E", "titulo": "E2E fiabilidad"}).json()
    rest = c.post(f"{API}/ev/programacion/actividades/{act['id']}/restricciones", json={
        "descripcion": "Perfiles de acero", "tipo": "MATERIALES",
        "responsable_id": resp_log.get("id"), "fecha_requerida": "2026-10-20"}).json()
    # Liberada 9 días TARDE, y se declara la fecha real (no la de captura).
    r = c.put(f"{API}/ev/programacion/restricciones/{rest.get('id')}",
              json={"liberada": True, "liberada_el": "2026-10-29"})
    lib = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT accion, latencia_dias FROM prog_restriccion_eventos "
                "WHERE restriccion_id=%s ORDER BY id", (rest.get("id"),))
    eventos = cur.fetchall()
    check("P90 liberar con fecha real deja la latencia congelada en la traza",
          r.status_code == 200 and str(lib.get("liberada_el")) == "2026-10-29"
          and [e[0] for e in eventos] == ["crear", "liberar"] and eventos[1][1] == 9,
          f"lib={lib.get('liberada_el')} eventos={eventos}")

    r = c.get(f"{API}/ev/fiabilidad/restricciones")
    fb = r.json() if r.status_code == 200 else {}
    mat = next((t for t in fb.get("por_tipo", []) if t["tipo"] == "MATERIALES"), {})
    logi = next((p for p in fb.get("por_responsable", []) if p["responsable"] == "LOGISTICA"), {})
    check("P91 el libro mayor mide la latencia por tipo y por responsable, con su n",
          r.status_code == 200 and mat.get("n_medidas", 0) >= 1
          and mat.get("mediana_dias") is not None and logi.get("n", 0) >= 1
          and mat.get("suficiente") is False,   # con 1 observación NO se presenta como referencia
          f"materiales={mat} logistica={logi}")

    # Reabrir borra la fecha real (la liberación se deshizo) pero el evento
    # anterior NO desaparece: sin eso, una restricción reabierta perdería la
    # historia de haberlo estado, que es justo lo que el libro mayor mide.
    c.put(f"{API}/ev/programacion/restricciones/{rest.get('id')}", json={"liberada": False})
    c.delete(f"{API}/ev/programacion/restricciones/{rest.get('id')}")
    cur.execute("SELECT accion FROM prog_restriccion_eventos WHERE restriccion_id=%s ORDER BY id",
                (rest.get("id"),))
    tras_borrar = [e[0] for e in cur.fetchall()]
    cur.execute("SELECT count(*) FROM prog_restricciones WHERE id=%s", (rest.get("id"),))
    quedan = cur.fetchone()[0]
    check("P92 la traza sobrevive a reabrir y al borrado de la restricción",
          tras_borrar == ["crear", "liberar", "reabrir", "eliminar"] and quedan == 0,
          f"eventos={tras_borrar} filas={quedan}")

    # ── P93-P96 · Bandeja de liberación en lote ──────────────────────────
    # La revisión del viernes: ver lo que sigue restringido y declarar, fila por
    # fila, qué día se liberó DE VERDAD. Las fechas requeridas van dos en el
    # pasado y una en el futuro para que la bandeja tenga que distinguirlas.
    rs = []
    for desc, req in (("Fierro corrugado", "2026-06-01"), ("Plano estructural", "2026-06-05"),
                      ("Grúa de 30t", "2026-12-01")):
        rs.append(c.post(f"{API}/ev/programacion/actividades/{act['id']}/restricciones", json={
            "descripcion": desc, "tipo": "MATERIALES",
            "responsable_id": resp_log.get("id"), "fecha_requerida": req}).json())
    r = c.get(f"{API}/ev/fiabilidad/pendientes")
    band = r.json() if r.status_code == 200 else {}
    orden = [p["id"] for p in band.get("pendientes", [])]
    pend = {p["id"]: p for p in band.get("pendientes", [])}
    check("P93 la bandeja lista lo abierto, lo más urgente primero y con el plazo calculado",
          r.status_code == 200 and all(x["id"] in pend for x in rs)
          and pend[rs[0]["id"]]["dias"] > 0 and pend[rs[2]["id"]]["dias"] < 0
          and orden.index(rs[0]["id"]) < orden.index(rs[2]["id"]),
          f"n={len(orden)} dias={[pend.get(x['id'], {}).get('dias') for x in rs]}")

    r = c.post(f"{API}/ev/fiabilidad/liberar", json={"items": [
        {"id": rs[0]["id"], "liberada_el": "2026-06-10"},     # 9 días tarde
        {"id": rs[1]["id"], "liberada_el": "2026-06-05"}]})   # el mismo día = a tiempo
    lote = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT restriccion_id, latencia_dias FROM prog_restriccion_eventos "
                "WHERE restriccion_id = ANY(%s) AND accion = 'liberar'",
                ([rs[0]["id"], rs[1]["id"]],))
    lat = dict(cur.fetchall())
    sigue = {p["id"] for p in c.get(f"{API}/ev/fiabilidad/pendientes").json().get("pendientes", [])}
    check("P94 el lote libera cada una con SU fecha real y las saca de la bandeja",
          r.status_code == 200 and len(lote.get("liberadas", [])) == 2
          and lat.get(rs[0]["id"]) == 9 and lat.get(rs[1]["id"]) == 0
          and rs[0]["id"] not in sigue and rs[1]["id"] not in sigue,
          f"lote={lote} latencias={lat}")

    # Control: un año tecleado de más no se nota en la lista pero deja la mediana
    # del responsable inservible. Tiene que rebotar Y no tocar la restricción.
    r = c.post(f"{API}/ev/fiabilidad/liberar",
               json={"items": [{"id": rs[2]["id"], "liberada_el": "2099-01-01"}]})
    cur.execute("SELECT liberada FROM prog_restricciones WHERE id = %s", (rs[2]["id"],))
    intacta = cur.fetchone()
    check("P95 una fecha de liberación futura se rechaza y no toca la restricción",
          r.status_code == 422 and intacta is not None and intacta[0] is False,
          f"status={r.status_code} liberada={intacta}")

    r = c.post(f"{API}/ev/fiabilidad/liberar",
               json={"items": [{"id": rs[0]["id"], "liberada_el": "2026-06-20"}]})
    out = r.json() if r.status_code == 200 else {}
    cur.execute("SELECT liberada_el FROM prog_restricciones WHERE id = %s", (rs[0]["id"],))
    ya_estaba = cur.fetchone()[0]
    check("P96 volver a liberar no pisa la fecha real ya declarada",
          r.status_code == 200 and out.get("omitidas") == [rs[0]["id"]]
          and out.get("liberadas") == [] and str(ya_estaba) == "2026-06-10",
          f"out={out} fecha={ya_estaba}")

    # ── P97-P99 · Fecha de detección, duración y proyecto (0045) ─────────
    # Una con la fecha DECLARADA y otra sin ella: la segunda tiene que caer al
    # sello de captura y decir que lo está haciendo, no fingir que la sabe.
    declarada = c.post(f"{API}/ev/programacion/actividades/{act['id']}/restricciones", json={
        "descripcion": "Valvulas importadas", "tipo": "MATERIALES",
        "responsable_id": resp_log.get("id"), "fecha_requerida": "2026-07-15",
        "detectada_el": "2026-05-20"}).json()
    sin_declarar = c.post(f"{API}/ev/programacion/actividades/{act['id']}/restricciones", json={
        "descripcion": "Certificado de calidad", "tipo": "INFORMACION",
        "responsable_id": resp_log.get("id"), "fecha_requerida": "2026-07-15"}).json()
    band = c.get(f"{API}/ev/fiabilidad/pendientes").json()
    pend = {p["id"]: p for p in band.get("pendientes", [])}
    pa, pb = pend.get(declarada.get("id"), {}), pend.get(sin_declarar.get("id"), {})
    hoy_api = band.get("hoy")
    check("P97 la detección declarada manda; sin ella cae al sello y se marca derivada",
          pa.get("detectada_el") == "2026-05-20" and pa.get("detectada_derivada") is False
          and pa.get("antiguedad", 0) > 0
          and pb.get("detectada_el") == hoy_api and pb.get("detectada_derivada") is True
          and pb.get("antiguedad") == 0 and pa.get("otm_id") == "OTM-E2E",
          f"declarada={pa} sin_declarar={pb}")

    # Vivió del 20-may al 30-jun = 41 días, y solo llegó 15 días tarde respecto
    # de lo pedido: el caso que justifica que duración y latencia sean columnas
    # distintas.
    r = c.post(f"{API}/ev/fiabilidad/liberar",
               json={"items": [{"id": declarada["id"], "liberada_el": "2026-06-30"}]})
    cur.execute("SELECT latencia_dias, duracion_dias FROM prog_restriccion_eventos "
                "WHERE restriccion_id=%s AND accion='liberar'", (declarada["id"],))
    ev = cur.fetchone()
    fb = c.get(f"{API}/ev/fiabilidad/restricciones").json()
    otms_fb = {o["otm_id"]: o for o in fb.get("por_otm", [])}
    fb_otm = c.get(f"{API}/ev/fiabilidad/restricciones?otm=OTM-E2E").json()
    # n_duracion == 1 es la parte que importa: las otras liberadas de este bloque
    # NO declararon detección, y con el sello de captura de respaldo darían una
    # duración negativa. Tienen que quedarse fuera, no entrar con signo cambiado.
    e2e_otm = otms_fb.get("OTM-E2E", {})
    check("P98 duración y latencia son distintas, y solo cuenta la detección declarada",
          r.status_code == 200 and ev == (-15, 41)
          and e2e_otm.get("duracion_mediana") == 41.0 and e2e_otm.get("n_duracion") == 1
          and e2e_otm.get("n_medidas", 0) > 1
          and 0 < fb_otm["total"]["n"] <= fb["total"]["n"],
          f"evento={ev} por_otm={e2e_otm} filtrado={fb_otm['total']['n']}")

    # Control de atomicidad: un lote con una fecha imposible NO puede dejar la
    # otra mitad aplicada.
    otra = c.post(f"{API}/ev/programacion/actividades/{act['id']}/restricciones", json={
        "descripcion": "Andamio E2E", "tipo": "EQUIPOS",
        "fecha_requerida": "2026-07-15", "detectada_el": "2026-07-01"}).json()
    r = c.post(f"{API}/ev/fiabilidad/liberar", json={"items": [
        {"id": sin_declarar["id"], "liberada_el": "2026-07-20"},
        {"id": otra["id"], "liberada_el": "2026-06-01"}]})   # antes de detectarse
    cur.execute("SELECT id, liberada FROM prog_restricciones WHERE id = ANY(%s)",
                ([sin_declarar["id"], otra["id"]],))
    estados = dict(cur.fetchall())
    check("P99 liberar antes de detectar rebota y deshace el lote ENTERO",
          r.status_code == 422 and estados.get(sin_declarar["id"]) is False
          and estados.get(otra["id"]) is False,
          f"status={r.status_code} estados={estados}")

    # ══════════════════════════════════════════════════════════════
    # F-CUADRILLA — el puente panel ↔ tareo (0046)
    #
    # Este bloque existe por un bug que ningún test unitario podía cazar: el
    # panel escribía en la tabla `cuadrillas` y la app de campo leía
    # `cuadrilla_otm`. Las dos mitades funcionaban perfectamente contra su
    # propia tabla y la cuadrilla creada en oficina no llegaba nunca al
    # teléfono. Lo único que lo detecta es cruzar los dos lados: escribir por
    # un endpoint y leer por el del otro.
    # ══════════════════════════════════════════════════════════════
    def cuadrillas_campo():
        """Lo que ve el TELÉFONO: {nombre: [miembros]}."""
        rr = c.get(f"{API}/ev/cuadrillas-plantilla", params={"supervisor_id": "SUPE2E"})
        return rr.json() if rr.status_code == 200 else {}

    def cuadrillas_panel():
        """Lo que ve OFICINA: [{id, nombre, miembros[]}]."""
        rr = c.get(f"{API}/api/cuadrillas/SUPE2E")
        return rr.json() if rr.status_code == 200 else []

    # C1 — oficina crea la cuadrilla; el orden que manda es el que se guarda.
    r = c.post(f"{API}/api/cuadrillas/SUPE2E",
               json={"nombre": "  E2E  Excavación ", "trab_ids": ["902", "901", "902"]})
    cua = r.json() if r.status_code == 200 else {}
    check("C1 el panel crea la cuadrilla (nombre normalizado, sin repetidos)",
          r.status_code == 200 and cua.get("nombre") == "E2E Excavación"
          and cua.get("total") == 2,
          f"status={r.status_code} body={r.text[:200]}")

    # C2 — EL CHECK QUE FALTABA: lo creado en el panel se ve en el teléfono.
    campo = cuadrillas_campo()
    miembros = campo.get("E2E Excavación", [])
    check("C2 esa cuadrilla llega al TAREO con su gente y en su orden",
          "E2E Excavación" in campo and len(miembros) == 2
          and [m["trabajador_id"] for m in miembros] == ["902", "901"]
          and miembros[0]["nombre"] and miembros[0]["cargo"],
          f"campo={campo}")

    # C3 — el nombre es único por supervisor: crear otra igual no machaca.
    r = c.post(f"{API}/api/cuadrillas/SUPE2E", json={"nombre": "E2E Excavación"})
    check("C3 nombre repetido rebota con 409 (no pisa la que está en uso)",
          r.status_code == 409 and len(cuadrillas_campo().get("E2E Excavación", [])) == 2,
          f"status={r.status_code}")

    # C4 — varias cuadrillas a la vez: es lo que pidió Jean.
    r = c.post(f"{API}/api/cuadrillas/SUPE2E",
               json={"nombre": "E2E Encofrado", "trab_ids": ["901"]})
    campo = cuadrillas_campo()
    check("C4 el supervisor tiene VARIAS cuadrillas a la vez",
          r.status_code == 200 and "E2E Excavación" in campo and "E2E Encofrado" in campo
          and len(campo["E2E Encofrado"]) == 1,
          f"status={r.status_code} campo={list(campo)}")

    # C5 — el circuito inverso: lo que el teléfono guarda se ve en oficina.
    # `otm_id` sigue viajando desde la app desplegada; ya no discrimina.
    r = c.post(f"{API}/ev/cuadrillas-plantilla", json={
        "supervisor_id": "SUPE2E", "otm_id": "OTM-E2E",
        "nombre": "E2E Del teléfono", "trabajadores": ["901", "902"]})
    panel = {g["nombre"]: g for g in cuadrillas_panel()}
    tel = panel.get("E2E Del teléfono", {})
    check("C5 la cuadrilla guardada en CAMPO se ve en el panel",
          r.status_code == 200 and tel.get("total") == 2
          and [m["trab_id"] for m in tel.get("miembros", [])] == ["901", "902"],
          f"status={r.status_code} panel={list(panel)}")

    # C6 — renombrar: antes un nombre mal escrito era permanente.
    gid = panel["E2E Excavación"]["id"]
    r = c.patch(f"{API}/api/cuadrilla-grupo/{gid}", json={"nombre": "E2E Excavación 2"})
    r409 = c.patch(f"{API}/api/cuadrilla-grupo/{gid}", json={"nombre": "E2E Encofrado"})
    campo = cuadrillas_campo()
    check("C6 renombrar llega al tareo y no permite chocar con otra",
          r.status_code == 200 and "E2E Excavación 2" in campo
          and "E2E Excavación" not in campo and r409.status_code == 409,
          f"status={r.status_code} dup={r409.status_code} campo={list(campo)}")

    # C7 — agregar y quitar gente de UNA cuadrilla sin tocar las otras.
    c.post(f"{API}/api/cuadrilla-grupo/{gid}/miembro/901")
    c.delete(f"{API}/api/cuadrilla-grupo/{gid}/miembro/902")
    campo = cuadrillas_campo()
    check("C7 alta y baja de un miembro afectan solo a su cuadrilla",
          [m["trabajador_id"] for m in campo.get("E2E Excavación 2", [])] == ["901"]
          and len(campo.get("E2E Encofrado", [])) == 1,
          f"campo={campo}")

    # C8 — eliminar: desaparece de los dos lados a la vez.
    r = c.delete(f"{API}/api/cuadrilla-grupo/{gid}")
    campo, panel = cuadrillas_campo(), {g["nombre"] for g in cuadrillas_panel()}
    check("C8 eliminar la cuadrilla la retira del panel Y del tareo",
          r.status_code == 200 and "E2E Excavación 2" not in campo
          and "E2E Excavación 2" not in panel and "E2E Encofrado" in campo,
          f"campo={list(campo)} panel={panel}")

    # C9 — nadie se lleva por delante la cuadrilla de otro supervisor.
    r = c.post(f"{API}/api/cuadrillas/SUPE2E", json={"nombre": ""})
    check("C9 cuadrilla sin nombre rebota con 422", r.status_code == 422,
          f"status={r.status_code}")

    # ══════════════════════════════════════════════════════════════
    # F-CATALOGO — las cuadrillas son listas LIBRES (0048)
    #
    # Corrección de modelo de Jean: la cuadrilla no es una unidad de mando que
    # rota de supervisor, es una lista de gente preestablecida para que el
    # registro diario sea de un toque. No es de nadie: la usa cualquiera.
    # ══════════════════════════════════════════════════════════════
    def catalogo():
        rr = c.get(f"{API}/api/cuadrillas")
        return {g["nombre"]: g for g in (rr.json() if rr.status_code == 200 else [])}

    def ve_el_telefono(sup):
        rr = c.get(f"{API}/ev/cuadrillas-plantilla", params={"supervisor_id": sup})
        return rr.json() if rr.status_code == 200 else {}

    # K1 — oficina crea una lista; nadie es su dueño.
    r = c.post(f"{API}/api/cuadrillas",
               json={"nombre": "E2E Compartida", "trab_ids": ["901", "902"]})
    cat = catalogo()
    check("K1 se crea una cuadrilla en el catálogo, sin dueño",
          r.status_code == 200 and cat.get("E2E Compartida", {}).get("total") == 2
          and cat["E2E Compartida"]["creada_por"] is None,
          f"status={r.status_code} body={r.text[:200]}")

    # K2 — EL PUNTO DEL CAMBIO: la lista que armó uno la ve el otro.
    # Antes «E2E Encofrado» era de SUPE2E y SUPE2F no la veía en su teléfono.
    de_2e, de_2f = ve_el_telefono("SUPE2E"), ve_el_telefono("SUPE2F")
    check("K2 todos los supervisores ven TODAS las cuadrillas",
          "E2E Encofrado" in de_2e and "E2E Encofrado" in de_2f
          and "E2E Compartida" in de_2e and "E2E Compartida" in de_2f,
          f"SUPE2E={list(de_2e)} SUPE2F={list(de_2f)}")

    # K3 — pero cada uno encuentra las SUYAS arriba: 30 listas sin orden es
    # peor que no tenerlas. Desde 0049 «las suyas» son sus HABITUALES, y armar
    # una la marca habitual de quien la armó (si no, caería al fondo mañana).
    r = c.post(f"{API}/api/cuadrillas/SUPE2F", json={"nombre": "E2E La de dos"})
    orden_2f = list(ve_el_telefono("SUPE2F"))
    orden_2e = list(ve_el_telefono("SUPE2E"))
    check("K3 cada supervisor ve primero sus habituales",
          r.status_code == 200 and orden_2f[0] == "E2E La de dos"
          and orden_2e[0] != "E2E La de dos",
          f"SUPE2F={orden_2f} SUPE2E={orden_2e}")

    # K4 — DUPLICAR: la variante de siempre sin volver a armarla.
    comp_id = catalogo()["E2E Compartida"]["id"]
    r = c.post(f"{API}/api/cuadrilla-grupo/{comp_id}/duplicar",
               json={"nombre": "E2E Compartida sin uno"})
    cat = catalogo()
    check("K4 duplicar copia la gente bajo otro nombre",
          r.status_code == 200 and cat.get("E2E Compartida sin uno", {}).get("total") == 2,
          f"status={r.status_code} body={r.text[:200]}")

    # K5 — la copia va por su cuenta.
    copia_id = cat["E2E Compartida sin uno"]["id"]
    c.delete(f"{API}/api/cuadrilla-grupo/{copia_id}/miembro/902")
    cat = catalogo()
    check("K5 editar la copia no toca al original",
          cat["E2E Compartida sin uno"]["total"] == 1
          and cat["E2E Compartida"]["total"] == 2,
          f"copia={cat['E2E Compartida sin uno']['total']} orig={cat['E2E Compartida']['total']}")

    # K6 — el nombre es único en TODA la empresa: se ven todas juntas y dos
    # «Encofrado» son indistinguibles justo al elegir.
    r1 = c.post(f"{API}/api/cuadrillas", json={"nombre": "E2E Compartida"})
    r2 = c.post(f"{API}/api/cuadrillas/SUPE2F", json={"nombre": "e2e compartida"})
    r3 = c.patch(f"{API}/api/cuadrilla-grupo/{copia_id}", json={"nombre": "E2E Compartida"})
    check("K6 nombre repetido: 409 aunque lo cree otro supervisor o cambie mayúsculas",
          r1.status_code == 409 and r2.status_code == 409 and r3.status_code == 409,
          f"oficina={r1.status_code} otro_sup={r2.status_code} renombrar={r3.status_code}")

    # K7 — el aviso de solapamiento sigue: 901 está en varias.
    m901 = [m for m in cat["E2E Compartida"]["miembros"] if m["trab_id"] == "901"]
    check("K7 avisa que esa persona está en más de una cuadrilla",
          len(m901) == 1 and m901[0]["en_otras"] >= 1, f"miembro={m901}")

    # K8 — PERSONAS SIN CUADRILLA (lo que pidió Jean): a estas no las encuentra
    # nadie en una lista guardada.
    r = c.get(f"{API}/api/trabajadores-sin-cuadrilla")
    sueltos = {t["id"] for t in (r.json() if r.status_code == 200 else [])}
    check("K8 lista a quien no está en ninguna cuadrilla, y no a quien sí",
          r.status_code == 200 and "901" not in sueltos and "902" not in sueltos
          and "900" in sueltos,          # el supervisor E2E no está en ninguna
          f"status={r.status_code} sueltos={sorted(sueltos)[:8]}")

    # K9 — al quitar a alguien de su última cuadrilla, aparece en esa lista.
    c.delete(f"{API}/api/cuadrilla-grupo/{copia_id}/miembro/901")
    for gid in [g["id"] for g in catalogo().values()]:
        c.delete(f"{API}/api/cuadrilla-grupo/{gid}/miembro/902")
    sueltos = {t["id"] for t in c.get(f"{API}/api/trabajadores-sin-cuadrilla").json()}
    check("K9 quien se queda fuera de todas aparece como sin cuadrilla",
          "902" in sueltos and "901" not in sueltos,   # 901 sigue en E2E Compartida
          f"sueltos={sorted(sueltos)[:8]}")

    # K10 — borrar y volver a crear con el mismo nombre REVIVE la lista con su
    # gente: el nombre sigue ocupado por la inactiva.
    r = c.delete(f"{API}/api/cuadrilla-grupo/{copia_id}")
    fuera = "E2E Compartida sin uno" not in catalogo()
    r2 = c.post(f"{API}/api/cuadrillas", json={"nombre": "E2E Compartida sin uno"})
    check("K10 borrar la esconde; recrearla con el mismo nombre la revive",
          r.status_code == 200 and fuera and r2.status_code == 200
          and "E2E Compartida sin uno" in catalogo(),
          f"borrar={r.status_code} fuera={fuera} recrear={r2.status_code}")

    # ══════════════════════════════════════════════════════════════
    # F-HABITUAL — cuáles usa NORMALMENTE cada supervisor (0049)
    #
    # Encargo de Jean: «poder asignarle a cada supervisor cuadrillas habituales,
    # que le aparezcan al inicio y como por separado, y luego las demás». Es un
    # atajo de PANTALLA: lo que hay que probar no es solo que suban, sino que
    # NADIE deje de ver ninguna por no tenerla marcada.
    # ══════════════════════════════════════════════════════════════
    def habituales_de(sup):
        rr = c.get(f"{API}/ev/cuadrillas-habituales", params={"supervisor_id": sup})
        return rr.json().get("habituales", []) if rr.status_code == 200 else []

    def asignacion():
        rr = c.get(f"{API}/api/cuadrillas-habituales")
        return {f["supervisor_id"]: f for f in (rr.json() if rr.status_code == 200 else [])}

    cat = catalogo()
    id_comp = cat["E2E Compartida"]["id"]

    # H1 — armar una cuadrilla la deja como habitual de quien la armó (K3).
    check("H1 quien arma una cuadrilla la recibe como habitual",
          "E2E La de dos" in habituales_de("SUPE2F"),
          f"SUPE2F={habituales_de('SUPE2F')}")

    # H2 — oficina se la asigna a otro: el caso de Jean, un supervisor nuevo que
    # no ha armado nada y necesita tener a mano las de siempre.
    r = c.put(f"{API}/api/supervisor/SUPE2E/cuadrillas-habituales",
              json={"grupos": [id_comp]})
    check("H2 oficina fija las habituales de un supervisor",
          r.status_code == 200 and habituales_de("SUPE2E") == ["E2E Compartida"],
          f"status={r.status_code} SUPE2E={habituales_de('SUPE2E')}")

    # H3 — y en su teléfono sale ARRIBA, que es todo el objetivo.
    orden_2e = list(ve_el_telefono("SUPE2E"))
    check("H3 la habitual encabeza la lista del teléfono",
          orden_2e[:1] == ["E2E Compartida"], f"SUPE2E={orden_2e}")

    # H4 — LO QUE NO DEBE PASAR: asignarla a uno no se la quita a los demás.
    # Si esto cae, «habitual» se convirtió en propiedad, que es justo el modelo
    # que Jean descartó.
    de_2f = ve_el_telefono("SUPE2F")
    check("H4 marcarla habitual de uno no se la esconde a nadie",
          "E2E Compartida" in de_2f and "E2E La de dos" in de_2f,
          f"SUPE2F={list(de_2f)}")

    # H5 — N:M de verdad: la misma cuadrilla es habitual de dos supervisores a
    # la vez (dos frentes, o el turno que se cubre entre dos).
    r = c.put(f"{API}/api/supervisor/SUPE2F/cuadrillas-habituales",
              json={"grupos": [id_comp, cat["E2E La de dos"]["id"]]})
    asig = asignacion()
    check("H5 la misma cuadrilla es habitual de dos supervisores",
          r.status_code == 200
          and id_comp in asig.get("SUPE2E", {}).get("grupos", [])
          and id_comp in asig.get("SUPE2F", {}).get("grupos", []),
          f"status={r.status_code} 2E={asig.get('SUPE2E')} 2F={asig.get('SUPE2F')}")

    # H6 — es un REEMPLAZO: la lista que llega es el estado final, no un alta.
    r = c.put(f"{API}/api/supervisor/SUPE2F/cuadrillas-habituales",
              json={"grupos": [id_comp]})
    check("H6 fijar habituales reemplaza, no acumula",
          r.status_code == 200 and habituales_de("SUPE2F") == ["E2E Compartida"],
          f"status={r.status_code} SUPE2F={habituales_de('SUPE2F')}")

    # H7 — vaciarlas es legítimo y devuelve el orden alfabético.
    r = c.put(f"{API}/api/supervisor/SUPE2F/cuadrillas-habituales", json={"grupos": []})
    orden_2f = list(ve_el_telefono("SUPE2F"))
    check("H7 sin habituales, el teléfono vuelve al orden alfabético",
          r.status_code == 200 and not habituales_de("SUPE2F")
          and orden_2f == sorted(orden_2f, key=str.lower),
          f"status={r.status_code} SUPE2F={orden_2f}")

    # H8 — una cuadrilla que no existe se ignora en silencio en vez de reventar:
    # el panel puede mandar un id que otro acaba de borrar.
    r = c.put(f"{API}/api/supervisor/SUPE2F/cuadrillas-habituales",
              json={"grupos": [id_comp, 999999]})
    check("H8 un id de cuadrilla inexistente se descarta, no rompe",
          r.status_code == 200 and r.json().get("grupos") == [id_comp],
          f"status={r.status_code} body={r.text[:160]}")

    # H9 — la asignación en bloque trae a TODOS los supervisores activos, tengan
    # o no habituales: la pantalla de oficina es la lista completa.
    asig = asignacion()
    check("H9 la asignación lista a todo supervisor activo",
          "SUPE2E" in asig and "SUPE2F" in asig
          and isinstance(asig["SUPE2E"]["grupos"], list),
          f"supervisores={sorted(asig)[:6]}")

    # H10 — el borrado de cuadrillas es LÓGICO (activo=false), así que la marca
    # de habitual sigue en la tabla. Ni el teléfono ni la pantalla de oficina
    # deben enseñarla: sería una habitual con un nombre que ya no existe.
    id_dos = catalogo()["E2E La de dos"]["id"]
    c.put(f"{API}/api/supervisor/SUPE2E/cuadrillas-habituales",
          json={"grupos": [id_comp, id_dos]})
    c.delete(f"{API}/api/cuadrilla-grupo/{id_dos}")
    asig = asignacion()
    check("H10 una cuadrilla borrada desaparece de las habituales",
          habituales_de("SUPE2E") == ["E2E Compartida"]
          and asig.get("SUPE2E", {}).get("grupos") == [id_comp],
          f"telefono={habituales_de('SUPE2E')} oficina={asig.get('SUPE2E')}")

    print()
    if _fallas:
        print(f"RESULTADO: {len(_fallas)} verificaciones FALLARON: {_fallas}")
        sys.exit(1)
    print("RESULTADO: TODAS las verificaciones pasaron -- flujo tareo->ISP OK")
    sys.exit(0)


if __name__ == "__main__":
    main()
