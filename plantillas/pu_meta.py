# ============================================================
# plantillas/pu_meta.py — plantilla del PRESUPUESTO META (PU)
#
# Es la única que no encaja en el modelo genérico de columnas: su estructura la
# fija el parser (`parsers/plantilla_pu.py`), que a su vez la heredó del Excel
# real del ex-gerente — dos hojas, cabeceras en filas fijas y un APU por bloques.
#
# Lo que se reparte hoy es literalmente el fixture de los tests
# (`tests/fixtures/pu_min.xls`): al usuario le llega un archivo que dice
# «OBRAS CIVILES TEST» y «Excavacion test». Esto lo sustituye por un ejemplo
# real, legible y con las instrucciones que nunca tuvo.
#
# Regla al tocar este archivo: se escriben VALORES, no fórmulas, en todo lo que
# el parser lee. openpyxl no calcula fórmulas, así que una fórmula sin valor
# cacheado llegaría al parser como celda vacía.
# ============================================================
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ._estilo import (AMBAR, AMBAR_BG, AMBAR_FUERTE, AZUL, AZUL_BG, DINERO,
                      DINERO_BG, FUENTE, GRIS, GRIS_BG, LINEA, PAPEL, TINTA,
                      TINTA2, TINTA3, WBS, WBS_BG)

# Filas donde el parser espera las cabeceras (0-based en el parser → +1 aquí).
FILA_CAB_PTOMETA = 9     # el parser lee datos desde la fila 10 (idx 9)
FILA_CAB_PUMETA = 8      # el parser lee datos desde la fila 9 (idx 8)
FILA_HH_DIA = 6          # 'HH - Dia' en idx 5, columna I


def _fina(color: str = LINEA) -> Side:
    return Side(style="thin", color=color)


def _titulo(ws: Worksheet, fila: int, col_ini: int, col_fin: int,
            titulo: str, sub: str) -> None:
    c = ws.cell(fila, col_ini, titulo)
    c.font = Font(name=FUENTE, size=16, bold=True, color=TINTA)
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    ws.row_dimensions[fila].height = 22
    s = ws.cell(fila + 1, col_ini, sub)
    s.font = Font(name=FUENTE, size=9, color=TINTA3)
    ws.merge_cells(start_row=fila + 1, start_column=col_ini,
                   end_row=fila + 1, end_column=col_fin)


def _guia(ws: Worksheet, fila: int, col_ini: int, col_fin: int, lineas: list) -> None:
    prim = fila
    for i, txt in enumerate(lineas):
        c = ws.cell(fila + i, col_ini, txt)
        c.font = Font(name=FUENTE, size=9.5,
                      bold=(i == 0), color=AZUL if i == 0 else TINTA2)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=fila + i, start_column=col_ini,
                       end_row=fila + i, end_column=col_fin)
        ws.row_dimensions[fila + i].height = 15
    relleno = PatternFill("solid", fgColor=AZUL_BG)
    for r in range(prim, prim + len(lineas)):
        for cc in range(col_ini, col_fin + 1):
            cel = ws.cell(r, cc)
            cel.fill = relleno
            cel.border = Border(
                top=_fina(AZUL) if r == prim else None,
                bottom=_fina(AZUL) if r == prim + len(lineas) - 1 else None,
                left=_fina(AZUL) if cc == col_ini else None,
                right=_fina(AZUL) if cc == col_fin else None)


def _cabecera(ws: Worksheet, fila: int, celdas: list, fondo: str, texto: str) -> None:
    """`celdas` = [(columna_1based, título, ancho, formato, ayuda)]."""
    for col, titulo, ancho, _fmt, ayuda in celdas:
        c = ws.cell(fila, col, titulo)
        c.font = Font(name=FUENTE, size=9, bold=True, color=texto)
        c.fill = PatternFill("solid", fgColor=fondo)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=texto),
                          left=_fina(PAPEL), right=_fina(PAPEL))
        if ayuda:
            com = Comment(f"{titulo}\n\n{ayuda}", "KAMPFER")
            com.width, com.height = 250, 110
            c.comment = com
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[fila].height = 28


def _fila_datos(ws: Worksheet, fila: int, valores: dict, formatos: dict,
                ejemplo: bool = True, negrita: bool = False) -> None:
    for col, val in valores.items():
        c = ws.cell(fila, col, val)
        c.font = Font(name=FUENTE, size=9.5, italic=ejemplo, bold=negrita,
                      color=TINTA3 if ejemplo else TINTA)
        if col in formatos:
            c.number_format = formatos[col]
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(vertical="center")
        c.border = Border(bottom=_fina(LINEA))
    ws.row_dimensions[fila].height = 16


def _hoja_ptometa(wb: Workbook, proyecto: str, hoy: date) -> None:
    """Jerarquía del presupuesto. El parser lee: D=item, A=fase, B=sub-fase,
    E=descripción, F=unidad, G=metrado, H=PU meta, I=parcial, L=PU oferta."""
    ws = wb.active
    ws.title = "PtoMeta"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBAR_FUERTE
    ws.column_dimensions["A"].width = 9

    _titulo(ws, 2, 1, 13, "Presupuesto META — jerarquía",
            f"KAMPFER · {proyecto or 'plantilla PU'} · generada el {hoy.strftime('%d/%m/%Y')}")
    # La cabecera va fija en la fila 9 (la busca el parser), así que la guía cabe
    # en cuatro líneas: la cuarta se gasta en dejar aire bajo el título.
    ws.row_dimensions[4].height = 8
    _guia(ws, 5, 1, 13, [
        "CÓMO LLENAR ESTA HOJA",
        "1.   Una fila por partida. La columna ITEM (D) manda la jerarquía: 01 es padre de 01.01.",
        "2.   Deja FASE vacía en las filas que solo agrupan; con FASE es una partida real que se mide,"
        " y esa necesita su bloque de APU en PU-Meta con el MISMO código.",
        "3.   No muevas ni borres la fila de cabeceras (la 9): el sistema la busca ahí exactamente.",
    ])

    cab = [
        (1,  "Fase",           9,  "", "Disciplina de la partida. VACÍA = la fila solo agrupa."),
        (2,  "Sub-fase",       10, "", "Segundo nivel de agrupación del presupuesto."),
        (3,  "Sub Item",       10, "", "Uso interno del presupuesto. Puede ir vacía."),
        (4,  "Item",           14, "", "Código jerárquico. Es la clave que enlaza con la hoja PU-Meta."),
        (5,  "Descripción",    42, "", "Nombre de la partida."),
        (6,  "Und.",           8,  "", "Unidad del metrado: m3, m2, kg, und…"),
        (7,  "Metrado Meta",   14, "#,##0.00", "Cantidad de la meta."),
        (8,  "Precio $ Meta",  14, "#,##0.0000",
             "Costo unitario directo. Tiene que coincidir con el CUD del APU en PU-Meta."),
        (9,  "Parcial $ Meta", 15, "#,##0.00", "Metrado × Precio meta."),
        (10, "Total",          14, "#,##0.00", "Solo en las filas padre: total del grupo."),
        (11, "",               3,  "", ""),
        (12, "Precio $ Oferta", 14, "#,##0.0000", "Precio unitario de venta, si lo manejas."),
        (13, "Parcial $ Oferta", 15, "#,##0.00", "Metrado × Precio oferta."),
    ]
    _cabecera(ws, FILA_CAB_PTOMETA, cab, AMBAR_FUERTE, PAPEL)

    fmt = {7: "#,##0.00", 8: "#,##0.0000", 9: "#,##0.00", 10: "#,##0.00",
           12: "#,##0.0000", 13: "#,##0.00"}
    # Padre (sin fase) + dos partidas hoja. Coherente con el APU de PU-Meta.
    _fila_datos(ws, 10, {4: "01", 5: "OBRAS CIVILES", 10: 4300.00}, fmt, negrita=True)
    _fila_datos(ws, 11, {1: "11", 2: "11.01", 3: "11.01", 4: "01.01",
                         5: "Excavación masiva con retroexcavadora", 6: "m3",
                         7: 500.0, 8: 6.50, 9: 3250.00, 12: 8.50, 13: 4250.00}, fmt)
    _fila_datos(ws, 12, {1: "11", 2: "11.02", 3: "11.02", 4: "01.02",
                         5: "Relleno compactado con material propio", 6: "m3",
                         7: 300.0, 8: 3.50, 9: 1050.00, 12: 4.60, 13: 1380.00}, fmt)

    for r in range(13, 200):
        ws.row_dimensions[r].height = 15
        for c in range(1, 14):
            ws.cell(r, c).border = Border(bottom=_fina("F2F5F9"))
            if c in fmt:
                ws.cell(r, c).number_format = fmt[c]

    ws.freeze_panes = "A10"
    for k in range(10, 13):
        m = ws.cell(k, 14, "◀ ejemplo: bórralo")
        m.font = Font(name=FUENTE, size=8.5, italic=True, color=AMBAR_FUERTE)
    ws.column_dimensions["N"].width = 18


def _hoja_pumeta(wb: Workbook, proyecto: str, hoy: date) -> None:
    """APU por bloques. El parser lee: H=marca de fila, I/J según la marca,
    D=tipo de recurso, L=und, M=cuadrilla, N=cantidad, O=precio, P=parcial."""
    ws = wb.create_sheet("PU-Meta")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DINERO
    # El parser exige 'HH - Dia' en la fila 6, columnas H/I. Así que el título va
    # solo en la fila 2 y la guía ocupa las filas 3-7 pero SIN pasar de la
    # columna G: si el merge llegara hasta la H, la celda de HH-Día quedaría
    # dentro de él y no se podría escribir.
    t = ws.cell(2, 1, "Presupuesto META — análisis de precios unitarios")
    t.font = Font(name=FUENTE, size=16, bold=True, color=TINTA)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    ws.row_dimensions[2].height = 22
    _guia(ws, 3, 1, 7, [
        "CÓMO LLENAR ESTA HOJA",
        "1.   Un BLOQUE por cada partida con fase de PtoMeta, en el mismo orden.",
        "2.   El bloque empieza con «Partida» en la columna H y el código en la I.",
        "3.   Dentro van las secciones Mano de Obra · Materiales · Equipos.",
        "4.   La suma de los PARCIALES debe dar el Costo Unitario Directo.",
    ])

    # 'HH - Dia' donde el parser lo busca: fila 6 (idx 5), columnas H e I.
    c = ws.cell(FILA_HH_DIA, 8, "HH - Dia")
    c.font = Font(name=FUENTE, size=9.5, bold=True, color=TINTA2)
    c.alignment = Alignment(horizontal="right", vertical="center")
    h = ws.cell(FILA_HH_DIA, 9, 10.0)
    h.font = Font(name=FUENTE, size=10, bold=True, color=AMBAR)
    h.fill = PatternFill("solid", fgColor=AMBAR_BG)
    h.border = Border(bottom=Side(style="medium", color=AMBAR))
    h.alignment = Alignment(horizontal="center", vertical="center")
    com = Comment("Horas de jornada al día.\nCon esto se convierten los rendimientos "
                  "en HH por unidad.", "KAMPFER")
    com.width, com.height = 250, 90
    h.comment = com

    cab = [
        (1,  "P or SP",   9,  "", "P = partida del presupuesto · SP = subpartida que se reutiliza."),
        (2,  "Area",      14, "", "Zona de la obra, si la manejas."),
        (3,  "Descripción Area", 18, "", "Opcional."),
        (4,  "Insumo",    10, "", "MO, MAT o EQ en las filas de recurso. Vacía en el resto."),
        (5,  "Fase",      8,  "", "La misma de PtoMeta."),
        (6,  "Sub-fase",  10, "", "La misma de PtoMeta."),
        (7,  "Sub Item",  10, "", "Uso interno."),
        (8,  "Código",    16, "", "Marca la fila: «Partida», «Rendimiento», o el código del recurso."),
        (9,  "Descripción", 40, "", "Código de la partida, nombre del recurso o nombre de la sección."),
        (10, "Rend. MO",  11, "#,##0.00", "Rendimiento de mano de obra por día."),
        (11, "",          3,  "", ""),
        (12, "Und.",      9,  "", "Unidad del recurso."),
        (13, "Cuadrilla", 10, "#,##0.00", "Número de personas o equipos."),
        (14, "Cantidad",  11, "#,##0.0000", "Cantidad por unidad de partida."),
        (15, "Precio",    12, "#,##0.0000", "Precio unitario del recurso."),
        (16, "Parcial",   12, "#,##0.0000", "Cantidad × Precio."),
        (17, "Metrado",   12, "#,##0.00", "Solo en la fila «Partida»: metrado total."),
        (18, "HH totales", 12, "#,##0.00", "Solo en mano de obra: HH de toda la partida."),
        (19, "Costo total", 13, "#,##0.00", "Solo en mano de obra: costo de toda la partida."),
    ]
    _cabecera(ws, FILA_CAB_PUMETA, cab, DINERO, PAPEL)

    fmt = {10: "#,##0.00", 13: "#,##0.00", 14: "#,##0.0000", 15: "#,##0.0000",
           16: "#,##0.0000", 17: "#,##0.00", 18: "#,##0.00", 19: "#,##0.00"}
    f = FILA_CAB_PUMETA + 1

    def bloque(codigo: str, desc: str, metrado: float, area: str, fase: str,
               sub_fase: str, rend: float, cud: float, recursos: list) -> None:
        nonlocal f
        _fila_datos(ws, f, {1: "P", 2: area, 5: fase, 6: sub_fase, 7: sub_fase,
                            8: "Partida", 9: codigo, 10: desc, 17: metrado},
                    fmt, negrita=True)
        for cc in range(1, 20):
            ws.cell(f, cc).fill = PatternFill("solid", fgColor=WBS_BG)
            ws.cell(f, cc).font = Font(name=FUENTE, size=9.5, bold=True, color=WBS)
        # Estos dos bloques también son ejemplo y también hay que borrarlos: sin
        # rótulo es fácil dejarlos y acabar importando «Excavación masiva» como
        # partida real de la obra.
        marca = ws.cell(f, 20, "◀ bloque de EJEMPLO: bórralo entero")
        marca.font = Font(name=FUENTE, size=8.5, italic=True, color=AMBAR_FUERTE)
        ws.column_dimensions["T"].width = 30
        f += 1
        _fila_datos(ws, f, {2: area, 5: fase, 6: sub_fase, 7: sub_fase,
                            8: "Rendimiento", 9: "MO.", 10: rend,
                            15: "Costo Unitario Directo", 16: cud}, fmt)
        f += 1
        seccion_actual = None
        for tipo, cod, nombre, und, cuad, cant, precio, parcial, hh_t, costo_t in recursos:
            nombre_seccion = {"MO": "Mano de Obra", "MAT": "Materiales",
                              "EQ": "Equipos"}[tipo]
            if nombre_seccion != seccion_actual:
                _fila_datos(ws, f, {2: area, 5: fase, 6: sub_fase, 7: sub_fase,
                                    9: nombre_seccion}, fmt)
                for cc in range(1, 20):
                    ws.cell(f, cc).fill = PatternFill("solid", fgColor=GRIS_BG)
                ws.cell(f, 9).font = Font(name=FUENTE, size=9, bold=True, color=GRIS)
                seccion_actual = nombre_seccion
                f += 1
            vals = {2: area, 4: tipo, 5: fase, 6: sub_fase, 7: sub_fase, 8: cod,
                    9: nombre, 12: und, 13: cuad, 14: cant, 15: precio, 16: parcial}
            if hh_t is not None:
                vals[18] = hh_t
                vals[19] = costo_t
            _fila_datos(ws, f, vals, fmt)
            f += 1
        # Fila de cierre: el parser la reconoce por el CUD en la columna P.
        _fila_datos(ws, f, {16: cud}, fmt, negrita=True)
        for cc in range(1, 20):
            ws.cell(f, cc).border = Border(top=_fina(DINERO))
        f += 2

    # ── Ejemplo 1: excavación (500 m3, CUD 6.50) ──
    bloque("01.01", "Excavación masiva con retroexcavadora", 500.0, "Area 1", "11", "11.01",
           40.0, 6.50, [
               ("MO", "470101", "Operario",           "hh",  1.0, 0.2000, 12.5000, 2.5000, 100.0, 1250.0),
               ("MO", "470102", "Peón",               "hh",  1.0, 0.2000,  8.0000, 1.6000, 100.0,  800.0),
               ("EQ", "490301", "Retroexcavadora 320", "hm", 1.0, 0.0400, 60.0000, 2.4000, None, None),
           ])
    # ── Ejemplo 2: relleno (300 m3, CUD 3.50) ──
    bloque("01.02", "Relleno compactado con material propio", 300.0, "Area 1", "11", "11.02",
           60.0, 3.50, [
               ("MO", "470102", "Peón",               "hh",  2.0, 0.2667,  8.0000, 2.1336, 80.0, 640.0),
               ("EQ", "490401", "Compactadora tipo plancha", "hm", 1.0, 0.0500, 27.3280, 1.3664, None, None),
           ])

    for r in range(f, f + 60):
        ws.row_dimensions[r].height = 15
        for cc in range(1, 20):
            if cc in fmt:
                ws.cell(r, cc).number_format = fmt[cc]

    ws.freeze_panes = ws.cell(FILA_CAB_PUMETA + 1, 1)


def _hoja_como_funciona(wb: Workbook) -> None:
    """La estructura del archivo explicada. Sin esto, un formato de dos hojas
    con bloques y filas-marca no lo adivina nadie."""
    ws = wb.create_sheet("CÓMO FUNCIONA")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AZUL
    ws.column_dimensions["A"].width = 2.6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 88

    _titulo(ws, 2, 2, 3, "Cómo está armada esta plantilla",
            "Dos hojas que se enlazan por el código de la partida.")

    filas = [
        ("LAS DOS HOJAS", ""),
        ("PtoMeta", "La lista de partidas con su metrado y su precio unitario meta. "
                    "Es el resumen del presupuesto."),
        ("PU-Meta", "El análisis de precios unitarios: de qué está hecho el precio de "
                    "cada partida (mano de obra, materiales, equipos)."),
        ("El enlace", "El código de la columna ITEM de PtoMeta tiene que aparecer igual "
                      "en la fila «Partida» de PU-Meta. Si no coincide, esa partida entra sin APU."),
        ("", ""),
        ("REGLAS QUE NO SE PUEDEN SALTAR", ""),
        ("Filas de cabecera", "En PtoMeta la cabecera va en la fila 9; en PU-Meta, en la 8. "
                              "El sistema las busca ahí. No insertes filas por encima."),
        ("Partidas hoja", "Toda fila de PtoMeta CON fase necesita su bloque de APU en PU-Meta."),
        ("Filas padre", "Las que solo agrupan van SIN fase y no necesitan APU."),
        ("El CUD manda", "La suma de los parciales de un bloque debe dar su Costo Unitario "
                         "Directo, y ese CUD debe coincidir con el «Precio $ Meta» de PtoMeta. "
                         "Si no cuadra, el import lo dice antes de guardar nada."),
        ("HH - Día", "En PU-Meta, celda I6. Es la jornada con la que se convierten los "
                     "rendimientos en horas-hombre."),
        ("", ""),
        ("SI YA TIENES TU PRESUPUESTO", ""),
        ("Archivos antiguos", "Si tu presupuesto está en el formato .xls de siempre, súbelo "
                              "tal cual: el sistema también lo lee. Esta plantilla es para "
                              "empezar de cero."),
        ("Antes de subir", "Revisa que los totales cuadren en tu Excel. El import verifica "
                           "el CUD y las HH, y te muestra los problemas antes de confirmar."),
    ]
    fila = 5
    for etiqueta, texto in filas:
        if etiqueta and not texto:
            c = ws.cell(fila, 2, etiqueta)
            c.font = Font(name=FUENTE, size=9, bold=True, color=PAPEL)
            c.fill = PatternFill("solid", fgColor=TINTA2)
            d = ws.cell(fila, 3)
            d.fill = PatternFill("solid", fgColor=TINTA2)
            ws.row_dimensions[fila].height = 20
        elif etiqueta:
            a = ws.cell(fila, 2, etiqueta)
            a.font = Font(name=FUENTE, size=9.5, bold=True, color=TINTA)
            a.alignment = Alignment(vertical="top", indent=1)
            a.border = Border(bottom=_fina(LINEA))
            b = ws.cell(fila, 3, texto)
            b.font = Font(name=FUENTE, size=9.5, color=TINTA2)
            b.alignment = Alignment(vertical="top", wrap_text=True)
            b.border = Border(bottom=_fina(LINEA))
            ws.row_dimensions[fila].height = 30
        else:
            ws.row_dimensions[fila].height = 8
        fila += 1


def construir_pu(proyecto: str = "", hoy: Optional[date] = None) -> Workbook:
    hoy = hoy or date.today()
    wb = Workbook()
    _hoja_ptometa(wb, proyecto, hoy)
    _hoja_pumeta(wb, proyecto, hoy)
    _hoja_como_funciona(wb)
    return wb
