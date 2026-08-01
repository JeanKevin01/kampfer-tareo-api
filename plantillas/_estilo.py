# ============================================================
# plantillas/_estilo.py — el estándar visual de las plantillas KAMPFER
#
# Encargo de Jean (2026-08-01): «que se vea lo más limpio y profesional,
# mejor que un estudio de ingeniería», y que se entienda sola cómo llenarla.
#
# Por qué vive en el API y no en el panel: la librería del panel (SheetJS
# community) DESCARTA los estilos de celda en silencio — se le puede pedir una
# cabecera ámbar y escribe una hoja plana. openpyxl sí escribe color, bordes,
# desplegables, formato condicional y paneles congelados; y aquí la plantilla
# puede nacer con los catálogos reales del proyecto dentro.
#
# La regla que hace que se entienda sin leer nada:
#     ÁMBAR = tienes que llenarlo · GRIS = si quieres · AZUL = no lo toques
# Son los mismos colores que el panel, con el mismo significado, para que se
# aprendan una sola vez.
# ============================================================
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ── Paleta ───────────────────────────────────────────────────
# Variante CLARA de los tokens de `index.css`: Excel siempre se ve sobre
# blanco. El relleno es un tinte del mismo tono que el texto — nunca dos
# familias de color en la misma celda.
TINTA      = "151B27"   # casi negro, para títulos
TINTA2     = "586274"   # texto secundario
TINTA3     = "878FA2"   # texto terciario / ejemplos
LINEA      = "E0E5EE"   # bordes finos
PAPEL      = "FFFFFF"

AMBAR      = "A86D08"; AMBAR_BG   = "FDF0D5"; AMBAR_FUERTE = "B8770C"
AZUL       = "2563EB"; AZUL_BG    = "EDF3FE"
GRIS       = "6B7488"; GRIS_BG    = "F4F6FA"
WBS        = "6D28D9"; WBS_BG     = "F1EBFE"
DINERO     = "0D9488"; DINERO_BG  = "E3F6F4"
VERDE      = "059669"; VERDE_BG   = "E6F6F0"
ROJO       = "DC2626"; ROJO_BG    = "FDEAEA"

FUENTE = "Calibri"

# Cada nivel de una columna: color del texto de la cabecera y su relleno.
# `obligatorio` va en sólido invertido porque es lo único que el ojo debe
# encontrar de un vistazo al abrir el archivo.
NIVELES = {
    "obligatorio": {"txt": PAPEL,  "bg": AMBAR_FUERTE, "borde": AMBAR},
    "opcional":    {"txt": GRIS,   "bg": GRIS_BG,      "borde": LINEA},
    "calculado":   {"txt": AZUL,   "bg": AZUL_BG,      "borde": AZUL},
}
# Matices para cabeceras de GRUPO (la banda que agrupa columnas).
TONOS = {
    "wbs":    {"txt": WBS,    "bg": WBS_BG},
    "dinero": {"txt": DINERO, "bg": DINERO_BG},
    "medida": {"txt": AMBAR,  "bg": AMBAR_BG},
    "info":   {"txt": AZUL,   "bg": AZUL_BG},
    "neutro": {"txt": GRIS,   "bg": GRIS_BG},
}

FILA_MARGEN = 1          # columna A / fila 1 = aire, no datos
COL_MARGEN = 1


@dataclass
class Col:
    """Una columna de la plantilla.

    `clave` es la cabecera EXACTA que espera el importador — no se toca sin
    romper los archivos que ya circulan. `titulo` es el nombre legible que va
    en el comentario emergente, que es donde puede ser todo lo claro que
    quiera sin cambiar el contrato.
    """
    clave: str
    titulo: str = ""
    nivel: str = "opcional"           # obligatorio | opcional | calculado
    ancho: int = 14
    formato: str = ""                 # number_format de Excel
    ayuda: str = ""                   # comentario emergente
    si_falla: str = ""                # «si te equivocas…» (hoja INSTRUCCIONES)
    lista: Optional[list] = None      # valores cerrados → desplegable
    catalogo: str = ""                # nombre del catálogo (lista larga)
    ejemplo: Any = ""
    # 2ª fila de ejemplo. `None` = repite la primera; `""` = VACÍA a propósito,
    # que es como se enseña «este campo va vacío en las filas que solo agrupan».
    ejemplo2: Any = None
    formula: str = ""                 # columna calculada; {f} = número de fila
    grupo: str = ""
    tono: str = "neutro"              # matiz de la cabecera de grupo
    alin: str = ""                    # left | center | right (por defecto según formato)


@dataclass
class Plantilla:
    clave: str
    archivo: str
    titulo: str
    proposito: str                    # una línea: para qué sirve
    hoja: str
    cols: list[Col]
    pasos: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)   # cosas que evitan un mal rato
    catalogos: dict[str, list[tuple]] = field(default_factory=dict)
    filas_libres: int = 300
    ejemplos: int = 1


# ── Piezas de dibujo ─────────────────────────────────────────
def _fina(color: str = LINEA) -> Side:
    return Side(style="thin", color=color)


def _sin_bordes() -> Border:
    return Border()


def _preparar(ws: Worksheet, ncols: int) -> None:
    """Lo que convierte una hoja de cálculo en un documento: sin cuadrícula,
    con margen a la izquierda y aire arriba."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBAR_FUERTE
    ws.column_dimensions["A"].width = 2.6
    ws.row_dimensions[1].height = 10
    ws.sheet_view.zoomScale = 100


def _banda_titulo(ws: Worksheet, titulo: str, subtitulo: str,
                  col_ini: int, col_fin: int, fila: int) -> int:
    """Título y subtítulo. Sin recuadros ni rellenos: el peso tipográfico
    basta, y es lo que distingue un documento de un formulario."""
    c = ws.cell(fila, col_ini, titulo)
    c.font = Font(name=FUENTE, size=17, bold=True, color=TINTA)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[fila].height = 24
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)

    s = ws.cell(fila + 1, col_ini, subtitulo)
    s.font = Font(name=FUENTE, size=9, color=TINTA3)
    s.alignment = Alignment(vertical="top")
    ws.row_dimensions[fila + 1].height = 14
    ws.merge_cells(start_row=fila + 1, start_column=col_ini,
                   end_row=fila + 1, end_column=col_fin)

    # Regla fina bajo el título: separa el encabezado del contenido sin ruido.
    ws.row_dimensions[fila + 2].height = 6
    for cc in range(col_ini, col_fin + 1):
        ws.cell(fila + 2, cc).border = Border(bottom=_fina(LINEA))
    return fila + 3


def _bloque_guia(ws: Worksheet, pasos: list[str], avisos: list[str],
                 col_ini: int, col_fin: int, fila: int) -> int:
    """La guía va ENCIMA de la tabla, no en una pestaña aparte: una hoja
    LEYENDA se lee poco; lo que está sobre los datos, se lee siempre."""
    fila += 1
    alto_paso = 15

    tit = ws.cell(fila, col_ini, "CÓMO LLENARLA")
    tit.font = Font(name=FUENTE, size=8.5, bold=True, color=AZUL)
    tit.alignment = Alignment(vertical="center")
    ws.row_dimensions[fila].height = 16
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    prim = fila
    fila += 1

    for p in pasos:
        c = ws.cell(fila, col_ini, p)
        c.font = Font(name=FUENTE, size=9.5, color=TINTA2)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[fila].height = alto_paso
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        fila += 1

    for a in avisos:
        c = ws.cell(fila, col_ini, a)
        c.font = Font(name=FUENTE, size=9.5, bold=True, color=AMBAR)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[fila].height = alto_paso
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        fila += 1

    # Fondo y contorno del bloque completo, aplicados celda a celda porque un
    # merge en openpyxl solo estila su celda superior izquierda.
    relleno = PatternFill("solid", fgColor=AZUL_BG)
    for r in range(prim, fila):
        for cc in range(col_ini, col_fin + 1):
            cel = ws.cell(r, cc)
            cel.fill = relleno
            cel.border = Border(
                top=_fina(AZUL) if r == prim else None,
                bottom=_fina(AZUL) if r == fila - 1 else None,
                left=_fina(AZUL) if cc == col_ini else None,
                right=_fina(AZUL) if cc == col_fin else None,
            )
    ws.row_dimensions[fila].height = 10
    return fila + 1


def _cabeceras(ws: Worksheet, cols: list[Col], col_ini: int, fila: int) -> int:
    """Cabecera de grupo (si la hay) + cabecera de columna."""
    hay_grupos = any(c.grupo for c in cols)
    if hay_grupos:
        ini = 0
        while ini < len(cols):
            fin = ini
            while fin + 1 < len(cols) and cols[fin + 1].grupo == cols[ini].grupo:
                fin += 1
            tono = TONOS.get(cols[ini].tono, TONOS["neutro"])
            c1, c2 = col_ini + ini, col_ini + fin
            cel = ws.cell(fila, c1, (cols[ini].grupo or "").upper())
            cel.font = Font(name=FUENTE, size=8, bold=True, color=tono["txt"])
            cel.alignment = Alignment(horizontal="center", vertical="center")
            for cc in range(c1, c2 + 1):
                x = ws.cell(fila, cc)
                x.fill = PatternFill("solid", fgColor=tono["bg"])
                x.border = Border(left=_fina(PAPEL) if cc == c1 else None,
                                  right=_fina(PAPEL) if cc == c2 else None)
            if c2 > c1:
                ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
            ini = fin + 1
        ws.row_dimensions[fila].height = 15
        fila += 1

    for i, col in enumerate(cols):
        est = NIVELES[col.nivel]
        cel = ws.cell(fila, col_ini + i, col.clave)
        cel.font = Font(name=FUENTE, size=9, bold=True, color=est["txt"])
        cel.fill = PatternFill("solid", fgColor=est["bg"])
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = Border(bottom=Side(style="medium", color=est["borde"]),
                            left=_fina(PAPEL), right=_fina(PAPEL))
        ayuda = col.titulo or col.clave
        if col.ayuda:
            ayuda += "\n\n" + col.ayuda
        if col.nivel == "obligatorio":
            ayuda += "\n\n● Campo obligatorio"
        elif col.nivel == "calculado":
            ayuda += "\n\n● No escribas aquí: lo calcula la plantilla"
        com = Comment(ayuda, "KAMPFER")
        com.width, com.height = 260, 120
        cel.comment = com
        ws.column_dimensions[get_column_letter(col_ini + i)].width = col.ancho
    ws.row_dimensions[fila].height = 30
    return fila + 1


def _alineacion(col: Col) -> Alignment:
    if col.alin:
        return Alignment(horizontal=col.alin, vertical="center")
    if col.formato and any(s in col.formato for s in ("#", "0")):
        return Alignment(horizontal="right", vertical="center")
    if col.lista or col.catalogo:
        return Alignment(horizontal="center", vertical="center")
    return Alignment(horizontal="left", vertical="center")


def _cuerpo(ws: Worksheet, p: Plantilla, col_ini: int, fila_datos: int) -> None:
    """Ejemplos + filas vacías listas: formato, desplegables, fórmulas y el
    aviso de lo que falta."""
    ncols = len(p.cols)
    col_fin = col_ini + ncols - 1
    ult = fila_datos + p.filas_libres - 1
    letra_ini, letra_fin = get_column_letter(col_ini), get_column_letter(col_fin)
    # Las fórmulas nombran COLUMNAS, no letras: `{HITO1_PESO}{f}` en vez de
    # `N{f}`. Escribir la letra a mano ya costó un error real —la columna de
    # control sumaba las descripciones en vez de los pesos— y volvería a fallar
    # en silencio al reordenar una columna.
    letras = {c.clave: get_column_letter(col_ini + i) for i, c in enumerate(p.cols)}

    for r in range(fila_datos, ult + 1):
        es_ej = r < fila_datos + p.ejemplos
        ws.row_dimensions[r].height = 17
        for i, col in enumerate(p.cols):
            cel = ws.cell(r, col_ini + i)
            if es_ej:
                if r == fila_datos:
                    val = col.ejemplo
                else:
                    val = col.ejemplo if col.ejemplo2 is None else col.ejemplo2
                if val != "" and not col.formula:
                    cel.value = val
                cel.font = Font(name=FUENTE, size=9.5, italic=True, color=TINTA3)
                cel.fill = PatternFill("solid", fgColor="FAFBFC")
            else:
                cel.font = Font(name=FUENTE, size=10, color=TINTA)
            if col.formula:
                cel.value = col.formula.format(f=r, **letras)
                cel.font = Font(name=FUENTE, size=10, bold=True, color=AZUL,
                                italic=es_ej)
            if col.formato:
                cel.number_format = col.formato
            cel.alignment = _alineacion(col)
            # Las filas aún vacías llevan una línea MÁS tenue: guían el ojo sin
            # que 250 filas preparadas parezcan una tabla rayada.
            cel.border = Border(bottom=_fina(LINEA if es_ej else "F1F4F9"))

    # Marca de la fila de ejemplo, en el margen: no ocupa una columna de datos
    # y aun así es imposible no verla.
    for k in range(p.ejemplos):
        m = ws.cell(fila_datos + k, COL_MARGEN, "▸")
        m.font = Font(name=FUENTE, size=11, bold=True, color=AMBAR_FUERTE)
        m.alignment = Alignment(horizontal="center", vertical="center")
        com = Comment("Fila de EJEMPLO.\nBórrala antes de subir el archivo.", "KAMPFER")
        com.width, com.height = 220, 70
        m.comment = com

    # Desplegables. Las listas largas apuntan a la hoja CATÁLOGOS (una lista
    # en línea no puede pasar de 255 caracteres).
    for i, col in enumerate(p.cols):
        letra = get_column_letter(col_ini + i)
        rango = f"{letra}{fila_datos}:{letra}{ult}"
        if col.lista:
            valores = ",".join(str(v) for v in col.lista)
            if len(valores) > 250:
                continue
            dv = DataValidation(type="list", formula1=f'"{valores}"',
                                allow_blank=True, showErrorMessage=True)
            dv.errorTitle = "Valor no válido"
            dv.error = f"Elige uno de: {', '.join(str(v) for v in col.lista)}"
            dv.promptTitle = col.clave
            dv.prompt = col.ayuda or col.titulo
            ws.add_data_validation(dv)
            dv.add(rango)
        elif col.catalogo and col.catalogo in p.catalogos:
            n = len(p.catalogos[col.catalogo])
            if not n:
                continue
            ini_cat = _fila_catalogo(p, col.catalogo)
            ref = f"CATALOGOS!$A${ini_cat}:$A${ini_cat + n - 1}"
            dv = DataValidation(type="list", formula1=ref, allow_blank=True,
                                showErrorMessage=True)
            dv.errorTitle = "Valor no válido"
            dv.error = ("Ese valor no está en el catálogo del proyecto. "
                        "Míralos en la hoja CATÁLOGOS.")
            ws.add_data_validation(dv)
            dv.add(rango)

    # Lo que falta, en rojo — pero SOLO en las filas ya empezadas. Pintar de
    # rojo todo lo vacío convertiría el archivo en un incendio y nadie miraría.
    obligatorias = [i for i, c in enumerate(p.cols) if c.nivel == "obligatorio"]
    if obligatorias:
        rojo = PatternFill("solid", fgColor=ROJO_BG)
        for i in obligatorias:
            letra = get_column_letter(col_ini + i)
            rango = f"{letra}{fila_datos}:{letra}{ult}"
            formula = (f"AND(LEN({letra}{fila_datos})=0,"
                       f"COUNTA(${letra_ini}{fila_datos}:${letra_fin}{fila_datos})>0)")
            ws.conditional_formatting.add(
                rango, FormulaRule(formula=[formula], fill=rojo, stopIfTrue=False))

    ws.auto_filter.ref = f"{letra_ini}{fila_datos - 1}:{letra_fin}{ult}"
    ws.freeze_panes = ws.cell(fila_datos, col_ini + 1)


def _fila_catalogo(p: Plantilla, nombre: str) -> int:
    """Fila donde arrancan los valores de ese catálogo en la hoja CATÁLOGOS.
    Debe coincidir exactamente con lo que dibuja `_hoja_catalogos`."""
    fila = 3
    for nom, filas in p.catalogos.items():
        fila += 1                      # título del bloque
        if nom == nombre:
            return fila
        fila += len(filas) + 2         # valores + aire
    return fila


def _hoja_catalogos(wb: Workbook, p: Plantilla) -> None:
    """Los valores válidos del PROYECTO, no de un manual: alimentan los
    desplegables y siempre están al día porque nacen en el servidor."""
    ws = wb.create_sheet("CATALOGOS")
    _preparar(ws, 3)
    ws.sheet_properties.tabColor = GRIS
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 62

    t = ws.cell(2, 1, "VALORES VÁLIDOS DE ESTE PROYECTO")
    t.font = Font(name=FUENTE, size=13, bold=True, color=TINTA)
    s = ws.cell(3, 1, "Los desplegables de la plantilla salen de aquí. No cambies esta hoja.")
    s.font = Font(name=FUENTE, size=9, color=TINTA3)

    fila = 4
    for nombre, filas in p.catalogos.items():
        c = ws.cell(fila, 1, nombre.upper())
        c.font = Font(name=FUENTE, size=9, bold=True, color=PAPEL)
        c.fill = PatternFill("solid", fgColor=AMBAR_FUERTE)
        c.alignment = Alignment(vertical="center", indent=1)
        d = ws.cell(fila, 2)
        d.fill = PatternFill("solid", fgColor=AMBAR_FUERTE)
        ws.row_dimensions[fila].height = 20
        fila += 1
        for valor, desc in filas:
            a = ws.cell(fila, 1, valor)
            a.font = Font(name=FUENTE, size=10, bold=True, color=TINTA)
            a.alignment = Alignment(vertical="center", indent=1)
            a.border = Border(bottom=_fina(LINEA))
            b = ws.cell(fila, 2, desc)
            b.font = Font(name=FUENTE, size=9.5, color=TINTA2)
            b.alignment = Alignment(vertical="center", wrap_text=True)
            b.border = Border(bottom=_fina(LINEA))
            fila += 1
        fila += 1


def _hoja_instrucciones(wb: Workbook, p: Plantilla) -> None:
    """Campo por campo, en tabla. La columna que hoy no existe en ningún sitio
    —y la que evita la llamada de teléfono— es «SI TE EQUIVOCAS»."""
    ws = wb.create_sheet("INSTRUCCIONES")
    _preparar(ws, 5)
    ws.sheet_properties.tabColor = AZUL
    anchos = [26, 15, 30, 24, 46]
    for i, a in enumerate(anchos):
        ws.column_dimensions[get_column_letter(2 + i)].width = a

    fila = _banda_titulo(ws, p.titulo, p.proposito, 2, 6, 2)

    # Leyenda de color con muestras reales, que es como se entiende de verdad.
    fila += 1
    ws.cell(fila, 2, "QUÉ SIGNIFICA CADA COLOR").font = Font(
        name=FUENTE, size=8.5, bold=True, color=TINTA3)
    fila += 1
    leyenda = [
        ("obligatorio", "OBLIGATORIO", "Sin esto la fila no se importa."),
        ("opcional",    "OPCIONAL",    "Puedes dejarlo vacío."),
        ("calculado",   "NO ESCRIBIR", "Lo calcula la plantilla sola."),
    ]
    for nivel, etiqueta, texto in leyenda:
        est = NIVELES[nivel]
        c = ws.cell(fila, 2, etiqueta)
        c.font = Font(name=FUENTE, size=9, bold=True, color=est["txt"])
        c.fill = PatternFill("solid", fgColor=est["bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="medium", color=est["borde"]))
        d = ws.cell(fila, 3, texto)
        d.font = Font(name=FUENTE, size=9.5, color=TINTA2)
        d.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=6)
        ws.row_dimensions[fila].height = 19
        fila += 1

    fila += 1
    cab = ["CAMPO", "¿OBLIGATORIO?", "QUÉ PONER", "EJEMPLO", "SI TE EQUIVOCAS"]
    for i, txt in enumerate(cab):
        c = ws.cell(fila, 2 + i, txt)
        c.font = Font(name=FUENTE, size=9, bold=True, color=PAPEL)
        c.fill = PatternFill("solid", fgColor=TINTA2)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 22
    fila += 1

    for col in p.cols:
        obligatorio = {"obligatorio": "Sí", "opcional": "No",
                       "calculado": "No — automático"}[col.nivel]
        if col.lista:
            que = " · ".join(str(v) for v in col.lista)
        elif col.catalogo:
            que = f"Un valor de «{col.catalogo}» (hoja CATÁLOGOS)"
        else:
            que = col.ayuda or col.titulo or "Texto libre"
        valores = [col.clave, obligatorio, que, str(col.ejemplo or "—"),
                   col.si_falla or "—"]
        for i, v in enumerate(valores):
            c = ws.cell(fila, 2 + i, v)
            negrita = i == 0
            color = TINTA if i == 0 else TINTA2
            if i == 1 and col.nivel == "obligatorio":
                color = AMBAR
                negrita = True
            c.font = Font(name=FUENTE, size=9.5, bold=negrita, color=color)
            c.alignment = Alignment(vertical="center", wrap_text=(i >= 2), indent=1)
            c.border = Border(bottom=_fina(LINEA))
        ws.row_dimensions[fila].height = 28
        fila += 1


def construir(p: Plantilla, proyecto: str = "", hoy: Optional[date] = None) -> Workbook:
    """Arma el libro completo: datos → instrucciones → catálogos."""
    wb = Workbook()
    ws = wb.active
    ws.title = p.hoja
    ncols = len(p.cols)
    col_ini = COL_MARGEN + 1
    col_fin = col_ini + ncols - 1
    _preparar(ws, ncols)

    fecha = (hoy or date.today()).strftime("%d/%m/%Y")
    sub = "KAMPFER" + (f" · {proyecto}" if proyecto else "") + f" · generada el {fecha}"
    fila = _banda_titulo(ws, p.titulo, sub, col_ini, col_fin, 2)

    # El paso 1 nombra la fila del ejemplo, así que hay que saberla ANTES de
    # escribir la guía. El bloque tiene una altura conocida: título + pasos +
    # avisos + aire, y después la cabecera (dos filas si hay grupos).
    hay_grupos = any(c.grupo for c in p.cols)
    n_pasos = len(p.pasos) + 3           # +2 fijos al inicio, +1 al final
    fila_datos = fila + 3 + n_pasos + len(p.avisos) + (2 if hay_grupos else 1)

    pasos = [
        f"Borra la fila de ejemplo (la {fila_datos}, en cursiva) y escribe los tuyos desde ahí.",
        "Las columnas NARANJA son obligatorias; las grises, opcionales; "
        "las azules las calcula la plantilla sola.",
        *p.pasos,
        "Guarda el archivo y súbelo en el panel: verás una vista previa antes de confirmar.",
    ]
    pasos = [f"{i + 1}.   {t}" for i, t in enumerate(pasos)]

    fila = _bloque_guia(ws, pasos, p.avisos, col_ini, col_fin, fila)
    assert fila + (1 if hay_grupos else 0) + 1 == fila_datos, (
        "la fila del ejemplo anunciada en el paso 1 no coincide con la real")
    fila = _cabeceras(ws, p.cols, col_ini, fila)
    _cuerpo(ws, p, col_ini, fila)

    _hoja_instrucciones(wb, p)
    if p.catalogos:
        _hoja_catalogos(wb, p)

    ws.sheet_view.selection[0].activeCell = f"{get_column_letter(col_ini)}{fila}"
    ws.sheet_view.selection[0].sqref = f"{get_column_letter(col_ini)}{fila}"
    return wb
